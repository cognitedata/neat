from collections.abc import Mapping
from pathlib import Path
from typing import ClassVar, cast

import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from cognite.neat._data_model.importers._base import DMSImporter
from cognite.neat._data_model.models.dms import (
    RequestSchema,
)
from cognite.neat._exceptions import DataModelImportException
from cognite.neat._issues import ModelSyntaxError
from cognite.neat._utils.text import humanize_collection
from cognite.neat._utils.useful_types import CellValueType, DataModelTableType
from cognite.neat._utils.validation import as_json_path, humanize_validation_error

from .data_classes import MetadataValue, TableDMS
from .reader import DMSTableReader
from .source import SpreadsheetReadContext, TableSource


class DMSTableImporter(DMSImporter):
    """Imports DMS from a table structure.

    The tables can are expected to be a dictionary where the keys are the table names and the values
    are lists of dictionaries representing the rows in the table.
    """

    # We can safely cast as we know the validation_alias is always set to a str.
    REQUIRED_SHEETS = tuple(
        cast(str, field_.validation_alias) for field_ in TableDMS.model_fields.values() if field_.is_required()
    )
    REQUIRED_SHEET_MESSAGES: ClassVar[Mapping[str, str]] = {
        f"Missing required column: {sheet!r}": f"Missing required sheet: {sheet!r}" for sheet in REQUIRED_SHEETS
    }
    MetadataSheet = cast(str, TableDMS.model_fields["metadata"].validation_alias)

    def __init__(self, tables: DataModelTableType, source: TableSource | None = None) -> None:
        self._table = tables
        self._source = source or TableSource("Unknown")

    def to_data_model(self) -> RequestSchema:
        tables = self._read_tables()

        space, version = self._read_defaults(tables.metadata)
        reader = DMSTableReader(space, version, self._source)
        return reader.read_tables(tables)

    @classmethod
    def from_yaml(cls, yaml_file: Path) -> "DMSTableImporter":
        """Create a DMSTableImporter from a YAML file."""
        source = cls._display_name(yaml_file)
        return cls(yaml.safe_load(yaml_file.read_text()), TableSource(source.as_posix()))

    @classmethod
    def from_json(cls, json_file: Path) -> "DMSTableImporter":
        """Create a DMSTableImporter from a JSON file."""
        return cls.from_yaml(json_file)

    @classmethod
    def from_excel(cls, excel_file: Path) -> "DMSTableImporter":
        """Create a DMSTableImporter from an Excel file."""
        tables: DataModelTableType = {}
        source = TableSource(cls._display_name(excel_file).as_posix())
        workbook = load_workbook(excel_file, read_only=True, data_only=True, rich_text=False)
        try:
            for column_id, column_info in TableDMS.model_fields.items():
                sheet_name = cast(str, column_info.validation_alias)
                if sheet_name not in workbook.sheetnames:
                    continue
                required_columns = TableDMS.get_sheet_columns(column_id, column_info, column_type="required")
                sheet = workbook[sheet_name]
                context = SpreadsheetReadContext()
                table_rows = cls._read_rows(sheet, required_columns, context)
                tables[sheet_name] = table_rows
                source.table_read[sheet_name] = context
            return cls(tables, source)
        finally:
            workbook.close()

    def _read_tables(self) -> TableDMS:
        try:
            # Check tables, columns, data type and entity syntax.
            table = TableDMS.model_validate(self._table)
        except ValidationError as e:
            errors = self._create_error_messages(e)
            raise DataModelImportException(errors) from None
        return table

    def _create_error_messages(self, error: ValidationError) -> list[ModelSyntaxError]:
        errors: list[ModelSyntaxError] = []
        seen: set[str] = set()
        for message in humanize_validation_error(
            error,
            humanize_location=self._location,
            field_name="column",
            missing_required_descriptor="missing",
        ):
            # Replace messages about missing required columns with missing required sheets.
            message = self.REQUIRED_SHEET_MESSAGES.get(message, message)
            if message in seen:
                # We treat all rows as the same, so we get duplicated errors for each row.
                continue
            seen.add(message)
            errors.append(ModelSyntaxError(message=message))
        return errors

    def _location(self, loc: tuple[str | int, ...]) -> str:
        if isinstance(loc[0], str) and len(loc) == 2:  # Sheet + row.
            # We skip the row as we treat all rows as the same. For example, if a required column is missing in one
            # row, it is missing in all rows.
            return f"{loc[0]} sheet"
        elif len(loc) == 3 and isinstance(loc[0], str) and isinstance(loc[1], int) and isinstance(loc[2], str):
            # This means there is something wrong in a specific cell.

            sheet = loc[0]
            row = loc[1]
            if self._source and sheet in self._source.table_read:
                context = self._source.table_read[sheet]
                row = context.adjusted_row_number(row) - 1

            return f"{sheet} sheet row {row + 1} column {loc[2]!r}"
        # This should be unreachable as the TableDMS model only has 2 levels.
        return as_json_path(loc)

    @staticmethod
    def _read_defaults(metadata: list[MetadataValue]) -> tuple[str, str]:
        """Reads the space and version from the metadata table."""
        default_space: str | None = None
        default_version: str | None = None
        missing = {"space", "version"}
        for meta in metadata:
            if meta.key == "space":
                default_space = str(meta.value)
                missing.remove("space")
            elif meta.key == "version":
                default_version = str(meta.value)
                missing.remove("version")
        if missing:
            error = ModelSyntaxError(message=f"In Metadata missing required values: {humanize_collection(missing)}")
            # If space or version is missing, we cannot continue parsing the model as these are used as defaults.
            raise DataModelImportException([error]) from None
        return str(default_space), str(default_version)

    @classmethod
    def _display_name(cls, filepath: Path) -> Path:
        """Get a display-friendly version of the file path."""
        cwd = Path.cwd()
        source = filepath
        if filepath.is_relative_to(cwd):
            source = filepath.relative_to(cwd)
        return source

    @classmethod
    def _read_rows(
        cls, sheet: Worksheet, required_columns: list[str], context: SpreadsheetReadContext
    ) -> list[dict[str, CellValueType]]:
        table_rows: list[dict[str, CellValueType]] = []
        # Metadata sheet is just a key-value pair of the first two columns.
        # For other sheets, we need to find the column header row first.
        columns: list[str] = [] if sheet.title != cls.MetadataSheet else required_columns
        for row_no, row in enumerate(sheet.iter_rows(values_only=True)):
            if columns:
                # We have found the column header row, read the data rows.
                if all(cell is None for cell in row):
                    context.empty_rows.append(row_no)
                else:
                    record = dict(zip(columns, row, strict=False))
                    # MyPy complains as it thinks DataTableFormula | ArrayFormula could be cell values,
                    # but as we used values_only=True, this is not the case.
                    table_rows.append(record)  # type: ignore[arg-type]
            else:
                # Look for the column header row.
                row_values = [str(cell) for cell in row]
                if set(row_values).intersection(required_columns):
                    columns = row_values
                    context.header_row = row_no

        return table_rows
