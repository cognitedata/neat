import itertools
from collections.abc import Mapping, Set
from dataclasses import dataclass
from functools import lru_cache
from typing import Literal, cast

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from cognite.neat._data_model._constants import (
    CDF_CDM_SPACE,
    CDF_CDM_VERSION,
    COGNITE_CONCEPTS_3D,
    COGNITE_CONCEPTS_ANNOTATIONS,
    COGNITE_CONCEPTS_CONFIGURATIONS,
    COGNITE_CONCEPTS_INTERFACES,
    COGNITE_CONCEPTS_MAIN,
)
from cognite.neat._data_model.importers._table_importer.data_classes import DMSContainer, DMSProperty, DMSView, TableDMS
from cognite.neat._data_model.models.dms import (
    DMS_DATA_TYPES,
    EnumProperty,
    FileCDFExternalIdReference,
    SequenceCDFExternalIdReference,
    TimeseriesCDFExternalIdReference,
)
from cognite.neat._utils.useful_types import CellValueType, DataModelTableType

MAIN_HEADERS_BY_SHEET_NAME: Mapping[str, str] = {
    "Properties": "Definition of Properties",
    "Views": "Definition of Views",
    "Containers": "Definition of Containers",
    "Nodes": "Definition of Nodes",
    "Enum": "Definition of Enum Collections",
}
MAX_COLUMN_WIDTH = 70.0
HEADER_ROWS = 2


@dataclass
class WorkbookOptions:
    """Options for creating an Excel workbook from a data model.

    Attributes:
        adjust_column_width (bool): Whether to adjust the column widths to fit the content. Default
            is True.
        style_headers (bool): Whether to style the header rows. Default is True.
        row_band_highlighting (bool): Whether to apply row band highlighting to the properties sheet.
            Default is True.
        separate_view_properties (bool): Whether to separate properties by view with an empty row.
            Default is False.
        add_dropdowns (bool): Whether to add drop-down menus for certain columns. Default is True.
        dropdown_implements (Set[Literal["main", "interface", "configuration", "annotation", "3D"]]):
            The types of Cognite concepts to include in the "implements" drop-down menu. Default is
            {"main", "interface"}.
        max_views (int): The maximum number of views to support in the drop-down menus. Default is 100.
        max_containers (int): The maximum number of containers to support in the drop-down menus. Default is 100.
        max_properties_per_view (int): The maximum number of properties per view to support in the
            drop-down menus. Default is 100.
        skip_properties_in_other_spaces (bool): Whether to skip properties that are in other spaces
            in the properties sheet. Default is True.
    """

    adjust_column_width: bool = True
    style_headers: bool = True
    row_band_highlighting: bool = True
    separate_view_properties: bool = False
    add_dropdowns: bool = True
    dropdown_implements: Set[Literal["main", "interface", "configuration", "annotation", "3D"]] = frozenset(
        {"main", "interface"}
    )
    max_views: int = 100
    max_containers: int = 100
    max_properties_per_view: int = 100
    skip_properties_in_other_spaces: bool = True


class WorkbookCreator:
    # skip types which require special handling (enum) or are surpassed by CDM (CDF references)
    DROPDOWN_DMS_TYPE_EXCLUDE = frozenset(
        {
            # MyPy does not understand that model_fields is in all pydantic classes.
            prop.model_fields["type"].default  # type: ignore[attr-defined]
            for prop in [
                EnumProperty,
                TimeseriesCDFExternalIdReference,
                SequenceCDFExternalIdReference,
                FileCDFExternalIdReference,
            ]
        }
    )

    # These classer are used to refer to sheets that needs to be explicitly checked for in the
    # workbook creation.
    class Sheets:
        metadata = cast(str, TableDMS.model_fields["metadata"].validation_alias)
        properties = cast(str, TableDMS.model_fields["properties"].validation_alias)
        views = cast(str, TableDMS.model_fields["views"].validation_alias)
        containers = cast(str, TableDMS.model_fields["containers"].validation_alias)
        dropdown_source = "_dropdown_source"

    # The following classes are used to refer to sheets and columns that are used
    # in dropdown creation.
    class PropertyColumns:
        view = cast(str, DMSProperty.model_fields["view"].validation_alias)
        value_type = cast(str, DMSProperty.model_fields["value_type"].validation_alias)
        immutable = cast(str, DMSProperty.model_fields["immutable"].validation_alias)
        container = cast(str, DMSProperty.model_fields["container"].validation_alias)

    class ContainerColumns:
        container = cast(str, DMSContainer.model_fields["container"].validation_alias)
        used_for = cast(str, DMSContainer.model_fields["used_for"].validation_alias)

    class ViewColumns:
        view = cast(str, DMSView.model_fields["view"].validation_alias)
        implements = cast(str, DMSView.model_fields["implements"].validation_alias)

    class DropdownSourceColumns:
        view = 1
        implements = 2
        value_type = 3
        container = 4
        in_model = 5
        immutable = 5
        used_for = 6

    def __init__(self, options: WorkbookOptions | None = None) -> None:
        options = options or WorkbookOptions()
        self._adjust_column_width = options.adjust_column_width
        self._style_headers = options.style_headers
        self._row_band_highlighting = options.row_band_highlighting
        self._separate_view_properties = options.separate_view_properties
        self._add_dropdowns = options.add_dropdowns
        self._dropdown_implements = options.dropdown_implements
        self._max_views = options.max_views
        self._max_containers = options.max_containers
        self._max_properties_per_view = options.max_properties_per_view

    def create_workbook(self, tables: DataModelTableType) -> Workbook:
        """Creates an Excel workbook from the data model.

        Args:
            tables (DataModelTableType): The data model in table
        """
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        index_by_sheet_name_column: dict[tuple[str, str], int] = {}
        for sheet_name, table in tables.items():
            if not table and sheet_name not in TableDMS.required_sheets():
                continue
            worksheet = workbook.create_sheet(title=sheet_name)
            if sheet_name == self.Sheets.metadata:
                self._write_metadata_to_worksheet(worksheet, table)
                continue
            if table:
                column_headers = list(table[0].keys())
            else:
                column_headers = TableDMS.get_sheet_column_by_name(sheet_name, column_type="all")
            self._write_table_to_worksheet(worksheet, table, MAIN_HEADERS_BY_SHEET_NAME[sheet_name], column_headers)
            for i, column in enumerate(column_headers, 1):
                index_by_sheet_name_column[(sheet_name, column)] = i

            if self._adjust_column_width:
                self._adjust_column_widths(worksheet)

        if self._add_dropdowns:
            self._add_drop_downs(workbook, index_by_sheet_name_column)
        return workbook

    @staticmethod
    def _write_metadata_to_worksheet(worksheet: Worksheet, table: list[dict[str, CellValueType]]) -> None:
        """Writes Metadata to the given worksheet.

        Metadata is written as key-value pairs without headers.
        """
        for row in table:
            worksheet.append(list(row.values()))

    def _write_table_to_worksheet(
        self, worksheet: Worksheet, table: list[dict[str, CellValueType]], main_header: str, column_headers: list[str]
    ) -> None:
        worksheet.append([main_header] + [""] * (len(column_headers) - 1))
        if self._style_headers:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(3, len(column_headers)))
            cell = worksheet.cell(row=1, column=1)
            cell.font = Font(bold=True, size=20)
            cell.fill = PatternFill(fgColor="FFC000", patternType="solid")

        worksheet.append(column_headers)
        header_row = 2 if main_header else 1
        if self._style_headers:
            for col_idx in range(1, len(column_headers) + 1):
                cell = worksheet.cell(row=header_row, column=col_idx)
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(fgColor="FFD966", patternType="solid")

        self._write_rows_to_worksheet(worksheet, table, column_headers)

        if self._style_headers:
            # openpyxl is not well typed
            worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)  # type: ignore[assignment]

    def _write_rows_to_worksheet(
        self, worksheet: Worksheet, table: list[dict[str, CellValueType]], headers: list[str]
    ) -> None:
        is_properties = worksheet.title == self.Sheets.properties
        fill_colors = itertools.cycle(["CADCFC", "FFFFFF"])
        fill_color = next(fill_colors)
        is_new_view = False
        last_view_value: CellValueType = None
        side = Side(style="thin")
        for row in table:
            if is_properties:
                is_new_view = row[self.PropertyColumns.view] != last_view_value and last_view_value is not None
            if is_new_view and is_properties and self._separate_view_properties:
                worksheet.append([None] * len(headers))  # Add an empty row between views
                if self._row_band_highlighting:
                    for cell in worksheet[worksheet.max_row]:
                        cell.border = Border(left=side, right=side, top=side, bottom=side)

            worksheet.append(list(row.values()))
            if self._row_band_highlighting and is_new_view and is_properties:
                fill_color = next(fill_colors)

            if self._row_band_highlighting and is_properties:
                for cell in worksheet[worksheet.max_row]:
                    cell.fill = PatternFill(fgColor=fill_color, fill_type="solid")
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

            if is_properties:
                last_view_value = row[self.PropertyColumns.view]

    @classmethod
    def _adjust_column_widths(cls, worksheet: Worksheet) -> None:
        for column_cells in worksheet.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            except ValueError:
                max_length = 0

            selected_column = column_cells[0]
            if isinstance(selected_column, MergedCell):
                selected_column = column_cells[1]

            current = worksheet.column_dimensions[selected_column.column_letter].width or (max_length + 0.5)  # type: ignore[union-attr]
            worksheet.column_dimensions[selected_column.column_letter].width = min(  # type: ignore[union-attr]
                max(current, max_length + 0.5), MAX_COLUMN_WIDTH
            )
        return None

    def _add_drop_downs(self, workbook: Workbook, index_by_sheet_name_column: dict[tuple[str, str], int]) -> None:
        """Adds drop down menus to specific columns for fast and accurate data entry

        Args:
            workbook: Workbook representation of the Excel file.
            index_by_sheet_name_column: A mapping of (sheet name, column name) to column index.
        """

        self._create_dropdown_source_sheet(workbook)

        property_sheet = workbook[self.Sheets.properties]
        self._add_validation(
            property_sheet,
            self.DropdownSourceColumns.view,
            self._max_views,
            index_by_sheet_name_column[(self.Sheets.properties, self.PropertyColumns.view)],
            self._max_views,
        )
        self._add_validation(
            property_sheet,
            self.DropdownSourceColumns.value_type,
            len(DMS_DATA_TYPES) - len(self.DROPDOWN_DMS_TYPE_EXCLUDE) + self._max_views,
            index_by_sheet_name_column[(self.Sheets.properties, self.PropertyColumns.value_type)],
            self._max_views * self._max_properties_per_view,
        )
        self._add_validation(
            property_sheet,
            self.DropdownSourceColumns.immutable,
            2,  # True, False
            index_by_sheet_name_column[(self.Sheets.properties, self.PropertyColumns.immutable)],
            self._max_views * self._max_properties_per_view,
        )
        self._add_validation(
            property_sheet,
            self.DropdownSourceColumns.container,
            self._max_containers,
            index_by_sheet_name_column[(self.Sheets.properties, self.PropertyColumns.container)],
            self._max_views * self._max_properties_per_view,
        )

        view_sheet = workbook[self.Sheets.views]
        self._add_validation(
            view_sheet,
            self.DropdownSourceColumns.implements,
            self._max_views + len(self._get_cognite_concepts()),
            index_by_sheet_name_column[(self.Sheets.views, self.ViewColumns.implements)],
            self._max_views,
        )
        container_sheet = workbook[self.Sheets.containers]
        self._add_validation(
            container_sheet,
            self.DropdownSourceColumns.used_for,
            3,  # node, edge, all
            index_by_sheet_name_column[(self.Sheets.containers, self.ContainerColumns.used_for)],
            self._max_containers,
        )

    def _add_validation(
        self, sheet: Worksheet, column_index: int, row_range: int, sheet_column_index: int, sheet_row_range: int
    ) -> None:
        """Adds data validation to a specific column in a sheet.

        Args:
            sheet: The worksheet to add the data validation to.
            column_index: The column index in the dropdown source sheet to use as the source for the drop-down.
            row_range: The number of rows in the dropdown source sheet to use as the source for the drop-down.
            sheet_column_index: The column index in the target sheet to add the data validation to.
            sheet_row_range: The number of rows in the target sheet to add the data validation to.
        """
        letter = get_column_letter(column_index)
        data_validation = DataValidation(
            type="list", formula1=f"={self.Sheets.dropdown_source}!${letter}$1:${letter}${row_range}"
        )
        sheet.add_data_validation(data_validation)
        target_letter = get_column_letter(sheet_column_index)
        data_validation.add(f"{target_letter}{HEADER_ROWS + 1}:{target_letter}{HEADER_ROWS + sheet_row_range}")

    def _create_dropdown_source_sheet(self, workbook: Workbook) -> None:
        """This methods creates a hidden sheet in the workbook which contains
        the source data for the drop-down menus.

        Args:
            workbook: Workbook representation of the Excel file.

        """
        dropdown_sheet = workbook.create_sheet(title=self.Sheets.dropdown_source)
        exclude = self.DROPDOWN_DMS_TYPE_EXCLUDE
        for no, dms_type in enumerate([type for type in DMS_DATA_TYPES.keys() if type not in exclude], 1):
            dropdown_sheet.cell(row=no, column=self.DropdownSourceColumns.value_type, value=dms_type)

        cognite_concepts = self._get_cognite_concepts()

        for i, concept in enumerate(cognite_concepts, 1):
            dropdown_sheet.cell(
                row=i,
                column=self.DropdownSourceColumns.implements,
                value=f"{CDF_CDM_SPACE}:{concept}(version={CDF_CDM_VERSION})",
            )

        dms_type_count = len(DMS_DATA_TYPES) - len(exclude)
        core_concept_count = len(cognite_concepts)
        for i in range(1, self._max_views + 1):
            source_row = i + HEADER_ROWS
            view_reference = f'=IF(ISBLANK({self.Sheets.views}!A{source_row}), "", {self.Sheets.views}!A{source_row})'
            dropdown_sheet.cell(row=i, column=self.DropdownSourceColumns.view, value=view_reference)
            dropdown_sheet.cell(
                row=i + core_concept_count, column=self.DropdownSourceColumns.implements, value=view_reference
            )
            dropdown_sheet.cell(
                row=i + dms_type_count, column=self.DropdownSourceColumns.value_type, value=view_reference
            )

        for i in range(1, self._max_containers + 1):
            source_row = i + HEADER_ROWS
            container_reference = (
                f'=IF(ISBLANK({self.Sheets.containers}!A{source_row}), "", {self.Sheets.containers}!A{source_row})'
            )
            dropdown_sheet.cell(row=i, column=self.DropdownSourceColumns.container, value=container_reference)

        for i, value in enumerate([True, False, None], 1):
            dropdown_sheet.cell(row=i, column=self.DropdownSourceColumns.in_model, value=value)

        for i, value in enumerate(["node", "edge", "all"], 1):
            dropdown_sheet.cell(row=i, column=self.DropdownSourceColumns.used_for, value=value)

        dropdown_sheet.sheet_state = "hidden"

    def _get_cognite_concepts(self) -> list[str]:
        """Gets the cognite concepts based on the dropdown_implements setting."""
        return _get_cognite_concepts(self._dropdown_implements)


@lru_cache(maxsize=1)
def _get_cognite_concepts(
    dropdown_implements: Set[Literal["main", "interface", "configuration", "annotation", "3D"]],
) -> list[str]:
    """Gets the cognite concepts based on the dropdown_implements setting.

    This is moved outside of the class to enable caching.
    """
    cognite_concepts: list[str] = []

    if "main" in dropdown_implements:
        cognite_concepts.extend(COGNITE_CONCEPTS_MAIN)
    if "interface" in dropdown_implements:
        cognite_concepts.extend(COGNITE_CONCEPTS_INTERFACES)
    if "configuration" in dropdown_implements:
        cognite_concepts.extend(COGNITE_CONCEPTS_CONFIGURATIONS)
    if "annotation" in dropdown_implements:
        cognite_concepts.extend(COGNITE_CONCEPTS_ANNOTATIONS)
    if "3D" in dropdown_implements:
        cognite_concepts.extend(COGNITE_CONCEPTS_3D)
    return cognite_concepts
