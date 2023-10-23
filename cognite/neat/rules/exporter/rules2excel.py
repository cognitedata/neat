from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from cognite.neat.rules.models.raw_rules import RawRules
from cognite.neat.rules.models.rules import Rules

from ._base import BaseExporter


class ExcelExporter(BaseExporter):
    """Class for exporting transformation rules object to excel file."""

    def __init__(self, rules: Rules | RawRules, filepath: Path, report_path: Path | None = None):
        super().__init__(rules, filepath, report_path)

    def export(self, filepath: Path | None = None):
        """Exports transformation rules to excel file."""
        filepath = filepath or self.filepath
        if not filepath:
            raise ValueError("No filepath given")

        data = self.generate_workbook()
        try:
            data.save(filepath)
            data.close()
        except Exception as e:
            data.close()
            raise e

        if self.report and self.report_path:
            self.report_path.write_text(self.report)

    def generate_workbook(self) -> Workbook:
        """Generates workbook from transformation rules."""
        data = Workbook()
        # Remove default sheet named "Sheet"
        data.remove(data["Sheet"])
        # Serialize the rules to the excel file

        # map rules metadata to excel
        metadata = self.rules.metadata
        metadata_sheet = data.create_sheet("Metadata")

        # add each metadata property to the sheet as a row
        metadata_sheet.append(["prefix", metadata.prefix])
        metadata_sheet.append(["namespace", metadata.namespace])
        metadata_sheet.append(["dataModelName", metadata.data_model_name])
        metadata_sheet.append(["cdfSpaceName", metadata.cdf_space_name])
        metadata_sheet.append(["title", metadata.title])
        metadata_sheet.append(["description", metadata.description])
        metadata_sheet.append(["version", metadata.version])
        metadata_sheet.append(["isCurrentVersion", metadata.is_current_version])
        metadata_sheet.append(
            [
                "creator",
                ",".join(metadata.creator if isinstance(metadata.creator, list) else [metadata.creator])
                if metadata.creator
                else "",
            ]
        )
        metadata_sheet.append(
            [
                "contributor",
                ",".join(metadata.contributor if isinstance(metadata.contributor, list) else [metadata.contributor])
                if metadata.contributor
                else "",
            ]
        )
        metadata_sheet.append(["created", metadata.created])
        metadata_sheet.append(["updated", metadata.updated])
        metadata_sheet.append(["rights", metadata.rights])

        # map classes to excel sheet named "Classes" and add each class as a row
        classes_sheet = data.create_sheet("Classes")

        classes_sheet.append(["Solution model", "", "", "Knowledge acquisition", "", "", ""])
        classes_sheet.merge_cells("A1:C1")
        classes_sheet.merge_cells("D1:G1")
        classes_sheet.append(
            ["Class", "Description", "Parent Class", "Source", "Source Entity Name", "Match Type", "Comment"]
        )  # A  ... # G

        for class_ in self.rules.classes.values():
            classes_sheet.append(
                [
                    class_.class_id,
                    class_.description,
                    class_.parent_class,
                    class_.source,
                    class_.source_entity_name,
                    class_.match_type,
                    class_.comment,
                ]
            )

        # map properties to excel sheet named "Properties" and add each property as a row
        properties_sheet = data.create_sheet("Properties")
        properties_sheet.append(
            [
                "Solution model",  # A
                "",  # B
                "",  # C
                "",  # D
                "",  # E
                "",  # F
                "Solution classic CDF resources",  # G
                "",  # H
                "",  # I
                "",  # J
                "",  # K
                "",  # L
                "Transformation rules",  # M
                "",  # N
                "Knowledge acquisition",  # O
                "",  # P
                "",  # Q
                "",  # R
            ]
        )
        properties_sheet.merge_cells("A1:F1")
        properties_sheet.merge_cells("G1:L1")
        properties_sheet.merge_cells("M1:N1")
        properties_sheet.merge_cells("O1:R1")
        properties_sheet.append(
            [
                "Class",  # A
                "Property",  # B
                "Description",  # C
                "Type",  # D
                "Min Count",  # E
                "Max Count",  # F
                "Resource Type",  # G
                "Resource Type Property",  # H
                "Relationship Source Type",  # I
                "Relationship Target Type",  # J
                "Relationship Label",  # K
                "Relationship ExternalID Rule",  # L
                "Rule Type",  # M
                "Rule",  # N
                "Source",  # O
                "Source Entity Name",  # P
                "Match Type",  # Q
                "Comment",  # R
            ]
        )

        for property_ in self.rules.properties.values():
            properties_sheet.append(
                [
                    property_.class_id,  # A
                    property_.property_id,  # B
                    property_.description,  # C
                    property_.expected_value_type,  # D
                    property_.min_count,  # E
                    property_.max_count,  # F
                    ",".join(property_.cdf_resource_type) if property_.cdf_resource_type else "",  # G
                    ",".join(property_.resource_type_property) if property_.resource_type_property else "",  # H
                    property_.source_type,  # I
                    property_.target_type,  # J
                    property_.label,  # K
                    property_.relationship_external_id_rule,  # L
                    property_.rule_type,  # M
                    property_.rule,  # N
                    property_.source,  # O
                    property_.source_entity_name,  # P
                    property_.match_type,  # Q
                    property_.comment,  # R
                ]
            )

        prefixes_sheet = data.create_sheet("Prefixes")
        prefixes_sheet.append(
            [
                "Prefix",  # A
                "URI",  # B
            ]
        )

        for prefix, uri in self.rules.prefixes.items():
            prefixes_sheet.append(
                [
                    prefix,  # A
                    uri,  # B
                ]
            )

        return self.set_header_style(data)

    @staticmethod
    def set_header_style(data: Workbook):
        """Sets the header style for all sheets in the self.workbook"""
        style = NamedStyle(name="header style")
        style.font = Font(bold=True, size=16)
        side = Side(style="thin", color="000000")
        style.border = Border(left=side, right=side, top=side, bottom=side)
        data.add_named_style(style)

        for sheet in data.sheetnames:
            if sheet == "Metadata":
                continue
            if sheet == "Classes" or sheet == "Properties":
                sheet_obj = data[sheet]
                if sheet == "Classes":
                    sheet_obj.freeze_panes = "A3"
                else:
                    sheet_obj.freeze_panes = "D3"

                for cell in sheet_obj[1]:
                    cell = cast(Cell, cell)  # type: ignore[index]
                    cell.style = style
                    cell.fill = PatternFill("solid", start_color="D5B2CF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for cell in sheet_obj[2]:
                    cell = cast(Cell, cell)  # type: ignore[index]
                    cell.style = style
                    cell.fill = PatternFill("solid", start_color="D5DBD5")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    adjusted_width = (len(str(cell.value)) + 5) * 1.2
                    data[sheet].column_dimensions[cell.column_letter].width = adjusted_width

        return data
