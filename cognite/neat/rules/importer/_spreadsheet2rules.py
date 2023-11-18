"""This module performs importing of graph to TransformationRules pydantic class.
In more details, it traverses the graph and abstracts class and properties, basically
generating a list of rules based on which nodes that form the graph are made.
"""


from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from cognite.neat.rules.importer._base import BaseImporter
from cognite.neat.utils.auxiliary import local_import


class ExcelImporter(BaseImporter):
    def __init__(self, filepath: Path):
        self.filepath = filepath

    def to_tables(self) -> dict[str, pd.DataFrame]:
        workbook: Workbook = load_workbook(self.filepath)

        return {
            sheet_name: pd.read_excel(
                self.filepath,
                sheet_name=sheet_name,
                header=None if sheet_name == "Metadata" else 0,
                skiprows=1 if sheet_name in ["Classes", "Properties", "Instances"] else None,
            )
            for sheet_name in workbook.sheetnames
        }


class GoogleSheetImporter(BaseImporter):
    def __init__(self, sheet_id: str):
        self.sheet_id = sheet_id

    def to_tables(self) -> dict[str, pd.DataFrame]:
        local_import("gspread", "google")
        import gspread  # type: ignore[import]

        client_google = gspread.service_account()
        google_sheet = client_google.open_by_key(self.sheet_id)

        return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in google_sheet.worksheets()}
