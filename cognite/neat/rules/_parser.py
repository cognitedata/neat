from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

import logging
import warnings
from typing import Any, Hashable, Literal, overload
from warnings import warn

from pydantic import field_validator
from pydantic_core import ErrorDetails, ValidationError

from openpyxl import Workbook, load_workbook
import pandas as pd
import requests

from rdflib import Namespace


from cognite.neat.utils.auxiliary import local_import
from cognite.neat.app.api.configuration import PREFIXES
from cognite.neat.rules import _exceptions
from cognite.neat.rules.models import Class, Metadata, Property, RuleModel, TransformationRules


@overload
def parse_rules_from_excel_file(filepath: Path, return_report: Literal[False] = False) -> TransformationRules:
    ...


@overload
def parse_rules_from_excel_file(
    filepath: Path, return_report: Literal[True]
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None]:
    ...


def parse_rules_from_excel_file(
    filepath: Path, return_report: bool = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from an Excel file.

    Parameters
    ----------
    filepath : Path
        Path to the Excel file
    return_report : bool, optional
        Whether to return a report, by default False

    Returns
    -------
        Either the transformation rules or a tuple with the rules, a list of errors and a list of warnings
    """
    return from_tables(read_excel_file_to_table_by_name(filepath), return_report)


def parse_rules_from_google_sheet(
    sheet_id: str, return_report: bool = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from a Google sheet.

    Parameters
    ----------
    sheet_id : str
        The identifier of the Google sheet with the rules.
    return_report : bool, optional
        Whether to return a report, by default False

    Returns
    -------
        Either the transformation rules or a tuple with the rules, a list of errors and a list of warnings
    """
    return from_tables(read_google_sheet_to_table_by_name(sheet_id), return_report)


def parse_rules_from_github_sheet(
    filepath: Path,
    personal_token: str,
    owner: str,
    repo: str,
    branch: str = "main",
    return_report: bool = False,
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from a sheet stored in private Github.

    Parameters
    ----------
    filepath : Path
        Path to the sheet in the Github repository.
    personal_token : str
        Personal access token to access the Github repository.
    owner : str
        Owner of the Github repository.
    repo : str
        Name of the Github repository.
    branch : str, optional
        Branch of the Github repository, by default "main".

    Returns
    -------
        Either the transformation rules or a tuple with the rules, a list of errors and a list of warnings
    """

    return from_tables(read_github_sheet_to_table_by_name(filepath, personal_token, owner, repo, branch), return_report)


def parse_rules_from_yaml(dirpath: Path) -> TransformationRules:
    """
    Load transformation rules from a yaml file.

    Args:
        dirpath (Path): Path to the yaml file.
    Returns:
        TransformationRules: The transformation rules.
    """
    return TransformationRules(**read_yaml_file_to_mapping_by_name(dirpath))


def from_tables(
    raw_dfs: dict[str, pd.DataFrame], return_report: bool = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    # the only way to suppress warnings from pylense
    validation_warnings = []
    try:
        with warnings.catch_warnings(record=True) as validation_warnings:
            raw_tables = RawTables.from_raw_dataframes(raw_dfs)
            rules_dict: dict[str, Any] = {
                "metadata": _parse_metadata(raw_tables.Metadata),
                "classes": _parse_classes(raw_tables.Classes),
                "properties": _parse_properties(raw_tables.Properties),
                "prefixes": PREFIXES if raw_tables.Prefixes.empty else _parse_prefixes(raw_tables.Prefixes),
            }

            rules_dict["instances"] = (
                None
                if raw_tables.Instances.empty
                else _parse_instances(raw_tables.Instances, rules_dict["metadata"], rules_dict["prefixes"])
            )
            rules = TransformationRules(**rules_dict)
        return (rules, None, _exceptions.wrangle_warnings(validation_warnings)) if return_report else rules

    except _exceptions.Error0 as e:
        validation_errors = [e.to_error_dict()]
        if return_report:
            return None, validation_errors, _exceptions.wrangle_warnings(validation_warnings)
        else:
            raise e
    except ValidationError as e:
        validation_errors = e.errors()
        if return_report:
            return None, validation_errors, _exceptions.wrangle_warnings(validation_warnings)
        else:
            raise e


def _parse_metadata(meta_df: pd.DataFrame) -> dict[str, Any]:
    metadata_dict = dict(zip(meta_df[0], meta_df[1]))
    metadata_dict["source"] = meta_df.source if "source" in dir(meta_df) else None
    if "namespace" in metadata_dict:
        metadata_dict["namespace"] = Namespace(metadata_dict["namespace"])
    return metadata_dict


def _parse_classes(classes_df: pd.DataFrame) -> dict[Any | None, dict[Hashable, Any]]:
    return {class_.get("Class"): class_ for class_ in classes_df.to_dict(orient="records")}


def _parse_properties(properties_df: pd.DataFrame) -> dict[str, dict[Hashable, Any]]:
    return {f"row {i+3}": property_ for i, property_ in enumerate(properties_df.to_dict(orient="records"))}


def _parse_prefixes(prefix_df: pd.DataFrame) -> dict[str, Namespace]:
    return {row["Prefix"]: Namespace(row["URI"]) for i, row in prefix_df.iterrows()}


def _parse_instances(
    instances_df: pd.DataFrame, metadata: dict[str, Any], prefixes: dict[str, Namespace]
) -> list[dict] | None:
    if "prefix" not in metadata or "namespace" not in metadata:
        logging.warning(_exceptions.Warning500().message)
        warn(_exceptions.Warning500().message)
        return None

    prefixes[metadata["prefix"]] = metadata["namespace"]

    instances = []
    for _, row in instances_df.iterrows():
        row_as_dict = row.to_dict()
        row_as_dict["namespace"] = metadata["namespace"]
        row_as_dict["prefixes"] = prefixes
        instances.append(row_as_dict)
    return instances


class RawTables(RuleModel):
    Metadata: pd.DataFrame
    Properties: pd.DataFrame
    Classes: pd.DataFrame
    Prefixes: pd.DataFrame = pd.DataFrame()
    Instances: pd.DataFrame = pd.DataFrame()

    @classmethod
    def from_raw_dataframes(cls, raw_dfs: dict[str, pd.DataFrame]) -> "RawTables":
        expected_tables = cls.mandatory()

        # Validate raw tables
        if missing_tables := (expected_tables - set(raw_dfs)):
            raise _exceptions.Error0(missing_tables)

        tables_dict = {
            Tables.metadata: raw_dfs[Tables.metadata],
            Tables.classes: raw_dfs[Tables.classes],
            Tables.properties: raw_dfs[Tables.properties],
        }

        if Tables.prefixes in raw_dfs:
            tables_dict[Tables.prefixes] = raw_dfs[Tables.prefixes]
        if Tables.instances in raw_dfs:
            tables_dict[Tables.instances] = raw_dfs[Tables.instances]

        return cls(**tables_dict)

    @field_validator("Metadata")
    def has_metadata_mandatory_rows(cls, v):
        given_rows = set(v[0].values)
        mandatory_rows = Metadata.mandatory()
        mandatory_rows_alias = Metadata.mandatory(use_alias=True)

        if not (mandatory_rows.issubset(given_rows) or mandatory_rows_alias.issubset(given_rows)):
            missing_rows = mandatory_rows_alias.difference(given_rows)
            raise _exceptions.Error51(missing_rows).to_pydantic_custom_error()
        return v

    @field_validator("Classes")
    def has_classes_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = Class.mandatory()
        mandatory_columns_alias = Class.mandatory(use_alias=True)

        if not (mandatory_columns.issubset(given_columns) or mandatory_columns_alias.issubset(given_columns)):
            missing_columns = mandatory_columns_alias.difference(given_columns)
            raise _exceptions.Error52(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Properties")
    def has_properties_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = Property.mandatory()
        mandatory_columns_alias = Property.mandatory(use_alias=True)

        if not (mandatory_columns.issubset(given_columns) or mandatory_columns_alias.issubset(given_columns)):
            missing_columns = mandatory_columns_alias.difference(given_columns)
            raise _exceptions.Error53(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Prefixes")
    def has_prefixes_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = {"Prefix", "URI"}

        if not mandatory_columns.issubset(given_columns):
            missing_columns = mandatory_columns.difference(given_columns)
            raise _exceptions.Error54(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Instances")
    def has_instances_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = {"Instance", "Property", "Value"}

        if not mandatory_columns.issubset(given_columns):
            missing_columns = mandatory_columns.difference(given_columns)
            raise _exceptions.Error55(missing_columns).to_pydantic_custom_error()
        return v


class Tables:
    prefixes = "Prefixes"
    properties = "Properties"
    classes = "Classes"
    metadata = "Metadata"
    instances = "Instances"


# readers:


def read_google_sheet_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if gspread is not installed
    local_import("gspread", "google")
    import gspread

    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def read_excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if openpyxl is not installed
    local_import("openpyxl", "excel")

    from openpyxl import Workbook, load_workbook

    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].source = filepath

    return sheets


def read_yaml_file_to_mapping_by_name(dirpath: Path, expected_files: set[str] | None = None) -> dict[str, dict]:
    # To trigger ImportError if yaml is not installed
    local_import("yaml", "yaml")
    from yaml import safe_load

    mapping_by_name = {}
    for filepath in dirpath.iterdir():
        if expected_files is not None and filepath.stem not in expected_files:
            continue
        mapping_by_name[filepath.stem] = safe_load(filepath.read_text())
    return mapping_by_name


def read_github_sheet_to_table_by_name(
    filepath: str, personal_token: str, owner: str, repo: str, branch: str = "main"
) -> dict[str, pd.DataFrame]:
    r = requests.get(
        f"https://api.github.com/repos/{owner}/{repo}/contents/{filepath}?ref={branch}",
        headers={"accept": "application/vnd.github.v3.raw", "authorization": f"token {personal_token}"},
    )

    loc = f"https://github.com/{owner}/{repo}/tree/{branch}"

    if r.status_code != 200:
        raise _exceptions.Error20(filepath, loc, r.reason)
    try:
        wb = load_workbook(BytesIO(r.content), data_only=True)
    except BadZipFile:
        raise _exceptions.Error21(filepath, loc)
    return _workbook_to_table_by_name(wb)


def _workbook_to_table_by_name(workbook: Workbook) -> dict[str, pd.DataFrame]:
    table = {}
    for sheet in workbook:
        sheetname = sheet.title
        data = sheet.values
        if sheetname == "Metadata":
            table[sheetname] = pd.DataFrame(data, columns=None)
        if sheetname in ["Classes", "Properties", "Instances"]:
            next(data)
            columns = next(data)[:]
            table[sheet.title] = pd.DataFrame(data, columns=columns).dropna(how="all")

    return table
