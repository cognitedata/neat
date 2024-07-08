from __future__ import annotations

import itertools
import json
from datetime import datetime, timezone
from pathlib import Path
from types import GenericAlias
from typing import Any, ClassVar, Literal, cast, get_args

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from cognite.neat.rules._shared import Rules
from cognite.neat.rules.models import (
    DataModelType,
    ExtensionCategory,
    RoleTypes,
    SchemaCompleteness,
    SheetEntity,
)
from cognite.neat.rules.models.dms import DMSMetadata
from cognite.neat.rules.models.domain import DomainMetadata
from cognite.neat.rules.models.information import InformationMetadata

from ._base import BaseExporter


class ExcelExporter(BaseExporter[Workbook]):
    """Export rules to Excel.

    Args:
        styling: The styling to use for the Excel file. Defaults to "default". See below for details
            on the different styles.
        output_role: The role to use for the exported spreadsheet. If provided, the rules will be converted to
            this role formate before being written to excel. If not provided, the role from the rules will be used.
        dump_as: This determines how the rules are written to the Excel file. An Excel file has up to three sets of
           sheets: user, last, and reference. The user sheets are used for inputting rules from a user. The last sheets
           are used for the last version of the same model as the user, while the reference sheets are used for
           the model the user is building on. The options are:
             * "user": The rules are written to the user sheets. This is used when you want to modify the rules
                directly and potentially change the model. This is useful when you have imported the data model
                from outside CDF and you want to modify it before you write it to CDF.
             * "last": The rules are written to the last sheets. This is used when you want to extend the rules,
               but have validation that you are not breaking the existing model. This is used when you want to
               change a model that has already been published to CDF and that model is in production.
             * "reference": The rules are written to the reference sheets. This is typically used when you want to build
               a new solution on top of an enterprise model.
        new_model_id: The new model ID to use for the exported spreadsheet. This is only applicable if the input
            rules have 'is_reference' set. If provided, the model ID will be used to automatically create the
            new metadata sheet in the Excel file. The model id is expected to be a tuple of (prefix, title)
            (space, external_id) for InformationRules and DMSRules respectively.

    The following styles are available:

    - "none":    No styling is applied.
    - "minimal": Column widths are adjusted to fit the content, and the header row(s) is frozen.
    - "default": Minimal + headers are bold, increased size, and colored.
    - "maximal": Default + alternating row colors in the properties sheet for each class in addition to extra
                 blank rows between classes and borders
    """

    Style = Literal["none", "minimal", "default", "maximal"]
    DumpOptions = Literal["user", "last", "reference"]
    _main_header_by_sheet_name: ClassVar[dict[str, str]] = {
        "Properties": "Definition of Properties per Class",
        "Classes": "Definition of Classes",
        "Views": "Definition of Views",
        "Containers": "Definition of Containers",
    }
    style_options = get_args(Style)
    dump_options = get_args(DumpOptions)

    def __init__(
        self,
        styling: Style = "default",
        output_role: RoleTypes | None = None,
        dump_as: DumpOptions = "user",
        new_model_id: tuple[str, str] | None = None,
    ):
        if styling not in self.style_options:
            raise ValueError(f"Invalid styling: {styling}. Valid options are {self.style_options}")
        if dump_as not in self.dump_options:
            raise ValueError(f"Invalid dump_as: {dump_as}. Valid options are {self.dump_options}")
        self.styling = styling
        self._styling_level = self.style_options.index(styling)
        self.output_role = output_role
        self.new_model_id = new_model_id
        self.dump_as = dump_as

    def export_to_file(self, rules: Rules, filepath: Path) -> None:
        """Exports transformation rules to excel file."""
        data = self.export(rules)
        try:
            data.save(filepath)
        finally:
            data.close()
        return None

    def export(self, rules: Rules) -> Workbook:
        rules = self._convert_to_output_role(rules, self.output_role)
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        dumped_user_rules: dict[str, Any]
        dumped_last_rules: dict[str, Any] | None = None
        dumped_reference_rules: dict[str, Any] | None = None
        if self.dump_as != "user":
            action = {"last": "update", "reference": "create"}[self.dump_as]
            metadata_creator = _MetadataCreator(action, self.new_model_id)  # type: ignore[arg-type]

            dumped_user_rules = {
                "Metadata": metadata_creator.create(rules.metadata),
            }

            if self.dump_as == "last":
                dumped_last_rules = rules.dump(by_alias=True)
                if rules.reference:
                    dumped_reference_rules = rules.reference.dump(by_alias=True, as_reference=True)
            elif self.dump_as == "reference":
                dumped_reference_rules = rules.dump(by_alias=True, as_reference=True)
        else:
            dumped_user_rules = rules.dump(by_alias=True)
            if rules.last:
                dumped_last_rules = rules.last.dump(by_alias=True)
            if rules.reference:
                dumped_reference_rules = rules.reference.dump(by_alias=True, as_reference=True)

        self._write_metadata_sheet(workbook, dumped_user_rules["Metadata"])
        self._write_sheets(workbook, dumped_user_rules, rules)
        if dumped_last_rules:
            self._write_sheets(workbook, dumped_last_rules, rules, sheet_prefix="Last")
            self._write_metadata_sheet(workbook, dumped_last_rules["Metadata"], sheet_prefix="Last")

        if dumped_reference_rules:
            self._write_sheets(workbook, dumped_reference_rules, rules, sheet_prefix="Ref")
            self._write_metadata_sheet(workbook, dumped_reference_rules["Metadata"], sheet_prefix="Ref")

        if self._styling_level > 0:
            self._adjust_column_widths(workbook)

        return workbook

    def _write_sheets(
        self,
        workbook: Workbook,
        dumped_rules: dict[str, Any],
        rules: Rules,
        sheet_prefix: str = "",
    ):
        for sheet_name, headers in rules.headers_by_sheet(by_alias=True).items():
            if sheet_name in ("Metadata", "Prefixes", "Reference", "Last"):
                continue
            sheet = workbook.create_sheet(f"{sheet_prefix}{sheet_name}")

            main_header = self._main_header_by_sheet_name[sheet_name]
            sheet.append([main_header] + [""] * (len(headers) - 1))
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            sheet.append(headers)

            fill_colors = itertools.cycle(["CADCFC", "FFFFFF"])
            fill_color = next(fill_colors)
            last_class: str | None = None
            item: dict[str, Any]
            for item in dumped_rules.get(sheet_name) or []:
                row = list(item.values())
                class_ = row[0]

                is_properties = sheet_name == "Properties"
                is_new_class = class_ != last_class and last_class is not None
                if self._styling_level > 2 and is_new_class and is_properties:
                    sheet.append([""] * len(headers))
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                    fill_color = next(fill_colors)

                sheet.append(row)
                if self._styling_level > 2 and is_properties:
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                last_class = class_

            if self._styling_level > 0:
                # This freezes all rows above the given row
                sheet.freeze_panes = sheet["A3"]

                sheet["A1"].alignment = Alignment(horizontal="center")

            if self._styling_level > 1:
                # Make the header row bold, larger, and colored
                sheet["A1"].font = Font(bold=True, size=20)
                sheet["A1"].fill = PatternFill(fgColor="FFC000", patternType="solid")
                for cell in sheet["2"]:
                    cell.font = Font(bold=True, size=14)

    def _write_metadata_sheet(self, workbook: Workbook, metadata: dict[str, Any], sheet_prefix: str = "") -> None:
        # Excel does not support timezone in datetime strings
        if isinstance(metadata.get("created"), datetime):
            metadata["created"] = metadata["created"].replace(tzinfo=None)
        if isinstance(metadata.get("updated"), datetime):
            metadata["updated"] = metadata["updated"].replace(tzinfo=None)

        metadata_sheet = workbook.create_sheet(f"{sheet_prefix}Metadata")
        for key, value in metadata.items():
            metadata_sheet.append([key, value])

        if self._styling_level > 1:
            for cell in metadata_sheet["A"]:
                cell.font = Font(bold=True, size=12)

    @classmethod
    def _get_item_class(cls, annotation: GenericAlias) -> type[SheetEntity]:
        if not isinstance(annotation, GenericAlias):
            raise ValueError(f"Expected annotation to be a GenericAlias, but got {type(annotation)}")
        args = get_args(annotation)
        if len(args) != 1:
            raise ValueError(f"Expected annotation to have exactly one argument, but got {len(args)}")
        arg = args[0]
        if not issubclass(arg, SheetEntity):
            raise ValueError(f"Expected annotation to have a BaseModel argument, but got {type(arg)}")
        return arg

    @classmethod
    def _adjust_column_widths(cls, workbook: Workbook) -> None:
        for sheet_ in workbook:
            sheet = cast(Worksheet, sheet_)
            for column_cells in sheet.columns:
                try:
                    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                except ValueError:
                    max_length = 0

                selected_column = column_cells[0]
                if isinstance(selected_column, MergedCell):
                    selected_column = column_cells[1]

                current = sheet.column_dimensions[selected_column.column_letter].width or (max_length + 0.5)
                sheet.column_dimensions[selected_column.column_letter].width = max(current, max_length + 0.5)
        return None


class _MetadataCreator:
    creator_name = "<YOUR NAME>"

    def __init__(
        self,
        action: Literal["create", "update"],
        new_model_id: tuple[str, str] | None = None,
    ):
        self.action = action
        self.new_model_id = new_model_id or ("YOUR_PREFIX", "YOUR_TITLE")

    def create(self, metadata: DomainMetadata | InformationMetadata | DMSMetadata) -> dict[str, Any]:
        now = datetime.now(timezone.utc).replace(microsecond=0, tzinfo=None)
        if self.action == "update":
            output = json.loads(metadata.model_dump_json(by_alias=True))
            # This is the same for Information and DMS
            output["updated"] = now.isoformat()
            output["schema"] = SchemaCompleteness.extended.value
            output["extension"] = ExtensionCategory.addition.value
            if value := output.get("creator"):
                output["creator"] = f"{value}, {self.creator_name}"
            else:
                output["creator"] = self.creator_name
            return output

        # Action "create"
        if isinstance(metadata, DomainMetadata):
            output = {field_alias: None for field_alias in metadata.model_dump(by_alias=True).keys()}
            output["role"] = metadata.role.value
            output["creator"] = self.creator_name
            return output

        new_metadata = self._create_new_info(now)
        if isinstance(metadata, DMSMetadata):
            from cognite.neat.rules.models.information._converter import (
                _InformationRulesConverter,
            )

            output_metadata: DMSMetadata | InformationMetadata = _InformationRulesConverter._convert_metadata_to_dms(
                new_metadata
            )
        elif isinstance(metadata, InformationMetadata):
            output_metadata = new_metadata
        else:
            raise ValueError(f"Bug in Neat: Unknown metadata type: {type(metadata)}")

        created = json.loads(output_metadata.model_dump_json(by_alias=True))
        created.pop("extension", None)
        return created

    def _create_new_info(self, now: datetime) -> InformationMetadata:
        prefix = self.new_model_id[0]
        return InformationMetadata(
            data_model_type=DataModelType.solution,
            schema_=SchemaCompleteness.complete,
            extension=ExtensionCategory.addition,
            prefix=prefix,
            namespace=f"http://purl.org/neat/{prefix}/",  # type: ignore[arg-type]
            description=None,
            version="1",
            created=now,
            updated=now,
            creator=[self.creator_name],
            name=self.new_model_id[1],
        )
