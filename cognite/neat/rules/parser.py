import logging
import warnings
from collections.abc import Hashable
from io import BytesIO
from pathlib import Path
from typing import Any, Literal, overload
from warnings import warn
from zipfile import BadZipFile

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from pydantic import field_validator
from pydantic_core import ErrorDetails, ValidationError
from rdflib import Namespace

from cognite.neat.constants import PREFIXES
from cognite.neat.exceptions import wrangle_warnings
from cognite.neat.rules import exceptions
from cognite.neat.rules.models import Class, Metadata, Property, RuleModel, TransformationRules
from cognite.neat.utils.auxiliary import local_import


@overload
def parse_rules_from_excel_file(filepath: Path, return_report: Literal[False] = False) -> TransformationRules:
    ...


@overload
def parse_rules_from_excel_file(
    filepath: Path, return_report: Literal[True]
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None]:
    ...


def parse_rules_from_excel_file(
    filepath: Path, return_report: Literal[True, False] = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from an Excel file.

    Args:
      filepath: Path to the Excel file
      return_report: Whether to return a report, by default False

    Returns:
        The transformation rules, and optionally one list of validation errors and a one list of warnings.
    """
    return from_tables(read_excel_file_to_table_by_name(filepath), return_report)


@overload
def parse_rules_from_google_sheet(sheet_id: str, return_report: Literal[False] = False) -> TransformationRules:
    ...


@overload
def parse_rules_from_google_sheet(
    sheet_id: str, return_report: Literal[True]
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None]:
    ...


def parse_rules_from_google_sheet(
    sheet_id: str, return_report: Literal[True, False] = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from a Google sheet.

    Args:
      sheet_id: The identifier of the Google sheet with the rules.
      return_report: Whether to return a report, by default False

    Returns:
        The transformation rules, and optionally one list of validation errors and a one list of warnings.
    """
    return from_tables(read_google_sheet_to_table_by_name(sheet_id), return_report)


@overload
def parse_rules_from_github_sheet(
    filepath: Path,
    personal_token: str,
    owner: str,
    repo: str,
    return_report: Literal[False] = False,
    branch: str = "main",
) -> TransformationRules:
    ...


@overload
def parse_rules_from_github_sheet(
    filepath: Path,
    personal_token: str,
    owner: str,
    repo: str,
    return_report: Literal[True],
    branch: str = "main",
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None]:
    ...


def parse_rules_from_github_sheet(
    filepath: Path,
    personal_token: str,
    owner: str,
    repo: str,
    return_report: Literal[True, False] = False,
    branch: str = "main",
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    """Parse transformation rules from a sheet stored in private GitHub.

    Args:
      filepath: Path to the sheet in the GitHub repository.
      personal_token: Personal access token to access the GitHub repository.
      owner: Owner of the GitHub repository.
      repo: Name of the GitHub repository.
      branch: Branch of the GitHub repository, by default "main".
      return_report: Whether to return a report, by default False

    Returns:
        The transformation rules, and optionally one list of validation errors and a one list of warnings.

    """
    tables = read_github_sheet_to_table_by_name(str(filepath), personal_token, owner, repo, branch)
    return from_tables(tables, return_report)


def parse_rules_from_yaml(folder_path: Path) -> TransformationRules:
    """Parse transformation rules from a yaml files.

    The yaml files must be named "metadata.yaml", "classes.yaml", "properties.yaml", "prefixes.yaml"
    and "instances.yaml". These must be located in the same directory.

    Args:
      folder_path: The directory where the yaml files are located.

    Returns:
      The transformation rules.

    """
    return TransformationRules(**read_yaml_file_to_mapping_by_name(folder_path))  # type: ignore[arg-type]


@overload
def from_tables(raw_dfs: dict[str, pd.DataFrame], return_report: Literal[False] = False) -> TransformationRules:
    ...


@overload
def from_tables(
    raw_dfs: dict[str, pd.DataFrame], return_report: Literal[True]
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None]:
    ...


def from_tables(
    raw_dfs: dict[str, pd.DataFrame], return_report: bool = False
) -> tuple[TransformationRules | None, list[ErrorDetails] | None, list | None] | TransformationRules:
    # the only way to suppress warnings from pylense
    validation_warnings = []
    try:
        with warnings.catch_warnings(record=True) as validation_warnings:
            raw_tables = RawTables.from_raw_dataframes(raw_dfs)
            rules_dict: dict[str, Any] = {
                "metadata": _parse_metadata(raw_tables.Metadata),
                "classes": _parse_classes(raw_tables.Classes),
                "properties": _parse_properties(raw_tables.Properties),
                "prefixes": PREFIXES if raw_tables.Prefixes.empty else _parse_prefixes(raw_tables.Prefixes),
            }

            rules_dict["instances"] = (
                None
                if raw_tables.Instances.empty
                else _parse_instances(raw_tables.Instances, rules_dict["metadata"], rules_dict["prefixes"])
            )
            rules = TransformationRules(**rules_dict)
        return (rules, None, wrangle_warnings(validation_warnings)) if return_report else rules

    except exceptions.ExcelFileMissingMandatorySheets as e:
        validation_errors = [e.to_error_dict()]
        if return_report:
            return None, validation_errors, wrangle_warnings(validation_warnings)
        else:
            raise e
    except ValidationError as e:
        validation_errors = e.errors()
        if return_report:
            return None, validation_errors, wrangle_warnings(validation_warnings)
        else:
            raise e


def _parse_metadata(meta_df: pd.DataFrame) -> dict[str, Any]:
    metadata_dict = dict(zip(meta_df[0], meta_df[1], strict=True))
    metadata_dict["source"] = meta_df.source if "source" in dir(meta_df) else None
    if "namespace" in metadata_dict:
        metadata_dict["namespace"] = Namespace(metadata_dict["namespace"])
    return metadata_dict


def _parse_classes(classes_df: pd.DataFrame) -> dict[Any | None, dict[Hashable, Any]]:
    return {class_.get("Class"): class_ for class_ in classes_df.to_dict(orient="records")}


def _parse_properties(properties_df: pd.DataFrame) -> dict[str, dict[Hashable, Any]]:
    return {f"row {i+3}": property_ for i, property_ in enumerate(properties_df.to_dict(orient="records"))}


def _parse_prefixes(prefix_df: pd.DataFrame) -> dict[str, Namespace]:
    return {row["Prefix"]: Namespace(row["URI"]) for i, row in prefix_df.iterrows()}


def _parse_instances(
    instances_df: pd.DataFrame, metadata: dict[str, Any], prefixes: dict[str, Namespace]
) -> list[dict] | None:
    if "prefix" not in metadata or "namespace" not in metadata:
        logging.warning(exceptions.MissingDataModelPrefixOrNamespace().message)
        warn(exceptions.MissingDataModelPrefixOrNamespace().message, stacklevel=2)
        return None

    prefixes[metadata["prefix"]] = metadata["namespace"]

    instances = []
    for _, row in instances_df.iterrows():
        row_as_dict = row.to_dict()
        row_as_dict["namespace"] = metadata["namespace"]
        row_as_dict["prefixes"] = prefixes
        instances.append(row_as_dict)
    return instances


class RawTables(RuleModel):
    Metadata: pd.DataFrame
    Properties: pd.DataFrame
    Classes: pd.DataFrame
    Prefixes: pd.DataFrame = pd.DataFrame()
    Instances: pd.DataFrame = pd.DataFrame()

    @classmethod
    def from_raw_dataframes(cls, raw_dfs: dict[str, pd.DataFrame]) -> "RawTables":
        expected_tables = cls.mandatory_fields()

        # Validate raw tables
        if missing_tables := (expected_tables - set(raw_dfs)):
            raise exceptions.ExcelFileMissingMandatorySheets(missing_tables)

        tables_dict = {
            Tables.metadata: raw_dfs[Tables.metadata],
            Tables.classes: cls.drop_non_string_columns(raw_dfs[Tables.classes]),
            Tables.properties: cls.drop_non_string_columns(raw_dfs[Tables.properties]),
        }

        if Tables.prefixes in raw_dfs:
            tables_dict[Tables.prefixes] = cls.drop_non_string_columns(raw_dfs[Tables.prefixes])
        if Tables.instances in raw_dfs:
            tables_dict[Tables.instances] = cls.drop_non_string_columns(raw_dfs[Tables.instances])

        return cls(**tables_dict)

    @field_validator("Metadata")
    def has_metadata_mandatory_rows(cls, v):
        given_rows = set(v[0].values)
        mandatory_rows = Metadata.mandatory_fields()
        mandatory_rows_alias = Metadata.mandatory_fields(use_alias=True)

        if not (mandatory_rows.issubset(given_rows) or mandatory_rows_alias.issubset(given_rows)):
            missing_rows = mandatory_rows_alias.difference(given_rows)
            raise exceptions.MetadataSheetMissingMandatoryFields(missing_rows).to_pydantic_custom_error()
        return v

    @field_validator("Classes")
    def has_classes_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = Class.mandatory_fields()
        mandatory_columns_alias = Class.mandatory_fields(use_alias=True)

        if not (mandatory_columns.issubset(given_columns) or mandatory_columns_alias.issubset(given_columns)):
            missing_columns = mandatory_columns_alias.difference(given_columns)
            raise exceptions.ClassesSheetMissingMandatoryColumns(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Properties")
    def has_properties_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = Property.mandatory_fields()
        mandatory_columns_alias = Property.mandatory_fields(use_alias=True)

        if not (mandatory_columns.issubset(given_columns) or mandatory_columns_alias.issubset(given_columns)):
            missing_columns = mandatory_columns_alias.difference(given_columns)
            raise exceptions.PropertiesSheetMissingMandatoryColumns(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Prefixes")
    def has_prefixes_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = {"Prefix", "URI"}

        if not mandatory_columns.issubset(given_columns):
            missing_columns = mandatory_columns.difference(given_columns)
            raise exceptions.PrefixesSheetMissingMandatoryColumns(missing_columns).to_pydantic_custom_error()
        return v

    @field_validator("Instances")
    def has_instances_mandatory_columns(cls, v):
        given_columns = set(v.columns)
        mandatory_columns = {"Instance", "Property", "Value"}

        if not mandatory_columns.issubset(given_columns):
            missing_columns = mandatory_columns.difference(given_columns)
            raise exceptions.InstancesSheetMissingMandatoryColumns(missing_columns).to_pydantic_custom_error()
        return v

    @staticmethod
    def drop_non_string_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Drop non-string columns as this can cause issue when loading rules

        Args:
            df: data frame

        Returns:
            dataframe with removed non string columns
        """
        columns = [column for column in df.columns[df.columns.notna()] if isinstance(column, str)]

        return df[columns]


class Tables:
    prefixes = "Prefixes"
    properties = "Properties"
    classes = "Classes"
    metadata = "Metadata"
    instances = "Instances"


# readers:


def read_google_sheet_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if gspread is not installed
    local_import("gspread", "google")
    import gspread  # type: ignore[import]

    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def read_excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if openpyxl is not installed
    local_import("openpyxl", "excel")

    from openpyxl import Workbook, load_workbook

    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].source = filepath

    return sheets


def read_yaml_file_to_mapping_by_name(dirpath: Path, expected_files: set[str] | None = None) -> dict[str, dict]:
    # To trigger ImportError if yaml is not installed
    local_import("yaml", "yaml")
    from yaml import safe_load

    mapping_by_name = {}
    for filepath in dirpath.iterdir():
        if expected_files is not None and filepath.stem not in expected_files:
            continue
        mapping_by_name[filepath.stem] = safe_load(filepath.read_text())
    return mapping_by_name


def read_github_sheet_to_workbook(
    filepath: str, personal_token: str, owner: str, repo: str, branch: str = "main"
) -> Workbook:
    r = requests.get(
        f"https://api.github.com/repos/{owner}/{repo}/contents/{filepath}?ref={branch}",
        headers={"accept": "application/vnd.github.v3.raw", "authorization": f"token {personal_token}"},
    )

    loc = f"https://github.com/{owner}/{repo}/tree/{branch}"

    if r.status_code != 200:
        raise exceptions.UnableToDownloadExcelFile(filepath, loc, r.reason)
    try:
        wb = load_workbook(BytesIO(r.content), data_only=True)
    except BadZipFile as e:
        raise exceptions.NotExcelFile(filepath, loc) from e
    return wb


def read_github_sheet_to_table_by_name(
    filepath: str, personal_token: str, owner: str, repo: str, branch: str = "main"
) -> dict[str, pd.DataFrame]:
    wb = read_github_sheet_to_workbook(filepath, personal_token, owner, repo, branch)
    return workbook_to_table_by_name(wb)


def workbook_to_table_by_name(workbook: Workbook) -> dict[str, pd.DataFrame]:
    table = {}
    for sheet in workbook:
        sheetname = sheet.title
        data = sheet.values
        if sheetname == "Metadata":
            table[sheetname] = pd.DataFrame(data, columns=None)
        if sheetname == "Prefixes":
            columns = next(data)[:]
            table[sheetname] = pd.DataFrame(data, columns=columns).dropna(how="all")
        if sheetname in ["Classes", "Properties", "Instances"]:
            next(data)
            columns = next(data)[:]
            table[sheet.title] = pd.DataFrame(data, columns=columns).dropna(how="all")
    return table
