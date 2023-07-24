from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile
from openpyxl import Workbook, load_workbook
import pandas as pd
import requests

from cognite.neat.core.utils.auxiliary import local_import
from cognite.neat.core.rules import _exceptions


def google_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if gspread is not installed
    local_import("gspread", "google")
    import gspread

    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if openpyxl is not installed
    local_import("openpyxl", "excel")

    from openpyxl import Workbook, load_workbook

    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].source = filepath

    return sheets


def yaml_file_to_mapping_by_name(dirpath: Path, expected_files: set[str] | None = None) -> dict[str, dict]:
    # To trigger ImportError if yaml is not installed
    local_import("yaml", "yaml")
    from yaml import safe_load

    mapping_by_name = {}
    for filepath in dirpath.iterdir():
        if expected_files is not None and filepath.stem not in expected_files:
            continue
        mapping_by_name[filepath.stem] = safe_load(filepath.read_text())
    return mapping_by_name


def github_to_table_by_name(
    filepath: str, personal_token: str, owner: str, repo: str, branch: str = "main"
) -> dict[str, pd.DataFrame]:
    r = requests.get(
        f"https://api.github.com/repos/{owner}/{repo}/contents/{filepath}?ref={branch}",
        headers={"accept": "application/vnd.github.v3.raw", "authorization": f"token {personal_token}"},
    )

    loc = f"https://github.com/{owner}/{repo}/tree/{branch}"

    if r.status_code != 200:
        raise _exceptions.Error20(filepath, loc, r.reason)
    try:
        wb = load_workbook(BytesIO(r.content), data_only=True)
    except BadZipFile:
        raise _exceptions.Error21(filepath, loc)
    return _workbook_to_table_by_name(wb)


def _workbook_to_table_by_name(workbook: Workbook) -> dict[str, pd.DataFrame]:
    table = {}
    for sheet in workbook:
        sheetname = sheet.title
        data = sheet.values
        if sheetname == "Metadata":
            table[sheetname] = pd.DataFrame(data, columns=None)
        if sheetname in ["Classes", "Properties", "Instances"]:
            next(data)
            columns = next(data)[:]
            table[sheet.title] = pd.DataFrame(data, columns=columns).dropna(how="all")

    return table
