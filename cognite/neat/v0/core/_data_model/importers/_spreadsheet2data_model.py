"""This module performs importing of data model from spreadsheets."""

import tempfile
from collections import UserDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Literal, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import ExcelFile
from pydantic import ValidationError
from rdflib import Namespace, URIRef

from cognite.neat.v0.core._data_model._constants import SPLIT_ON_COMMA_PATTERN
from cognite.neat.v0.core._data_model._shared import (
    ImportedDataModel,
    T_UnverifiedDataModel,
)
from cognite.neat.v0.core._data_model.models import (
    UNVERIFIED_DATA_MODEL_BY_ROLE,
    VERIFIED_DATA_MODEL_BY_ROLE,
    RoleTypes,
    SchemaCompleteness,
)
from cognite.neat.v0.core._data_model.models._import_contexts import SpreadsheetContext
from cognite.neat.v0.core._data_model.models.entities._single_value import ContainerConstraintEntity, ContainerEntity
from cognite.neat.v0.core._issues import IssueList, MultiValueError
from cognite.neat.v0.core._issues.errors import (
    FileMissingRequiredFieldError,
    FileNotFoundNeatError,
    FileReadError,
)
from cognite.neat.v0.core._issues.warnings import FileMissingRequiredFieldWarning
from cognite.neat.v0.core._utils.spreadsheet import (
    SpreadsheetRead,
    find_column_and_row_with_value,
    read_individual_sheet,
)
from cognite.neat.v0.core._utils.text import humanize_collection

from ._base import BaseImporter

SOURCE_SHEET__TARGET_FIELD__HEADERS = [
    (
        "Properties",
        "Properties",
        {
            RoleTypes.information: ["Concept", "Property"],
            RoleTypes.dms: ["View", "View Property"],
        },
    ),
    ("Concepts", "Concepts", ["Concept"]),
    ("Containers", "Containers", ["Container"]),
    ("Views", "Views", ["View"]),
    ("Enum", "Enum", ["Collection"]),
    ("Nodes", "Nodes", ["Node"]),
]


MANDATORY_SHEETS_BY_ROLE: dict[RoleTypes, set[str]] = {
    role_type: {
        str(sheet_name)
        for sheet_name in (
            VERIFIED_DATA_MODEL_BY_ROLE.get(role_type).mandatory_fields(use_alias=True)  # type: ignore
            if VERIFIED_DATA_MODEL_BY_ROLE.get(role_type)
            else []
        )
        if sheet_name is not None
    }
    for role_type in RoleTypes.__members__.values()
}


class MetadataRaw(UserDict):
    @classmethod
    def from_excel(cls, excel_file: ExcelFile, metadata_sheet_name: str) -> "MetadataRaw":
        return cls(pd.read_excel(excel_file, metadata_sheet_name, header=None).replace(float("nan"), None).values)

    @property
    def has_role_field(self) -> bool:
        return self.get("role") in [role.value for role in RoleTypes.__members__.values()]

    @property
    def role(self) -> RoleTypes:
        return RoleTypes(self["role"])

    @property
    def has_schema_field(self) -> bool:
        return self.get("schema") in [schema.value for schema in SchemaCompleteness.__members__.values()]

    @property
    def schema(self) -> SchemaCompleteness | None:
        if not self.has_schema_field:
            return None
        return SchemaCompleteness(self["schema"])

    def is_valid(self, issue_list: IssueList, filepath: Path) -> bool:
        if not self.has_role_field:
            issue_list.append(FileMissingRequiredFieldError(filepath, "metadata", "role"))
            return False

        return True


@dataclass
class ReadResult:
    sheets: dict[str, dict | list]
    read_info_by_sheet: dict[str, SpreadsheetRead]
    metadata: MetadataRaw
    prefixes: dict[str, Namespace] | None = None

    @property
    def role(self) -> RoleTypes:
        return self.metadata.role

    @property
    def schema(self) -> SchemaCompleteness | None:
        return self.metadata.schema


class SpreadsheetReader:
    def __init__(
        self,
        issue_list: IssueList,
        required: bool = True,
        metadata: MetadataRaw | None = None,
        sheet_prefix: Literal["", "Last", "Ref", "CDMRef"] = "",
    ):
        self.issue_list = issue_list
        self.required = required
        self.metadata = metadata
        self._sheet_prefix = sheet_prefix
        self._seen_files: set[Path] = set()
        self._seen_sheets: set[str] = set()

    @property
    def metadata_sheet_name(self) -> str:
        return f"{self._sheet_prefix}Metadata"

    @property
    def prefixes_sheet_name(self) -> str:
        return "Prefixes"

    @property
    def seen_sheets(self) -> set[str]:
        if not self._seen_files:
            raise ValueError("No files have been read yet.")
        return self._seen_sheets

    def sheet_names(self, role: RoleTypes) -> set[str]:
        names = MANDATORY_SHEETS_BY_ROLE[role]
        return {f"{self._sheet_prefix}{sheet_name}" for sheet_name in names if sheet_name != "Metadata"}

    def read(self, excel_file: pd.ExcelFile, filepath: Path) -> None | ReadResult:
        self._seen_files.add(filepath)
        self._seen_sheets.update(map(str, excel_file.sheet_names))
        metadata: MetadataRaw | None
        if self.metadata is not None:
            metadata = self.metadata
        else:
            metadata = self._read_metadata(excel_file, filepath)
            if metadata is None:
                # The reading of metadata failed, so we can't continue
                return None

        sheets, read_info_by_sheet = self._read_sheets(excel_file, metadata.role)
        if sheets is None or self.issue_list.has_errors:
            return None
        sheets["Metadata"] = dict(metadata)

        # Special case for reading prefixes as they are suppose to be read only once
        if (
            self.prefixes_sheet_name in excel_file.sheet_names
            and not self._sheet_prefix
            and (prefixes := self._read_prefixes(excel_file, filepath))
        ):
            sheets["Prefixes"] = prefixes

        return ReadResult(sheets, read_info_by_sheet, metadata)

    def _read_metadata(self, excel_file: ExcelFile, filepath: Path) -> MetadataRaw | None:
        if self.metadata_sheet_name not in excel_file.sheet_names:
            if self.required:
                self.issue_list.append(FileMissingRequiredFieldError(filepath, "sheet", self.metadata_sheet_name))
            return None

        metadata = MetadataRaw.from_excel(excel_file, self.metadata_sheet_name)

        if not metadata.is_valid(self.issue_list, filepath):
            return None
        return metadata

    def _read_prefixes(self, excel_file: ExcelFile, filepath: Path) -> dict[str, Namespace] | None:
        if self.prefixes_sheet_name not in excel_file.sheet_names:
            return None

        else:
            prefixes = {}

            for row in read_individual_sheet(excel_file, "Prefixes", expected_headers=["Prefix", "Namespace"]):
                if "Prefix" in row and "Namespace" in row:
                    prefixes[row["Prefix"]] = row["Namespace"]
                else:
                    if "Prefix" not in row:
                        self.issue_list.append(FileMissingRequiredFieldWarning(filepath, "prefixes", "prefix"))
                    if "Namespace" not in row:
                        self.issue_list.append(FileMissingRequiredFieldWarning(filepath, "prefixes", "namespace"))
                    return None

            return prefixes

    def _read_sheets(
        self, excel_file: ExcelFile, read_role: RoleTypes
    ) -> tuple[dict[str, dict | list] | None, dict[str, SpreadsheetRead]]:
        read_info_by_sheet: dict[str, SpreadsheetRead] = defaultdict(SpreadsheetRead)

        sheets: dict[str, dict | list] = {}

        expected_sheet_names = self.sheet_names(read_role)

        if missing_sheets := expected_sheet_names.difference(set(excel_file.sheet_names)):
            if self.required:
                self.issue_list.append(
                    FileMissingRequiredFieldError(
                        cast(Path, excel_file.io), "sheets", humanize_collection(missing_sheets)
                    )
                )
            return None, read_info_by_sheet

        for source_sheet_name, target_sheet_name, headers_input in SOURCE_SHEET__TARGET_FIELD__HEADERS:
            source_sheet_name = f"{self._sheet_prefix}{source_sheet_name}"

            if source_sheet_name not in excel_file.sheet_names:
                continue
            if isinstance(headers_input, dict):
                headers = headers_input[read_role]
            else:
                headers = headers_input

            try:
                sheets[target_sheet_name], read_info_by_sheet[source_sheet_name] = read_individual_sheet(
                    excel_file,
                    source_sheet_name,
                    return_read_info=True,
                    expected_headers=headers,
                )
            except Exception as e:
                self.issue_list.append(FileReadError(cast(Path, excel_file.io), str(e)))
                continue

        return sheets, read_info_by_sheet


class ExcelImporter(BaseImporter[T_UnverifiedDataModel]):
    """Import data_model from an Excel file.

    Args:
        filepath (Path): The path to the Excel file.
    """

    def __init__(self, filepath: Path):
        self.filepath = filepath

    def to_data_model(self) -> ImportedDataModel[T_UnverifiedDataModel]:
        issue_list = IssueList(title=f"'{self.filepath.name}'")
        if not self.filepath.exists():
            raise FileNotFoundNeatError(self.filepath)

        self.filepath = self._make_forward_compatible_spreadsheet(self.filepath)

        with pd.ExcelFile(self.filepath) as excel_file:
            user_reader = SpreadsheetReader(issue_list)
            user_read = user_reader.read(excel_file, self.filepath)

        issue_list.trigger_warnings()
        if issue_list.has_errors:
            raise MultiValueError(issue_list.errors)

        if user_read is None:
            return ImportedDataModel(None, None)

        sheets = user_read.sheets
        original_role = user_read.role
        read_info_by_sheet = user_read.read_info_by_sheet

        data_model_cls = UNVERIFIED_DATA_MODEL_BY_ROLE[original_role]
        data_model = cast(T_UnverifiedDataModel, data_model_cls.load(sheets))

        # Delete the temporary file if it was created
        if "temp_neat_file" in self.filepath.name:
            try:
                self.filepath.unlink()
            except Exception as e:
                issue_list.append(FileReadError(self.filepath, f"Failed to delete temporary file: {e}"))

        return ImportedDataModel(data_model, SpreadsheetContext(read_info_by_sheet))

    @property
    def description(self) -> str:
        return f"Excel file {self.filepath.name} read as unverified data model"

    @property
    def source_uri(self) -> URIRef:
        return URIRef(f"file://{self.filepath.name}")

    def _make_forward_compatible_spreadsheet(self, filepath: Path) -> Path:
        """Makes the spreadsheet forward compatible by renaming legacy class with concept

        Args:
            filepath (Path): The path to the Excel file.

        """

        workbook = load_workbook(filepath, data_only=True)

        if "Classes" in workbook.sheetnames:
            print(
                (
                    "You are using a legacy spreadsheet format, "
                    "which we will support until v1.0 of neat."
                    " Please update your spreadsheet to the new format."
                ),
            )
            _replace_class_with_concept_cell(workbook["Classes"])
            sheet = workbook["Classes"]
            sheet.title = "Concepts"

            if "Properties" in workbook.sheetnames:
                _replace_class_with_concept_cell(workbook["Properties"])

        elif "Containers" in workbook.sheetnames:
            _replace_legacy_constraint_form(workbook["Containers"])
            _replace_legacy_constraint_form(workbook["Properties"])

        else:
            return filepath

        with tempfile.NamedTemporaryFile(prefix="temp_neat_file", suffix=".xlsx", delete=False) as temp_file:
            workbook.save(temp_file.name)
            return Path(temp_file.name)


def _replace_class_with_concept_cell(sheet: Worksheet) -> None:
    """
    Replaces the word "Class" with "Concept" in the first row of the given sheet.

    Args:
        sheet (Worksheet): The sheet in which to replace the word "Class".
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Class":
                cell.value = "Concept"


def _replace_legacy_constraint_form(sheet: Worksheet) -> None:
    """
    Replaces the legacy form of container constraints with the new form in the given sheet.

    Args:
        sheet (Worksheet): The sheet in which to replace the old form of container constraints.
    """
    column, row = find_column_and_row_with_value(sheet, "Constraint", False)

    if not column or not row:
        return None

    # Iterate over values in the constraint column and replace old form with new form
    replaced: bool = False
    for cell_row in sheet.iter_rows(min_row=row + 1, min_col=column, max_col=column):
        cell = cell_row[0]
        if cell.value is not None:  # Skip empty cells
            # Container sheet update
            if sheet.title.lower() == "containers":
                constraints = []
                for constraint in SPLIT_ON_COMMA_PATTERN.split(str(cell.value)):
                    # latest format, do nothing
                    if "require" in constraint.lower():
                        constraints.append(constraint)
                        continue

                    try:
                        container = ContainerEntity.load(constraint, space="default")
                        container_str = container.external_id if container.space == "default" else str(container)
                        constraints.append(
                            f"requires:{container.space}_{container.external_id}(require={container_str})"
                        )
                        replaced = True
                    except ValidationError:
                        constraints.append(constraint)

                cell.value = ",".join(constraints)

            # Properties sheet update
            elif sheet.title.lower() == "properties":
                constraints = []
                for constraint in SPLIT_ON_COMMA_PATTERN.split(str(cell.value)):
                    try:
                        constraint_entity = ContainerConstraintEntity.load(constraint)

                        if constraint_entity.prefix in ["uniqueness", "requires"]:
                            constraints.append(constraint)

                        # Replace old format with new format
                        else:
                            constraints.append(f"uniqueness:{constraint}")
                            replaced = True

                    # If the constraint is not valid, we keep it as is
                    # to be caught by validation later
                    except ValidationError:
                        constraints.append(constraint)

                cell.value = ",".join(constraints)

    if replaced:
        print(
            (
                "You are using a legacy container constraints format "
                f"in the {sheet.title} sheet. Please update to the latest format."
            ),
        )
