from __future__ import annotations

import itertools
from datetime import datetime
from pathlib import Path
from types import GenericAlias
from typing import Any, ClassVar, Literal, cast, get_args, overload

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from rdflib import Namespace

from cognite.neat.v0.core._constants import BASE_MODEL, get_base_concepts
from cognite.neat.v0.core._data_model._constants import get_internal_properties
from cognite.neat.v0.core._data_model._shared import VerifiedDataModel
from cognite.neat.v0.core._data_model.models import (
    SheetRow,
)
from cognite.neat.v0.core._data_model.models._base_verified import (
    BaseVerifiedMetadata,
    RoleTypes,
)
from cognite.neat.v0.core._data_model.models.conceptual._verified import (
    ConceptualDataModel,
)
from cognite.neat.v0.core._data_model.models.data_types import (
    _DATA_TYPE_BY_DMS_TYPE,
)
from cognite.neat.v0.core._data_model.models.physical._verified import PhysicalDataModel
from cognite.neat.v0.core._utils.spreadsheet import (
    find_column_and_row_with_value,
    generate_data_validation,
)

from ._base import BaseExporter

MAX_COLUMN_WIDTH = 70.0


class ExcelExporter(BaseExporter[VerifiedDataModel, Workbook]):
    """Export data_model to Excel.

    Args:
        styling: The styling to use for the Excel file. Defaults to "default". See below for details
            on the different styles.
        new_model_id: The new model ID to use for the exported spreadsheet. This is only applicable if the input
            data_model have 'is_reference' set. If provided, the model ID will be used to automatically create the
            new metadata sheet in the Excel file. The model id is expected to be a tuple of (prefix, title)
            (space, external_id) for Conceptual and Physical Data Model respectively.

        sheet_prefix: The prefix to use for the sheet names in the Excel file. Defaults to an empty string.

    The following styles are available:

    - "none":    No styling is applied.
    - "minimal": Column widths are adjusted to fit the content, and the header row(s) is frozen.
    - "default": Minimal + headers are bold, increased size, and colored.
    - "maximal": Default + alternating row colors in the properties sheet for each class in addition to extra
                 blank rows between classes and borders
    """

    Style = Literal["none", "minimal", "default", "maximal"]
    DumpOptions = Literal["user", "last", "reference"]
    _helper_sheet_name: str = "_helper"
    _main_header_by_sheet_name: ClassVar[dict[str, str]] = {
        "Properties": "Definition of Properties",
        "Concepts": "Definition of Concepts",
        "Views": "Definition of Views",
        "Containers": "Definition of Containers",
        "Nodes": "Definition of Nodes",
        "Enum": "Definition of Enum Collections",
    }
    _helper_sheet_column_indexes_by_names: ClassVar[dict[str, int]] = {
        "Concept": 1,
        "View": 1,
        "Implements": 2,
        "Value Type": 3,
        "Container": 4,
        "In Model": 5,
        "Immutable": 5,
        "Used For": 6,
    }
    style_options = get_args(Style)
    dump_options = get_args(DumpOptions)

    def __init__(
        self,
        styling: Style = "default",
        new_model_id: tuple[str, str] | None = None,
        sheet_prefix: str | None = None,
        reference_data_model_with_prefix: tuple[VerifiedDataModel, str] | None = None,
        add_empty_rows: bool = False,
        hide_internal_columns: bool = True,
        include_properties: Literal["same-space", "all"] = "all",
        add_drop_downs: bool = True,
        base_model: BASE_MODEL | None = None,
        total_concepts: int | None = None,
    ):
        self.sheet_prefix = sheet_prefix or ""
        if styling not in self.style_options:
            raise ValueError(f"Invalid styling: {styling}. Valid options are {self.style_options}")
        self.styling = styling
        self._styling_level = self.style_options.index(styling)
        self.new_model_id = new_model_id
        self.reference_data_model_with_prefix = reference_data_model_with_prefix
        self.add_empty_rows = add_empty_rows
        self.hide_internal_columns = hide_internal_columns
        self.include_properties = include_properties
        self.add_drop_downs = add_drop_downs
        self.base_model = base_model
        self.total_concepts = total_concepts

    @property
    def description(self) -> str:
        return "Export verified model to Excel."

    def export_to_file(self, data_model: VerifiedDataModel, filepath: Path) -> None:
        """Exports transformation data_model to excel file."""
        data = self.export(data_model)
        try:
            data.save(filepath)
        finally:
            data.close()
        return None

    @overload
    def template(self, role: RoleTypes, filepath: Path) -> None: ...

    @overload
    def template(self, role: RoleTypes, filepath: Path | None = None) -> None: ...

    def template(self, role: RoleTypes, filepath: Path | None = None) -> None | Workbook:
        """This method will create an spreadsheet template for data modeling depending on the role.

        Args:
            role: The role for which the template is created. Can be either "dms" or "information".
            filepath: The path to the file where the template will be saved.

        """
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        data_model_model = PhysicalDataModel if role == RoleTypes.dms else ConceptualDataModel

        headers_by_sheet = data_model_model.headers_by_sheet(by_alias=True)
        headers_by_sheet.pop("Metadata")

        self._write_metadata_sheet(
            workbook,
            cast(BaseVerifiedMetadata, data_model_model.model_fields["metadata"].annotation).default().model_dump(),
        )

        for sheet_name, headers in headers_by_sheet.items():
            if sheet_name in ("Metadata", "Prefixes", "Reference", "Last"):
                continue
            sheet = self._create_sheet_with_header(workbook, headers, sheet_name)
            self._style_sheet_header(sheet, headers)

        self._adjust_column_widths(workbook)
        self._hide_internal_columns(workbook)

        self._add_drop_downs(role, workbook)

        if filepath:
            try:
                workbook.save(filepath)
            finally:
                workbook.close()
            return None

        return workbook

    def export(self, data_model: VerifiedDataModel) -> Workbook:
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        dumped_user_data_model: dict[str, Any] = data_model.dump(by_alias=True)

        self._write_metadata_sheet(workbook, dumped_user_data_model["Metadata"], sheet_prefix=self.sheet_prefix)
        self._write_sheets(workbook, dumped_user_data_model, data_model, sheet_prefix=self.sheet_prefix)
        if self.reference_data_model_with_prefix:
            reference_data_model, prefix = self.reference_data_model_with_prefix
            dumped_reference_data_model = reference_data_model.dump(entities_exclude_defaults=False, by_alias=True)
            self._write_sheets(workbook, dumped_reference_data_model, reference_data_model, sheet_prefix=prefix)
            self._write_metadata_sheet(workbook, dumped_reference_data_model["Metadata"], sheet_prefix=prefix)

        if isinstance(data_model, ConceptualDataModel) and data_model.prefixes:
            self._write_prefixes_sheet(workbook, data_model.prefixes)

        if self._styling_level > 0:
            self._adjust_column_widths(workbook)

        if self.hide_internal_columns:
            self._hide_internal_columns(workbook)

        # Only add drop downs if the data_model are Physical Data Model
        if self.add_drop_downs:
            self._add_drop_downs(data_model.metadata.role, workbook)

        return workbook

    def _hide_internal_columns(self, workbook: Workbook) -> None:
        """Hides internal columns in workbook sheets.

        Args:
            workbook: Workbook representation of the Excel file.

        """
        for sheet in workbook.sheetnames:
            if sheet.lower() == "metadata":
                continue
            ws = workbook[sheet]
            for col in get_internal_properties():
                column_letter = find_column_and_row_with_value(ws, col)[0]
                if column_letter:
                    ws.column_dimensions[column_letter].hidden = True

    def _add_drop_downs(
        self,
        role: RoleTypes,
        workbook: Workbook,
    ) -> None:
        """Adds drop down menus to specific columns for fast and accurate data entry
        for both Conceptual and Physical Data Model.

        Args:
            role: The role for which the drop downs are created. Can be either "dms" or "information".
            workbook: Workbook representation of the Excel file.

        !!! note "Why defining individual data validation per desired column?
            This is due to the internal working of openpyxl. Adding same validation to
            different column leads to unexpected behavior when the openpyxl workbook is exported
            as and Excel file. Probably, the validation is not copied to the new column,
            but instead reference to the data validation object is added.

        !!! note "Why 100*100 rows?"
            This is due to the fact that we need to add drop down menus for all properties
            in the sheet. The maximum number of properties per view/concept is 100. So considering
            maximum number of views/concepts is 100, we need to add 100*100 rows.
        """

        self._make_helper_sheet(role, workbook, 100)

        data_validators: dict[str, DataValidation] = {}

        # We need create individual data validation and cannot re-use the same one due
        # the internals of openpyxl

        self._add_data_validation(
            workbook,
            sheet_name="Views" if role == RoleTypes.dms else "Concepts",
            column_name="Implements",
            data_validator_name="implements",
            data_validators=data_validators,
            validation_range=100 + 100,  # base + user concepts (max)
            total_rows=100,
        )

        self._add_data_validation(
            workbook,
            sheet_name="Properties",
            column_name="Value Type",
            data_validator_name="value_type",
            data_validators=data_validators,
            validation_range=150,  # primitive types + classes
            total_rows=100 * 100,  # 100 views/classes * 100 properties (max properties per view/class)
        )

        self._add_data_validation(
            workbook,
            sheet_name="Properties",
            column_name="View" if role == RoleTypes.dms else "Concept",
            data_validator_name="views_or_concepts",
            data_validators=data_validators,
            validation_range=100,
            total_rows=100 * 100,
        )

        if role == RoleTypes.dms:
            self._add_data_validation(
                workbook,
                sheet_name="Properties",
                column_name="Container",
                data_validator_name="container",
                data_validators=data_validators,
                validation_range=100,
                total_rows=100 * 100,
            )

            self._add_data_validation(
                workbook,
                sheet_name="Properties",
                column_name="Immutable",
                data_validator_name="immutable",
                data_validators=data_validators,
                validation_range=2,
                total_rows=100 * 100,
            )

            self._add_data_validation(
                workbook,
                sheet_name="Views",
                column_name="In Model",
                data_validator_name="in_model",
                data_validators=data_validators,
                validation_range=2,
                total_rows=100,  # 100 views
            )

            self._add_data_validation(
                workbook,
                sheet_name="Containers",
                column_name="Used For",
                data_validator_name="used_for",
                data_validators=data_validators,
                validation_range=3,
                total_rows=100,  # 100 views
            )

    def _make_helper_sheet(
        self,
        role: RoleTypes,
        workbook: Workbook,
        no_rows: int,
    ) -> None:
        """This helper sheet is used as source of data for drop down menus creation

        Args:
            role: The role for which the helper sheet is created. Can be either "dms" or "information".
            workbook: Workbook representation of the Excel file.
            no_rows: number of rows to add data too that will form base for drop down menus.
        """

        workbook.create_sheet(title=self._helper_sheet_name)

        value_type_counter = 0

        for value_type_counter, value_type in enumerate(_DATA_TYPE_BY_DMS_TYPE.values()):
            value_type_as_str = value_type.dms._type.casefold() if role == RoleTypes.dms else value_type.xsd
            # skip types which require special handling or are surpassed by CDM
            if value_type_as_str in ["enum", "timeseries", "sequence", "file"]:
                continue
            workbook[self._helper_sheet_name].cell(
                row=value_type_counter + 1,
                column=self._helper_sheet_column_indexes_by_names["Value Type"],
                value=value_type_as_str,
            )

        value_type_counter += 1

        concept_counter = 0
        if self.base_model and (concepts := get_base_concepts(self.base_model, self.total_concepts)):
            for concept_counter, concept in enumerate(concepts):
                workbook[self._helper_sheet_name].cell(
                    row=concept_counter + 1,
                    column=self._helper_sheet_column_indexes_by_names["Implements"],
                    value=concept,
                )
            concept_counter += 1

        views_or_concepts_sheet = "Views" if role == RoleTypes.dms else "Concepts"
        view_or_concept_column = "View" if role == RoleTypes.dms else "Concept"

        for i in range(no_rows):
            workbook[self._helper_sheet_name].cell(
                row=i + 1,
                column=self._helper_sheet_column_indexes_by_names[view_or_concept_column],
                value=f'=IF(ISBLANK({views_or_concepts_sheet}!A{i + 3}), "", {views_or_concepts_sheet}!A{i + 3})',
            )

            workbook[self._helper_sheet_name].cell(
                row=concept_counter + i + 1,
                column=self._helper_sheet_column_indexes_by_names["Implements"],
                value=f'=IF(ISBLANK({views_or_concepts_sheet}!A{i + 3}), "", {views_or_concepts_sheet}!A{i + 3})',
            )

            workbook[self._helper_sheet_name].cell(
                row=value_type_counter + i + 1,
                column=self._helper_sheet_column_indexes_by_names["Value Type"],
                value=f'=IF(ISBLANK({views_or_concepts_sheet}!A{i + 3}), "", {views_or_concepts_sheet}!A{i + 3})',
            )

            if role == RoleTypes.dms:
                workbook[self._helper_sheet_name].cell(
                    row=i + 1,
                    column=self._helper_sheet_column_indexes_by_names["Container"],
                    value=f'=IF(ISBLANK(Containers!A{i + 3}), "", Containers!A{i + 3})',
                )

        if role == RoleTypes.dms:
            for i, value in enumerate([True, False, ""]):
                workbook[self._helper_sheet_name].cell(
                    row=i + 1,
                    column=self._helper_sheet_column_indexes_by_names["In Model"],
                    value=cast(bool | str, value),
                )

            for i, value in enumerate(["node", "edge", "all"]):
                workbook[self._helper_sheet_name].cell(
                    row=i + 1,
                    column=self._helper_sheet_column_indexes_by_names["Used For"],
                    value=value,
                )

        workbook[self._helper_sheet_name].sheet_state = "hidden"

    def _add_data_validation(
        self,
        workbook: Workbook,
        sheet_name: str,
        column_name: str,
        data_validator_name: str,
        data_validators: dict,
        validation_range: int,
        total_rows: int,
    ) -> None:
        """Adds data validation to a column in a sheet.

        Args:
            workbook: Workbook representation of the Excel file.
            sheet_name: The name of the sheet to add the data validation to.
            column_name: The name of the column to add the data validation to.
            data_validator_name: The name of the data validation to add.
            data_validators: A dictionary to store the data validators.
            validation_range: The total number of validation values to add.
            total_rows: The number of rows to add the data validation to.

        !!! note "Why defining individual data validation per desired column?"
            This is due to the internal working of openpyxl. Adding same validation to
            different column leads to unexpected behavior when the openpyxl workbook is exported
            as and Excel file.

        !!! note "Why starting at row 3?"
            This is due to the header rows in the sheet. The first two rows are reserved for the header.
        """
        # CREATE VALIDATOR
        data_validators[data_validator_name] = generate_data_validation(
            self._helper_sheet_name,
            get_column_letter(self._helper_sheet_column_indexes_by_names[column_name]),
            total_header_rows=0,
            validation_range=validation_range,
        )

        # REGISTER VALIDATOR TO SPECIFIC WORKBOOK SHEET
        workbook[sheet_name].add_data_validation(data_validators[data_validator_name])

        # APPLY VALIDATOR TO SPECIFIC COLUMN
        if column_letter := find_column_and_row_with_value(workbook[sheet_name], column_name)[0]:
            data_validators[data_validator_name].add(f"{column_letter}{3}:{column_letter}{3 + total_rows}")

    def _create_sheet_with_header(
        self,
        workbook: Workbook,
        headers: list[str],
        sheet_name: str,
        sheet_prefix: str = "",
    ) -> Worksheet:
        """Creates an empty sheet with the given headers.

        Args:
            workbook: The workbook to add the sheet to.
            headers: The headers to add to the sheet.
            sheet_name: The name of the sheet.
            sheet_prefix: The prefix to add to the sheet name, if any.
        """

        sheet = workbook.create_sheet(f"{sheet_prefix}{sheet_name}")
        main_header = self._main_header_by_sheet_name[sheet_name]
        sheet.append([main_header] + [""] * (len(headers) - 1))

        if headers[0] == "Neat ID":
            # Move the Neat ID to the end of the columns
            headers = headers[1:] + ["Neat ID"]

        # Append the headers to the sheet
        sheet.append(headers)

        return sheet

    def _style_sheet_header(self, sheet: Worksheet, headers: list[str]) -> None:
        """Styles the sheet with the given headers.

        Args:
            sheet: The sheet to style.
            headers: The headers to style.
        """
        if self._styling_level > 0:
            # This freezes all rows above the given row
            sheet.freeze_panes = sheet["A3"]

            sheet["A1"].alignment = Alignment(horizontal="left")

        if self._styling_level > 1:
            # Make the header row bold, larger, and colored
            for cell, *_ in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                cell.font = Font(bold=True, size=20)
                cell.fill = PatternFill(fgColor="FFC000", patternType="solid")
            for cell in sheet["2"]:
                cell.font = Font(bold=True, size=14)

    def _write_sheets(
        self,
        workbook: Workbook,
        dumped_data_model: dict[str, Any],
        data_model: VerifiedDataModel,
        sheet_prefix: str = "",
    ) -> None:
        for sheet_name, headers in data_model.headers_by_sheet(by_alias=True).items():
            if sheet_name in ("Metadata", "Prefixes", "Reference", "Last"):
                continue

            sheet = self._create_sheet_with_header(workbook, headers, sheet_name, sheet_prefix)

            fill_colors = itertools.cycle(["CADCFC", "FFFFFF"])
            fill_color = next(fill_colors)
            last_concept: str | None = None
            item: dict[str, Any]
            for item in dumped_data_model.get(sheet_name) or []:
                if "Neat ID" in item:
                    # Move the Neat ID to the end of the columns
                    item["Neat ID"] = item.pop("Neat ID")
                row = list(item.values())
                concept = row[0]

                is_properties = sheet_name == "Properties"
                is_new_class = concept != last_concept and last_concept is not None
                if self._styling_level > 2 and is_new_class and is_properties:
                    if self.add_empty_rows:
                        sheet.append([""] * len(headers))
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                    fill_color = next(fill_colors)

                if is_properties and self.include_properties == "same-space":
                    space = concept.split(":")[0] if ":" in concept else data_model.metadata.space
                    if space != data_model.metadata.space:
                        continue

                sheet.append(row)
                if self._styling_level > 2 and is_properties:
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                last_concept = concept

            self._style_sheet_header(sheet, headers)

    def _write_metadata_sheet(self, workbook: Workbook, metadata: dict[str, Any], sheet_prefix: str = "") -> None:
        # Excel does not support timezone in datetime strings
        if isinstance(metadata.get("created"), datetime):
            metadata["created"] = metadata["created"].replace(tzinfo=None)
        if isinstance(metadata.get("updated"), datetime):
            metadata["updated"] = metadata["updated"].replace(tzinfo=None)

        metadata_sheet = workbook.create_sheet(f"{sheet_prefix}Metadata")
        for key, value in metadata.items():
            metadata_sheet.append([key, value])

        if self._styling_level > 1:
            for cell in metadata_sheet["A"]:
                cell.font = Font(bold=True, size=12)

    def _write_prefixes_sheet(self, workbook: Workbook, prefixes: dict[str, Namespace]) -> None:
        metadata_sheet = workbook.create_sheet("Prefixes")
        metadata_sheet.append(["Prefix", "Namespace"])
        for key, value in prefixes.items():
            metadata_sheet.append([key, value])

        if self._styling_level > 1:
            for cell in metadata_sheet["A"]:
                cell.font = Font(bold=True, size=12)

    @classmethod
    def _get_item_class(cls, annotation: GenericAlias) -> type[SheetRow]:
        if not isinstance(annotation, GenericAlias):
            raise ValueError(f"Expected annotation to be a GenericAlias, but got {type(annotation)}")
        args = get_args(annotation)
        if len(args) != 1:
            raise ValueError(f"Expected annotation to have exactly one argument, but got {len(args)}")
        arg = args[0]
        if not issubclass(arg, SheetRow):
            raise ValueError(f"Expected annotation to have a BaseModel argument, but got {type(arg)}")
        return arg

    @classmethod
    def _adjust_column_widths(cls, workbook: Workbook) -> None:
        for sheet_ in workbook:
            sheet = cast(Worksheet, sheet_)
            for column_cells in sheet.columns:
                try:
                    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                except ValueError:
                    max_length = 0

                selected_column = column_cells[0]
                if isinstance(selected_column, MergedCell):
                    selected_column = column_cells[1]

                current = sheet.column_dimensions[selected_column.column_letter].width or (max_length + 0.5)  # type: ignore[union-attr]
                sheet.column_dimensions[selected_column.column_letter].width = min(  # type: ignore[union-attr]
                    max(current, max_length + 0.5), MAX_COLUMN_WIDTH
                )
        return None
