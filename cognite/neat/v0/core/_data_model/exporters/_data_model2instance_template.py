import uuid
from pathlib import Path
from typing import Literal, cast

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cognite.neat.v0.core._data_model._constants import EntityTypes
from cognite.neat.v0.core._data_model.analysis import DataModelAnalysis
from cognite.neat.v0.core._data_model.models.conceptual._verified import (
    ConceptualDataModel,
)
from cognite.neat.v0.core._data_model.models.entities._single_value import ConceptEntity

from ._base import BaseExporter


class InstanceTemplateExporter(BaseExporter[ConceptualDataModel, Workbook]):
    """
    Converts conceptual data model to a templated spreadsheet meant for capturing
    instances based on concept definitions in the data model.

    Args:
        no_rows: number of rows for processing, by default 1000
        auto_identifier_type: type of automatic identifier, by default "index" based, alternative is "uuid" based
        add_drop_down_list: Add drop down selection for columns that contain linking properties, by default True

    !!! note "no_rows parameter"
        no_rows should be set to the maximum expected number of instances of any of the classes.
        By default, it is set to 1000, increase it accordingly if you have more instances.

    """

    def __init__(
        self,
        no_rows: int = 1000,
        auto_identifier_type: Literal["index", "uuid"] = "index",
        add_drop_down_list: bool = True,
    ):
        self.no_rows = no_rows
        self.auto_identifier_type = auto_identifier_type
        self.add_drop_down_list = add_drop_down_list

    @property
    def description(self) -> str:
        return "Export verified information instance template to Excel."

    def export(
        self,
        data_model: ConceptualDataModel,
    ) -> Workbook:
        workbook = Workbook()

        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        for concept, properties in DataModelAnalysis(data_model).properties_by_id_by_concept().items():
            workbook.create_sheet(title=concept.suffix)

            # Add header rows
            workbook[concept.suffix].append(["identifier", *list(properties.keys())])

            if self.auto_identifier_type == "uuid":
                _add_uuid_identifiers(workbook, concept.suffix, self.no_rows)
            else:
                # Default to index-based identifier
                _add_index_identifiers(workbook, concept.suffix, self.no_rows)

            for i, property_ in enumerate(properties.values()):
                if property_.type_ == EntityTypes.object_property and self.add_drop_down_list:
                    _add_drop_down_list(
                        workbook,
                        concept.suffix,
                        get_column_letter(i + 2),
                        self.no_rows,
                        cast(ConceptEntity, property_.value_type).suffix,
                        "A",
                    )

        _adjust_column_width(workbook)
        _set_header_style(workbook)

        return workbook

    def export_to_file(self, data_model: ConceptualDataModel, filepath: Path) -> None:
        """Exports graph capturing sheet to excel file."""
        data = self.export(data_model)
        try:
            data.save(filepath)
        finally:
            data.close()
        return None


def _add_index_identifiers(workbook: Workbook, sheet: str, no_rows: int) -> None:
    """Adds index-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        workbook[sheet][f"A{i + 2}"] = f'=IF(ISBLANK(B{i + 2}), "","{sheet}-{i + 1}")'


def _add_uuid_identifiers(workbook: Workbook, sheet: str, no_rows: int) -> None:
    """Adds UUID-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        workbook[sheet][f"A{i + 2}"] = f'=IF(ISBLANK(B{i + 2}), "","{sheet}-{uuid.uuid4()}")'


def _add_drop_down_list(
    workbook: Workbook,
    sheet: str,
    column: str,
    no_rows: int,
    value_sheet: str,
    value_column: str,
) -> None:
    """Adds a drop down list to a column"""
    drop_down_list = DataValidation(
        type="list",
        formula1=f"={value_sheet}!{value_column}$2:{value_column}${no_rows}",
    )

    workbook[sheet].add_data_validation(drop_down_list)

    for i in range(no_rows):
        drop_down_list.add(workbook[sheet][f"{column}{i + 2}"])


def _adjust_column_width(workbook: Workbook) -> None:
    """Adjusts the column width based on the content"""
    for sheet in workbook.sheetnames:
        for cell_tuple in workbook[sheet].columns:
            # Wrong type annotation in openpyxl
            cell = cast(Cell, cell_tuple[0])  # type: ignore[index]
            if cell.value:
                adjusted_width = (len(str(cell.value)) + 5) * 1.2
                workbook[sheet].column_dimensions[cell.column_letter].width = adjusted_width


def _set_header_style(workbook: Workbook) -> None:
    """Sets the header style for all sheets in the workbook"""
    style = NamedStyle(name="header style")
    style.font = Font(bold=True, size=16)
    side = Side(style="thin", color="000000")
    style.border = Border(left=side, right=side, top=side, bottom=side)
    workbook.add_named_style(style)

    for sheet in workbook.sheetnames:
        for cell_tuple in workbook[sheet].columns:
            # Wrong type annotation in openpyxl
            cell = cast(Cell, cell_tuple[0])  # type: ignore[index]
            workbook[sheet][f"{cell.column_letter}1"].style = style
            if f"{cell.column_letter}1" == "A1":
                workbook[sheet][f"{cell.column_letter}1"].fill = PatternFill("solid", start_color="2FB5F2")
            else:
                workbook[sheet][f"{cell.column_letter}1"].fill = PatternFill("solid", start_color="FFB202")
            workbook[sheet][f"{cell.column_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
