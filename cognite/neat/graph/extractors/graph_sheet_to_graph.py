import logging
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from rdflib import RDF, XSD, Literal, Namespace

from cognite.neat.graph import exceptions
from cognite.neat.graph.exceptions import NamespaceRequired
from cognite.neat.rules.analysis import get_defined_classes, to_class_property_pairs
from cognite.neat.rules.models import TransformationRules


def extract_graph_from_sheet(
    filepath: Path, transformation_rule: TransformationRules, separator: str = ",", namespace: str | None = None
) -> list[tuple]:
    """Converts a graph capturing sheet to RDF triples that define data model instances

    Args:
        filepath : Path to the graph capturing sheet
        transformation_rule : Transformation rules which holds data model that is used to validate
                              the graph capturing sheet and extract data model instances from it (i.e. RDF triples)
        separator : Multi value separator at cell level. Defaults to ",".
        namespace : Optional custom namespace to use for extracted triples that define data
                    model instances. Defaults to None.

    Returns:
        List of RDF triples, represented as tuples `(subject, predicate, object)`, that define data model instances
    """

    graph_capturing_sheet = read_graph_excel_file_to_table_by_name(filepath)

    return sheet2triples(graph_capturing_sheet, transformation_rule, separator, namespace)


def sheet2triples(
    graph_capturing_sheet: dict[str, pd.DataFrame],
    transformation_rule: TransformationRules,
    separator: str = ",",
    namespace: str | None = None,
) -> list[tuple]:
    """Converts a graph capturing sheet represented as dictionary of dataframes to rdf triples

    Args:
        graph_capturing_sheet : Graph capturing sheet provided as dictionary of dataframes
        transformation_rule : Transformation rules which holds data model that is used to validate
                             the graph capturing sheet and extract data model instances from it (i.e. RDF triples)
        separator : Multi value separator at cell level. Defaults to ",".
        namespace : Optional custom namespace to use for extracted triples that define
                    data model instances. Defaults to None.

    Returns:
        List of RDF triples, represented as tuples `(subject, predicate, object)`, that define data model instances
    """

    # Validation that everything is in order before proceeding
    validate_if_graph_capturing_sheet_empty(graph_capturing_sheet)
    validate_rules_graph_pair(graph_capturing_sheet, transformation_rule)

    # get class property pairs
    class_property_pairs = to_class_property_pairs(transformation_rule)

    # namespace selection
    if namespace is None and transformation_rule.metadata.namespace is not None:
        instance_namespace = transformation_rule.metadata.namespace
    elif namespace:
        instance_namespace = Namespace(namespace)
    else:
        raise NamespaceRequired("Extact instances from sheet")

    if transformation_rule.metadata.namespace is not None:
        model_namespace = Namespace(transformation_rule.metadata.namespace)
    else:
        raise NamespaceRequired("Extact instances from sheet")

    # Now create empty graph
    triples: list[tuple] = []

    # Add triples from the capturing sheet to the graph by iterating over the capturing sheet
    # iterate over sheets
    for sheet_name, df in graph_capturing_sheet.items():
        # iterate over sheet rows
        for _, row in df.iterrows():
            if row.identifier is None:
                msg = f"Missing identifier in sheet {sheet_name} at row {row.name}! Skipping..."
                logging.warning(msg)
                warnings.warn(msg, stacklevel=2)
                continue

            # iterate over sheet rows properties
            for property_name, value in row.to_dict().items():
                # Setting RDF type of the instance
                if property_name == "identifier":
                    triples.append((instance_namespace[row.identifier], RDF.type, model_namespace[sheet_name]))
                    continue
                elif not value:
                    continue

                property_ = class_property_pairs[sheet_name][property_name]
                is_one_to_many = (property_.max_count or 1) > 1 and separator
                values = value.split(separator) if is_one_to_many else [value]

                # Adding object properties
                if property_.property_type == "ObjectProperty":
                    triples.extend(
                        (
                            instance_namespace[row.identifier],
                            model_namespace[property_name],
                            instance_namespace[v.strip()],
                        )
                        for v in values
                    )
                # Adding data properties
                elif property_.property_type == "DatatypeProperty":
                    triples.extend(
                        (
                            instance_namespace[row.identifier],
                            model_namespace[property_name],
                            Literal(v.strip(), datatype=XSD[property_.expected_value_type]),
                        )
                        for v in values
                    )
                else:
                    raise exceptions.UnsupportedPropertyType(property_.property_type)
    return triples


def validate_if_graph_capturing_sheet_empty(graph_capturing_sheet: dict[str, pd.DataFrame]):
    """Validate if the graph capturing sheet is empty

    Args:
        graph_capturing_sheet : Graph capturing sheet
    """
    if all(df.empty for df in graph_capturing_sheet.values()):
        msg = "Graph capturing sheet is empty! Aborting!"
        logging.error(msg)
        raise ValueError(msg)


def validate_rules_graph_pair(graph_capturing_sheet: dict[str, pd.DataFrame], transformation_rule: TransformationRules):
    """Validate if the graph capturing sheet is based on the transformation rules

    Args:
        graph_capturing_sheet : Graph capturing sheet
        transformation_rule : Transformation rules
    """
    intersection = set(graph_capturing_sheet.keys()).intersection(set(get_defined_classes(transformation_rule)))

    if not intersection:
        msg = "Graph capturing sheet is not based on transformation rules! Aborting!"
        logging.error(msg)
        raise ValueError(msg)

    elif len(intersection) == len(graph_capturing_sheet.keys()):
        logging.info("All classes in the graph capturing sheet are defined in the transformation rules!")

    elif len(intersection) < len(graph_capturing_sheet.keys()):
        msg = "Graph capturing sheet contains classes that are not defined in the transformation rules! Proceeding..."
        logging.warning(msg)
        warnings.warn(msg, stacklevel=2)

    elif len(intersection) < len(get_defined_classes(transformation_rule)):
        msg = "Transformation rules contain classes that are not present in the graph capturing sheet! Proceeding..."
        logging.warning(msg)
        warnings.warn(msg, stacklevel=2)


def read_graph_excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    workbook: Workbook = load_workbook(filepath)

    parsed_sheets = {
        sheetname: pd.read_excel(filepath, sheet_name=sheetname, header=0) for sheetname in workbook.sheetnames
    }

    for sheetname, df in parsed_sheets.items():
        if "identifier" in df.columns:
            parsed_sheets[sheetname] = df.drop(df[df.identifier == 0].index)
            parsed_sheets[sheetname] = df.replace({np.nan: None})
        else:
            logging.error(f"Sheet {sheetname} does not have an identifier column")
            raise ValueError(f"Sheet {sheetname} does not have an identifier column")

    return parsed_sheets
