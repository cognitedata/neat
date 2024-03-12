import logging
import uuid
import warnings
from pathlib import Path
from typing import cast

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rdflib import RDF, XSD, Literal, Namespace, URIRef

from cognite.neat.graph import exceptions
from cognite.neat.graph.exceptions import NamespaceRequired
from cognite.neat.graph.models import Triple
from cognite.neat.rules.analysis import get_defined_classes, to_class_property_pairs
from cognite.neat.rules.exporter._rules2rules import to_dms_name
from cognite.neat.rules.models.rules import Rules

from ._base import BaseExtractor


class GraphCapturingSheet(BaseExtractor):
    """
    Graph capturing sheet class that provides methods for creating a graph capturing sheet and extracting RDF triples.

    Args:
        rules: Transformation rules which holds data model that is used to validate
                the graph capturing sheet and extract data model instances from it (i.e. RDF triples)
        filepath: File path to save the sheet to. Defaults to None.
        separator: Multi value separator at cell level. Defaults to ",".
        namespace: Optional custom namespace to use for extracted triples that define data
                    model instances. Defaults to None, meaning namespace of rules will be used.
        store_graph_capturing_sheet: Whether to store the graph capturing sheet in the object. Will be stored in the
                                     `sheet` attribute. Defaults to False.
        use_source_ids : Whether to use source ids for properties and classes stored in Source column if they exist.
                         Defaults to False, meaning that the source ids will be ignored.

    """

    def __init__(
        self,
        rules: Rules,
        filepath: Path | str | None = None,
        separator: str = ",",
        namespace: str | None = None,
        store_graph_capturing_sheet: bool = False,
        use_source_ids: bool = False,
    ):
        self.rules = rules
        self.filepath = Path(filepath) if isinstance(filepath, str | Path) else None
        self.separator = separator
        self.namespace = namespace
        self.store_graph_capturing_sheet = store_graph_capturing_sheet
        self.use_source_ids = use_source_ids
        self.sheet: dict[str, pd.DataFrame] = {}

    def create_template(self, filepath: Path | None = None, overwrite: bool = False) -> None:
        """
        Creates a graph capturing sheet template based on the transformation rules.

        Args:
            filepath: File path to save the sheet to. Defaults to None.
            overwrite: Overwrite existing file. Defaults to False.
        """
        if filepath is None:
            filepath = self.filepath
        if filepath is None:
            raise ValueError("File path to the graph capturing sheet is not provided!")
        if filepath.exists() and not overwrite:
            raise FileExistsError(f"File {filepath} already exists! Set overwrite to True to overwrite it!")
        rules2graph_capturing_sheet(self.rules, filepath)

    def extract(self) -> list[Triple]:
        """
        Extracts RDF triples from the graph capturing sheet.

        Returns:
            List of RDF triples, represented as tuples `(subject, predicate, object)`, that define data model instances
        """
        if self.filepath is None:
            raise ValueError("File path to the graph capturing sheet is not provided!")
        graph_capturing_sheet = read_graph_excel_file_to_table_by_name(self.filepath)
        if self.store_graph_capturing_sheet:
            self.sheet = graph_capturing_sheet

        print(self.namespace)

        return sheet2triples(graph_capturing_sheet, self.rules, self.separator, self.namespace, self.use_source_ids)


def extract_graph_from_sheet(
    filepath: Path, transformation_rule: Rules, separator: str = ",", namespace: str | None = None
) -> list[Triple]:
    """Converts a graph capturing sheet to RDF triples that define data model instances

    Args:
        filepath : Path to the graph capturing sheet
        transformation_rule : Transformation rules which holds data model that is used to validate
                              the graph capturing sheet and extract data model instances from it (i.e. RDF triples)
        separator : Multi value separator at cell level. Defaults to ",".
        namespace : Optional custom namespace to use for extracted triples that define data
                    model instances. Defaults to None, meaning namespace of rules will be used.

    Returns:
        List of RDF triples, represented as tuples `(subject, predicate, object)`, that define data model instances
    """

    graph_capturing_sheet = read_graph_excel_file_to_table_by_name(filepath)

    return sheet2triples(graph_capturing_sheet, transformation_rule, separator, namespace)


def sheet2triples(
    graph_capturing_sheet: dict[str, pd.DataFrame],
    rules: Rules,
    separator: str = ",",
    namespace: str | None = None,
    use_source_ids: bool = False,
) -> list[Triple]:
    """Converts a graph capturing sheet represented as dictionary of dataframes to rdf triples

    Args:
        graph_capturing_sheet : Graph capturing sheet provided as dictionary of dataframes
        transformation_rule : Transformation rules which holds data model that is used to validate
                             the graph capturing sheet and extract data model instances from it (i.e. RDF triples)
        separator : Multi value separator at cell level. Defaults to ",".
        namespace : Optional custom namespace to use for extracted triples that define
                    data model instances. Defaults to None, meaning namespace of rules will be used.
        use_source_ids : Whether to use source ids for properties and classes stored in Source column if they exist.
                         Defaults to False, meaning that the source ids will be ignored.

    Returns:
        List of RDF triples, represented as tuples `(subject, predicate, object)`, that define data model instances
    """

    # Validation that everything is in order before proceeding
    validate_if_graph_capturing_sheet_empty(graph_capturing_sheet)
    validate_rules_graph_pair(graph_capturing_sheet, rules)

    # get class property pairs
    class_property_pairs = to_class_property_pairs(rules)

    # namespace selection
    if namespace is None and rules.metadata.namespace is not None:
        instance_namespace = rules.metadata.namespace
    elif namespace:
        instance_namespace = Namespace(namespace)
    else:
        raise NamespaceRequired("Extract instances from sheet")

    if rules.metadata.namespace is not None:
        model_namespace = Namespace(rules.metadata.namespace)
    else:
        raise NamespaceRequired("Extract instances from sheet")

    # Now create empty graph
    triples: list[Triple] = []

    # Add triples from the capturing sheet to the graph by iterating over the capturing sheet
    # iterate over sheets
    for sheet_name, df in graph_capturing_sheet.items():
        # iterate over sheet rows

        class_uri = (
            URIRef(str(rules.classes[sheet_name].source))
            if use_source_ids and rules.classes[sheet_name].source
            else model_namespace[sheet_name]
        )

        for _, row in df.iterrows():
            if row.identifier is None:
                msg = f"Missing identifier in sheet {sheet_name} at row {row.name}! Skipping..."
                logging.warning(msg)
                warnings.warn(msg, stacklevel=2)
                continue

            # iterate over sheet rows properties
            for property_name, value in row.to_dict().items():
                # Setting RDF type of the instance
                if property_name == "identifier":
                    triples.append((instance_namespace[row.identifier], RDF.type, class_uri))
                    continue
                elif not value:
                    continue

                property_uri = (
                    URIRef(str(class_property_pairs[sheet_name][property_name].source))
                    if use_source_ids and class_property_pairs[sheet_name][property_name].source
                    else model_namespace[property_name]
                )

                property_ = class_property_pairs[sheet_name][property_name]

                is_one_to_many = separator and (
                    (property_.max_count and property_.max_count > 1) or not property_.max_count
                )

                values = value.split(separator) if is_one_to_many else [value]

                # Adding object properties
                if property_.property_type == "ObjectProperty":
                    triples.extend(
                        (
                            instance_namespace[row.identifier],
                            property_uri,
                            instance_namespace[v.strip()],
                        )
                        for v in values
                    )
                # Adding data properties
                elif property_.property_type == "DatatypeProperty":
                    for v in values:
                        try:
                            literal_value = v.strip()
                        except AttributeError:
                            literal_value = v

                        triples.append(
                            (
                                instance_namespace[row.identifier],
                                property_uri,
                                Literal(literal_value, datatype=XSD[property_.expected_value_type.suffix]),
                            )
                        )

                else:
                    raise exceptions.UnsupportedPropertyType(property_.property_type)
    return triples


def validate_if_graph_capturing_sheet_empty(graph_capturing_sheet: dict[str, pd.DataFrame]):
    """Validate if the graph capturing sheet is empty

    Args:
        graph_capturing_sheet : Graph capturing sheet
    """
    if all(df.empty for df in graph_capturing_sheet.values()):
        msg = "Graph capturing sheet is empty! Aborting!"
        logging.error(msg)
        raise ValueError(msg)


def validate_rules_graph_pair(graph_capturing_sheet: dict[str, pd.DataFrame], transformation_rule: Rules):
    """Validate if the graph capturing sheet is based on the transformation rules

    Args:
        graph_capturing_sheet : Graph capturing sheet
        transformation_rule : Transformation rules
    """
    intersection = set(graph_capturing_sheet.keys()).intersection(set(get_defined_classes(transformation_rule)))

    if not intersection:
        msg = "Graph capturing sheet is not based on transformation rules! Aborting!"
        logging.error(msg)
        raise ValueError(msg)

    elif len(intersection) == len(graph_capturing_sheet.keys()):
        logging.info("All classes in the graph capturing sheet are defined in the transformation rules!")

    elif len(intersection) < len(graph_capturing_sheet.keys()):
        msg = "Graph capturing sheet contains classes that are not defined in the transformation rules! Proceeding..."
        logging.warning(msg)
        warnings.warn(msg, stacklevel=2)

    elif len(intersection) < len(get_defined_classes(transformation_rule)):
        msg = "Transformation rules contain classes that are not present in the graph capturing sheet! Proceeding..."
        logging.warning(msg)
        warnings.warn(msg, stacklevel=2)


def read_graph_excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    workbook: Workbook = load_workbook(filepath)

    parsed_sheets = {
        sheetname: pd.read_excel(filepath, sheet_name=sheetname, header=0) for sheetname in workbook.sheetnames
    }

    for sheetname, df in parsed_sheets.items():
        if "identifier" in df.columns:
            parsed_sheets[sheetname] = df.drop(df[df.identifier == 0].index)
            parsed_sheets[sheetname] = df.replace({np.nan: None})
        else:
            logging.error(f"Sheet {sheetname} does not have an identifier column")
            raise ValueError(f"Sheet {sheetname} does not have an identifier column")

    return parsed_sheets


def rules2graph_capturing_sheet(
    rules: Rules,
    file_path: Path,
    no_rows: int = 1000,
    auto_identifier_type: str = "index-based",
    add_drop_down_list: bool = True,
):
    """
    Converts a TransformationRules object to a graph capturing sheet

    Args:
        rules: The TransformationRules object to convert to the graph capturing sheet
        file_path: File path to save the sheet to
        no_rows: Number of rows for processing, by default 1000
        auto_identifier_type: Type of automatic identifier, by default "index" based, alternative is "uuid" based
        add_drop_down_list: Add drop down selection for columns that contain linking properties, by default True

    !!! note "no_rows parameter"
        no_rows should be set to the maximum expected number of instances of any of the classes.
        By default, it is set to 1000, increase it accordingly if you have more instances.

    """

    workbook = Workbook()
    # Remove default sheet named "Sheet"
    workbook.remove(workbook["Sheet"])

    for class_, properties in to_class_property_pairs(rules).items():
        workbook.create_sheet(title=class_)

        # Add header rows
        workbook[class_].append(["identifier", *list(properties.keys())])

        if auto_identifier_type and auto_identifier_type == "index-based":  # default, easy to read
            logging.debug(f"Configuring index-based automatic identifiers for sheet {class_}")
            _add_index_identifiers(workbook, class_, no_rows)
        elif auto_identifier_type and auto_identifier_type == "uuid-based":
            logging.debug(f"Configuring UUID-based automatic identifiers for sheet {class_}")
            _add_uuid_identifiers(workbook, class_, no_rows)
        else:
            logging.debug(f"No automatic identifier set for sheet {class_}!")

        for i, property_ in enumerate(properties.values()):
            if property_.property_type == "ObjectProperty" and add_drop_down_list:
                _add_drop_down_list(
                    workbook, class_, get_column_letter(i + 2), no_rows, property_.expected_value_type.suffix, "A"
                )

    _adjust_column_width(workbook)
    _set_header_style(workbook)

    logging.info(f"Graph capturing sheet generated and stored at {file_path}!")
    workbook.save(file_path)
    workbook.close()


def _add_index_identifiers(workbook: Workbook, sheet: str, no_rows: int):
    """Adds index-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        prefix = to_dms_name(sheet, "class", True)
        workbook[sheet][f"A{i+2}"] = f'=IF(ISBLANK(B{i+2}), "","{prefix}-{i+1}")'


def _add_uuid_identifiers(workbook: Workbook, sheet: str, no_rows: int):
    """Adds UUID-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        prefix = to_dms_name(sheet, "class", True)
        workbook[sheet][f"A{i+2}"] = f'=IF(ISBLANK(B{i+2}), "","{prefix}-{uuid.uuid4()}")'


def _add_drop_down_list(workbook: Workbook, sheet: str, column: str, no_rows: int, value_sheet: str, value_column: str):
    """Adds a drop down list to a column"""
    drop_down_list = DataValidation(type="list", formula1=f"={value_sheet}!{value_column}$2:{value_column}${no_rows}")

    workbook[sheet].add_data_validation(drop_down_list)

    for i in range(no_rows):
        drop_down_list.add(workbook[sheet][f"{column}{i+2}"])


def _adjust_column_width(workbook: Workbook):
    """Adjusts the column width based on the content"""
    for sheet in workbook.sheetnames:
        for cell_tuple in workbook[sheet].columns:
            # Wrong type annotation in openpyxl
            cell = cast(Cell, cell_tuple[0])  # type: ignore[index]
            if cell.value:
                adjusted_width = (len(str(cell.value)) + 5) * 1.2
                workbook[sheet].column_dimensions[cell.column_letter].width = adjusted_width


def _set_header_style(workbook: Workbook):
    """Sets the header style for all sheets in the workbook"""
    style = NamedStyle(name="header style")
    style.font = Font(bold=True, size=16)
    side = Side(style="thin", color="000000")
    style.border = Border(left=side, right=side, top=side, bottom=side)
    workbook.add_named_style(style)

    for sheet in workbook.sheetnames:
        for cell_tuple in workbook[sheet].columns:
            # Wrong type annotation in openpyxl
            cell = cast(Cell, cell_tuple[0])  # type: ignore[index]
            workbook[sheet][f"{cell.column_letter}1"].style = style
            if f"{cell.column_letter}1" == "A1":
                workbook[sheet][f"{cell.column_letter}1"].fill = PatternFill("solid", start_color="2FB5F2")
            else:
                workbook[sheet][f"{cell.column_letter}1"].fill = PatternFill("solid", start_color="FFB202")
            workbook[sheet][f"{cell.column_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
