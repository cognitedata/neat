from __future__ import annotations

import itertools
from datetime import datetime
from pathlib import Path
from types import GenericAlias
from typing import Any, ClassVar, Literal, cast, get_args

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from rdflib import Namespace

from cognite.neat._constants import COGNITE_CONCEPTS
from cognite.neat._rules._constants import get_internal_properties
from cognite.neat._rules._shared import VerifiedRules
from cognite.neat._rules.models import (
    SheetRow,
)
from cognite.neat._rules.models._base_rules import BaseMetadata, RoleTypes
from cognite.neat._rules.models.data_types import (
    _DATA_TYPE_BY_DMS_TYPE,
)
from cognite.neat._rules.models.dms._rules import DMSRules
from cognite.neat._rules.models.information._rules import InformationRules
from cognite.neat._utils.spreadsheet import (
    find_column_with_value,
    generate_data_validation,
)

from ._base import BaseExporter

MAX_COLUMN_WIDTH = 70.0


class ExcelExporter(BaseExporter[VerifiedRules, Workbook]):
    """Export rules to Excel.

    Args:
        styling: The styling to use for the Excel file. Defaults to "default". See below for details
            on the different styles.
        new_model_id: The new model ID to use for the exported spreadsheet. This is only applicable if the input
            rules have 'is_reference' set. If provided, the model ID will be used to automatically create the
            new metadata sheet in the Excel file. The model id is expected to be a tuple of (prefix, title)
            (space, external_id) for InformationRules and DMSRules respectively.

        sheet_prefix: The prefix to use for the sheet names in the Excel file. Defaults to an empty string.

    The following styles are available:

    - "none":    No styling is applied.
    - "minimal": Column widths are adjusted to fit the content, and the header row(s) is frozen.
    - "default": Minimal + headers are bold, increased size, and colored.
    - "maximal": Default + alternating row colors in the properties sheet for each class in addition to extra
                 blank rows between classes and borders
    """

    Style = Literal["none", "minimal", "default", "maximal"]
    DumpOptions = Literal["user", "last", "reference"]
    _helper_sheet_name: str = "_helper"
    _main_header_by_sheet_name: ClassVar[dict[str, str]] = {
        "Properties": "Definition of Properties",
        "Classes": "Definition of Classes",
        "Views": "Definition of Views",
        "Containers": "Definition of Containers",
        "Nodes": "Definition of Nodes",
        "Enum": "Definition of Enum Collections",
    }
    style_options = get_args(Style)
    dump_options = get_args(DumpOptions)

    def __init__(
        self,
        styling: Style = "default",
        new_model_id: tuple[str, str] | None = None,
        sheet_prefix: str | None = None,
        reference_rules_with_prefix: tuple[VerifiedRules, str] | None = None,
        add_empty_rows: bool = False,
        hide_internal_columns: bool = True,
        include_properties: Literal["same-space", "all"] = "all",
        add_drop_downs: bool = True,
    ):
        self.sheet_prefix = sheet_prefix or ""
        if styling not in self.style_options:
            raise ValueError(f"Invalid styling: {styling}. Valid options are {self.style_options}")
        self.styling = styling
        self._styling_level = self.style_options.index(styling)
        self.new_model_id = new_model_id
        self.reference_rules_with_prefix = reference_rules_with_prefix
        self.add_empty_rows = add_empty_rows
        self.hide_internal_columns = hide_internal_columns
        self.include_properties = include_properties
        self.add_drop_downs = add_drop_downs

    @property
    def description(self) -> str:
        return "Export verified model to Excel."

    def export_to_file(self, rules: VerifiedRules, filepath: Path) -> None:
        """Exports transformation rules to excel file."""
        data = self.export(rules)
        try:
            data.save(filepath)
        finally:
            data.close()
        return None

    def template(self, role: RoleTypes, filepath: Path | None = None) -> None | Workbook:
        """This method will create an spreadsheet template for data modeling depending on the role.

        Args:
            role: The role for which the template is created. Can be either "dms" or "information".
            filepath: The path to the file where the template will be saved.

        """
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        rules_model = DMSRules if role == RoleTypes.dms else InformationRules

        headers_by_sheet = rules_model.headers_by_sheet(by_alias=True)
        headers_by_sheet.pop("Metadata")

        self._write_metadata_sheet(
            workbook,
            cast(BaseMetadata, rules_model.model_fields["metadata"].annotation).default().model_dump(),
        )

        for sheet_name, headers in headers_by_sheet.items():
            if sheet_name in ("Metadata", "Prefixes", "Reference", "Last"):
                continue
            sheet = self._create_sheet_with_header(workbook, headers, sheet_name)
            self._style_sheet_header(sheet, headers)

        self._adjust_column_widths(workbook)
        self._hide_internal_columns(workbook)

        if role == RoleTypes.dms:
            self._add_dms_drop_downs(workbook)
        else:
            self._add_info_drop_downs(workbook)

        if filepath:
            try:
                workbook.save(filepath)
            finally:
                workbook.close()
            return None

        return workbook

    def export(self, rules: VerifiedRules) -> Workbook:
        workbook = Workbook()
        # Remove default sheet named "Sheet"
        workbook.remove(workbook["Sheet"])

        dumped_user_rules: dict[str, Any] = rules.dump(by_alias=True)

        self._write_metadata_sheet(workbook, dumped_user_rules["Metadata"], sheet_prefix=self.sheet_prefix)
        self._write_sheets(workbook, dumped_user_rules, rules, sheet_prefix=self.sheet_prefix)
        if self.reference_rules_with_prefix:
            reference_rules, prefix = self.reference_rules_with_prefix
            dumped_reference_rules = reference_rules.dump(entities_exclude_defaults=False, by_alias=True)
            self._write_sheets(workbook, dumped_reference_rules, reference_rules, sheet_prefix=prefix)
            self._write_metadata_sheet(workbook, dumped_reference_rules["Metadata"], sheet_prefix=prefix)

        if isinstance(rules, InformationRules) and rules.prefixes:
            self._write_prefixes_sheet(workbook, rules.prefixes)

        if self._styling_level > 0:
            self._adjust_column_widths(workbook)

        if self.hide_internal_columns:
            self._hide_internal_columns(workbook)

        # Only add drop downs if the rules are DMSRules
        if self.add_drop_downs and isinstance(rules, DMSRules):
            self._add_dms_drop_downs(workbook)
        elif self.add_drop_downs and isinstance(rules, InformationRules):
            self._add_info_drop_downs(workbook)

        return workbook

    def _hide_internal_columns(self, workbook: Workbook) -> None:
        """Hides internal columns in workbook sheets.

        Args:
            workbook: Workbook representation of the Excel file.

        """
        for sheet in workbook.sheetnames:
            if sheet.lower() == "metadata":
                continue
            ws = workbook[sheet]
            for col in get_internal_properties():
                column_letter = find_column_with_value(ws, col)
                if column_letter:
                    ws.column_dimensions[column_letter].hidden = True

    def _add_info_drop_downs(self, workbook: Workbook, no_rows: int = 100) -> None:
        """Adds drop down menus to specific columns for fast and accurate data entry
        in the Information rules.

        Args:
            workbook: Workbook representation of the Excel file.
            no_rows: number of rows to add drop down menus. Defaults to 100*100.

        !!! note "Why no_rows=100?"
            Maximum number of views per data model is 100, thus this value is set accordingly

        !!! note "Why defining individual data validation per desired column?
            This is due to the internal working of openpyxl. Adding same validation to
            different column leads to unexpected behavior when the openpyxl workbook is exported
            as and Excel file. Probably, the validation is not copied to the new column,
            but instead reference to the data validation object is added.
        """
        self._make_helper_info_sheet(workbook, no_rows)

        # We need create individual data validation and cannot re-use the same one due
        # the internals of openpyxl
        dv_classes = generate_data_validation(self._helper_sheet_name, "A", no_header_rows=0, no_rows=no_rows)
        dv_value_types = generate_data_validation(self._helper_sheet_name, "B", no_header_rows=0, no_rows=no_rows)
        dv_implements = generate_data_validation(
            self._helper_sheet_name,
            "C",
            no_header_rows=0,
            no_rows=no_rows + len(COGNITE_CONCEPTS),
        )

        workbook["Properties"].add_data_validation(dv_classes)
        workbook["Properties"].add_data_validation(dv_value_types)
        workbook["Classes"].add_data_validation(dv_implements)

        # we multiply no_rows with 100 since a view can have max 100 properties per view
        if column := find_column_with_value(workbook["Properties"], "Class"):
            dv_classes.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Properties"], "Value Type"):
            dv_value_types.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Classes"], "Implements"):
            dv_implements.add(f"{column}{3}:{column}{no_rows}")

    def _make_helper_info_sheet(self, workbook: Workbook, no_rows: int) -> None:
        """This helper Information sheet is used as source of data for drop down menus creation"""
        workbook.create_sheet(title=self._helper_sheet_name)

        for dtype_counter, dtype in enumerate(_DATA_TYPE_BY_DMS_TYPE.values()):
            # skip types which require special handling or are surpassed by CDM
            if dtype.xsd in ["enum", "timeseries", "sequence", "file", "json"]:
                continue
            workbook[self._helper_sheet_name].cell(row=dtype_counter + 1, column=2, value=dtype.xsd)

        # Add Cognite Core Data Views:
        for concept_counter, concept in enumerate(COGNITE_CONCEPTS):
            workbook[self._helper_sheet_name].cell(
                row=concept_counter + 1,
                column=3,
                value=f"cdf_cdm:{concept}(version=v1)",
            )

        for i in range(no_rows):
            workbook[self._helper_sheet_name].cell(
                row=i + 1,
                column=1,
                value=f'=IF(ISBLANK(Classes!A{i + 3}), "", Classes!A{i + 3})',
            )
            workbook[self._helper_sheet_name].cell(
                row=dtype_counter + i + 2,
                column=2,
                value=f'=IF(ISBLANK(Classes!A{i + 3}), "", Classes!A{i + 3})',
            )
            workbook[self._helper_sheet_name].cell(
                row=concept_counter + i + 2,
                column=3,
                value=f'=IF(ISBLANK(Classes!A{i + 3}), "", Classes!A{i + 3})',
            )

        workbook[self._helper_sheet_name].sheet_state = "hidden"

    def _add_dms_drop_downs(self, workbook: Workbook, no_rows: int = 100) -> None:
        """Adds drop down menus to specific columns for fast and accurate data entry
        in the DMS rules.

        Args:
            workbook: Workbook representation of the Excel file.
            no_rows: number of rows to add drop down menus. Defaults to 100*100.

        !!! note "Why no_rows=100?"
            Maximum number of views per data model is 100, thus this value is set accordingly

        !!! note "Why defining individual data validation per desired column?
            This is due to the internal working of openpyxl. Adding same validation to
            different column leads to unexpected behavior when the openpyxl workbook is exported
            as and Excel file. Probably, the validation is not copied to the new column,
            but instead reference to the data validation object is added.
        """

        self._make_helper_dms_sheet(workbook, no_rows)

        # We need create individual data validation and cannot re-use the same one due
        # the internals of openpyxl
        dv_views = generate_data_validation(self._helper_sheet_name, "A", no_header_rows=0, no_rows=no_rows)
        dv_containers = generate_data_validation(self._helper_sheet_name, "B", no_header_rows=0, no_rows=no_rows)
        dv_value_types = generate_data_validation(self._helper_sheet_name, "C", no_header_rows=0, no_rows=no_rows)
        dv_implements = generate_data_validation(
            self._helper_sheet_name,
            "F",
            no_header_rows=0,
            no_rows=no_rows + len(COGNITE_CONCEPTS),
        )

        dv_immutable = generate_data_validation(self._helper_sheet_name, "D", no_header_rows=0, no_rows=3)
        dv_in_model = generate_data_validation(self._helper_sheet_name, "D", no_header_rows=0, no_rows=3)
        dv_used_for = generate_data_validation(self._helper_sheet_name, "E", no_header_rows=0, no_rows=3)

        workbook["Properties"].add_data_validation(dv_views)
        workbook["Properties"].add_data_validation(dv_containers)
        workbook["Properties"].add_data_validation(dv_value_types)
        workbook["Properties"].add_data_validation(dv_immutable)
        workbook["Views"].add_data_validation(dv_in_model)
        workbook["Views"].add_data_validation(dv_implements)
        workbook["Containers"].add_data_validation(dv_used_for)

        # we multiply no_rows with 100 since a view can have max 100 properties per view
        if column := find_column_with_value(workbook["Properties"], "View"):
            dv_views.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Properties"], "Container"):
            dv_containers.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Properties"], "Value Type"):
            dv_value_types.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Properties"], "Immutable"):
            dv_immutable.add(f"{column}{3}:{column}{no_rows * 100}")

        if column := find_column_with_value(workbook["Views"], "In Model"):
            dv_in_model.add(f"{column}{3}:{column}{no_rows}")

        if column := find_column_with_value(workbook["Views"], "Implements"):
            dv_implements.add(f"{column}{3}:{column}{no_rows}")

        if column := find_column_with_value(workbook["Containers"], "Used For"):
            dv_used_for.add(f"{column}{3}:{column}{no_rows}")

    def _make_helper_dms_sheet(self, workbook: Workbook, no_rows: int) -> None:
        """This helper DMS sheet is used as source of data for drop down menus creation"""
        workbook.create_sheet(title=self._helper_sheet_name)

        for dtype_counter, dtype in enumerate(_DATA_TYPE_BY_DMS_TYPE):
            if dtype in ["enum", "timeseries", "sequence", "file", "json"]:
                continue
            workbook[self._helper_sheet_name].cell(row=dtype_counter + 1, column=3, value=dtype)

        # Add Cognite Core Data Views:
        for concept_counter, concept in enumerate(COGNITE_CONCEPTS):
            workbook[self._helper_sheet_name].cell(
                row=concept_counter + 1,
                column=6,
                value=f"cdf_cdm:{concept}(version=v1)",
            )

        for i in range(no_rows):
            workbook[self._helper_sheet_name].cell(
                row=i + 1,
                column=1,
                value=f'=IF(ISBLANK(Views!A{i + 3}), "", Views!A{i + 3})',
            )
            workbook[self._helper_sheet_name].cell(
                row=i + 1,
                column=2,
                value=f'=IF(ISBLANK(Containers!A{i + 3}), "", Containers!A{i + 3})',
            )
            workbook[self._helper_sheet_name].cell(
                row=dtype_counter + i + 2,
                column=3,
                value=f'=IF(ISBLANK(Views!A{i + 3}), "", Views!A{i + 3})',
            )
            workbook[self._helper_sheet_name].cell(
                row=concept_counter + i + 2,
                column=6,
                value=f'=IF(ISBLANK(Views!A{i + 3}), "", Views!A{i + 3})',
            )

        for i, value in enumerate([True, False, ""]):
            workbook[self._helper_sheet_name].cell(row=i + 1, column=4, value=cast(bool | str, value))

        for i, value in enumerate(["node", "edge", "all"]):
            workbook[self._helper_sheet_name].cell(row=i + 1, column=5, value=value)

        workbook[self._helper_sheet_name].sheet_state = "hidden"

    def _create_sheet_with_header(
        self,
        workbook: Workbook,
        headers: list[str],
        sheet_name: str,
        sheet_prefix: str = "",
    ) -> Worksheet:
        """Creates an empty sheet with the given headers.

        Args:
            workbook: The workbook to add the sheet to.
            headers: The headers to add to the sheet.
            sheet_name: The name of the sheet.
            sheet_prefix: The prefix to add to the sheet name, if any.
        """

        sheet = workbook.create_sheet(f"{sheet_prefix}{sheet_name}")
        main_header = self._main_header_by_sheet_name[sheet_name]
        sheet.append([main_header] + [""] * (len(headers) - 1))

        if headers[0] == "Neat ID":
            # Move the Neat ID to the end of the columns
            headers = headers[1:] + ["Neat ID"]

        # Append the headers to the sheet
        sheet.append(headers)

        return sheet

    def _style_sheet_header(self, sheet: Worksheet, headers: list[str]) -> None:
        """Styles the sheet with the given headers.

        Args:
            sheet: The sheet to style.
            headers: The headers to style.
        """
        if self._styling_level > 0:
            # This freezes all rows above the given row
            sheet.freeze_panes = sheet["A3"]

            sheet["A1"].alignment = Alignment(horizontal="left")

        if self._styling_level > 1:
            # Make the header row bold, larger, and colored
            for cell, *_ in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                cell.font = Font(bold=True, size=20)
                cell.fill = PatternFill(fgColor="FFC000", patternType="solid")
            for cell in sheet["2"]:
                cell.font = Font(bold=True, size=14)

    def _write_sheets(
        self,
        workbook: Workbook,
        dumped_rules: dict[str, Any],
        rules: VerifiedRules,
        sheet_prefix: str = "",
    ) -> None:
        for sheet_name, headers in rules.headers_by_sheet(by_alias=True).items():
            if sheet_name in ("Metadata", "Prefixes", "Reference", "Last"):
                continue

            sheet = self._create_sheet_with_header(workbook, headers, sheet_name, sheet_prefix)

            fill_colors = itertools.cycle(["CADCFC", "FFFFFF"])
            fill_color = next(fill_colors)
            last_class: str | None = None
            item: dict[str, Any]
            for item in dumped_rules.get(sheet_name) or []:
                if "Neat ID" in item:
                    # Move the Neat ID to the end of the columns
                    item["Neat ID"] = item.pop("Neat ID")
                row = list(item.values())
                class_ = row[0]

                is_properties = sheet_name == "Properties"
                is_new_class = class_ != last_class and last_class is not None
                if self._styling_level > 2 and is_new_class and is_properties:
                    if self.add_empty_rows:
                        sheet.append([""] * len(headers))
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                    fill_color = next(fill_colors)

                if is_properties and self.include_properties == "same-space":
                    space = class_.split(":")[0] if ":" in class_ else rules.metadata.space
                    if space != rules.metadata.space:
                        continue

                sheet.append(row)
                if self._styling_level > 2 and is_properties:
                    for cell in sheet[sheet.max_row]:
                        cell.fill = PatternFill(fgColor=fill_color, patternType="solid")
                        side = Side(style="thin", color="000000")
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                last_class = class_

            self._style_sheet_header(sheet, headers)

    def _write_metadata_sheet(self, workbook: Workbook, metadata: dict[str, Any], sheet_prefix: str = "") -> None:
        # Excel does not support timezone in datetime strings
        if isinstance(metadata.get("created"), datetime):
            metadata["created"] = metadata["created"].replace(tzinfo=None)
        if isinstance(metadata.get("updated"), datetime):
            metadata["updated"] = metadata["updated"].replace(tzinfo=None)

        metadata_sheet = workbook.create_sheet(f"{sheet_prefix}Metadata")
        for key, value in metadata.items():
            metadata_sheet.append([key, value])

        if self._styling_level > 1:
            for cell in metadata_sheet["A"]:
                cell.font = Font(bold=True, size=12)

    def _write_prefixes_sheet(self, workbook: Workbook, prefixes: dict[str, Namespace]) -> None:
        metadata_sheet = workbook.create_sheet("Prefixes")
        metadata_sheet.append(["Prefix", "Namespace"])
        for key, value in prefixes.items():
            metadata_sheet.append([key, value])

        if self._styling_level > 1:
            for cell in metadata_sheet["A"]:
                cell.font = Font(bold=True, size=12)

    @classmethod
    def _get_item_class(cls, annotation: GenericAlias) -> type[SheetRow]:
        if not isinstance(annotation, GenericAlias):
            raise ValueError(f"Expected annotation to be a GenericAlias, but got {type(annotation)}")
        args = get_args(annotation)
        if len(args) != 1:
            raise ValueError(f"Expected annotation to have exactly one argument, but got {len(args)}")
        arg = args[0]
        if not issubclass(arg, SheetRow):
            raise ValueError(f"Expected annotation to have a BaseModel argument, but got {type(arg)}")
        return arg

    @classmethod
    def _adjust_column_widths(cls, workbook: Workbook) -> None:
        for sheet_ in workbook:
            sheet = cast(Worksheet, sheet_)
            for column_cells in sheet.columns:
                try:
                    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                except ValueError:
                    max_length = 0

                selected_column = column_cells[0]
                if isinstance(selected_column, MergedCell):
                    selected_column = column_cells[1]

                current = sheet.column_dimensions[selected_column.column_letter].width or (max_length + 0.5)  # type: ignore[union-attr]
                sheet.column_dimensions[selected_column.column_letter].width = min(  # type: ignore[union-attr]
                    max(current, max_length + 0.5), MAX_COLUMN_WIDTH
                )
        return None
