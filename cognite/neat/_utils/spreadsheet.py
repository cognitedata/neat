import warnings
from dataclasses import dataclass, field
from typing import Any, Literal, cast, overload
from warnings import catch_warnings

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from cognite.neat._rules._constants import get_internal_properties


@dataclass
class SpreadsheetRead:
    """This class is used to store information about the source spreadsheet.

    It is used to adjust row numbers to account for header rows and empty rows
    such that the error/warning messages are accurate.
    """

    header_row: int = 1
    empty_rows: list[int] = field(default_factory=list)
    skipped_rows: list[int] = field(default_factory=list)
    is_one_indexed: bool = True

    def __post_init__(self):
        self.empty_rows = sorted(self.empty_rows)

    def adjusted_row_number(self, row_no: int) -> int:
        output = row_no
        for empty_row in self.empty_rows:
            if empty_row <= output:
                output += 1
            else:
                break

        for skipped_rows in self.skipped_rows:
            if skipped_rows <= output:
                output += 1
            else:
                break

        return output + self.header_row + (1 if self.is_one_indexed else 0)


@overload
def read_individual_sheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: Literal[True],
    expected_headers: list[str] | None = None,
) -> tuple[list[dict], SpreadsheetRead]: ...


@overload
def read_individual_sheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: Literal[False] = False,
    expected_headers: list[str] | None = None,
) -> list[dict]: ...


def read_individual_sheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: bool = False,
    expected_headers: list[str] | None = None,
) -> tuple[list[dict], SpreadsheetRead] | list[dict]:
    if expected_headers:
        with catch_warnings():
            # When reading spreadsheets produced by neat, they contain dropdowns. These
            # are not supported by openpyxl and will raise a warning as openpyxl cannot validate these.
            # We ignore these warnings as Neat will do the same checks.
            warnings.simplefilter("ignore")
            target_row = _get_row_number(cast(Worksheet, load_workbook(excel_file)[sheet_name]), expected_headers)
        skiprows = target_row - 1 if target_row is not None else 0
    else:
        skiprows = 0

    with catch_warnings():
        # When reading spreadsheets produced by neat, they contain dropdowns. These
        # are not supported by openpyxl and will raise a warning as openpyxl cannot validate these.
        # We ignore these warnings as Neat will do the same checks.
        warnings.simplefilter("ignore")
        raw = pd.read_excel(excel_file, sheet_name, skiprows=skiprows)
    is_na = raw.isnull().all(axis=1)
    skip_rows = _find_rows_to_skip(raw)
    empty_rows = is_na[is_na].index.tolist()

    if skip_rows:
        raw = raw.drop(skip_rows)

    raw.dropna(axis=0, how="all", inplace=True)

    if "Value Type" in raw.columns:
        # Special handling for Value Type column, #N/A is treated specially by NEAT it means Unknown
        raw["Value Type"] = raw["Value Type"].replace(float("nan"), "#N/A")

    output = raw.replace(float("nan"), None).to_dict(orient="records")
    if return_read_info:
        # If no rows are skipped, row 1 is the header row.
        return output, SpreadsheetRead(
            header_row=skiprows + 1,
            empty_rows=empty_rows,
            is_one_indexed=True,
            skipped_rows=skip_rows,
        )
    return output


def _find_rows_to_skip(
    df: pd.DataFrame,
) -> list:
    """Find rows which are having all values as None except for internal properties."""
    rows_to_skip = []

    internal_cols = {val.lower() for val in get_internal_properties()}
    for i, row in df.iterrows():
        user_cols_state = []
        internal_cols_state = []
        for col in df.columns:
            if col.lower() not in internal_cols:
                user_cols_state.append(row[col] == "#N/A" or row[col].__str__().lower() in ["none", "nan"])
            else:
                internal_cols_state.append(row[col] is not None)

        if all(user_cols_state) and any(internal_cols_state):
            rows_to_skip.append(i)

    return rows_to_skip


def _get_row_number(sheet: Worksheet, values_to_find: list[str]) -> int | None:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(value in row for value in values_to_find):
            return row_number
    return None


def find_column_with_value(sheet: Worksheet, value: Any) -> str | None:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.lower() == value.lower():
                return cell.column_letter  # type: ignore

    return None


def generate_data_validation(sheet: str, column: str, no_header_rows: int, no_rows: int) -> DataValidation:
    "Creates openpyxl data validation object for a cell in a sheet"

    return DataValidation(
        type="list",
        formula1=f"={sheet}!{column}${no_header_rows + 1}:{column}${no_rows}",
    )
