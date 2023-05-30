import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

__all__ = [
    "excel_file_to_table_by_name",
]


def excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    workbook: Workbook = load_workbook(filepath)

    parsed_sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=0,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname, df in parsed_sheets.items():
        if "identifier" in df.columns:
            parsed_sheets[sheetname] = df.drop(df[df.identifier == 0].index)
            parsed_sheets[sheetname] = df.replace({np.nan: None})
        else:
            logging.error(f"Sheet {sheetname} does not have an identifier column")
            raise ValueError(f"Sheet {sheetname} does not have an identifier column")

    return parsed_sheets
