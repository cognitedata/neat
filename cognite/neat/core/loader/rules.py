from pathlib import Path

import gspread
import pandas as pd
from openpyxl import Workbook, load_workbook

__all__ = [
    "google_to_table_by_name",
    "excel_file_to_table_by_name",
]


def google_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].filepath = filepath

    return sheets
