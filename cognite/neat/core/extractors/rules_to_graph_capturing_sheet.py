import logging
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cognite.neat.core.data_classes.transformation_rules import TransformationRules
from cognite.neat.core.extractors.rules_to_graphql import repair_name as to_graphql_name


def _add_index_identifiers(workbook: Workbook, sheet: str, no_rows: int):
    """Adds index-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        prefix = to_graphql_name(sheet, "class", True)
        workbook[sheet][f"A{i+2}"] = f'=IF(ISBLANK(B{i+2}), "","{prefix}-{i+1}")'


def _add_uuid_identifiers(workbook: Workbook, sheet: str, no_rows: int):
    """Adds UUID-based auto identifier to a sheet identifier column"""
    for i in range(no_rows):
        prefix = to_graphql_name(sheet, "class", True)
        workbook[sheet][f"A{i+2}"] = f'=IF(ISBLANK(B{i+2}), "","{prefix}-{uuid.uuid4()}")'


def _add_drop_down_list(workbook: Workbook, sheet: str, column: str, no_rows: int, value_sheet: str, value_column: str):
    """Adds a drop down list to a column"""
    drop_down_list = DataValidation(type="list", formula1=f"={value_sheet}!{value_column}2:{value_column}{no_rows}")

    workbook[sheet].add_data_validation(drop_down_list)

    for i in range(no_rows):
        drop_down_list.add(workbook[sheet][f"{column}{i+2}"])


def _adjust_column_width(workbook: Workbook):
    """Adjusts the column width based on the content"""
    for sheet in workbook.sheetnames:
        for column in workbook[sheet].columns:
            if column[0].value:
                adjusted_width = (len(str(column[0].value)) + 5) * 1.2
                workbook[sheet].column_dimensions[column[0].column_letter].width = adjusted_width


def _set_header_style(workbook: Workbook):
    """Sets the header style for all sheets in the workbook"""
    style = NamedStyle(name="header style")
    style.font = Font(bold=True, size=16)
    side = Side(style="thin", color="000000")
    style.border = Border(left=side, right=side, top=side, bottom=side)
    workbook.add_named_style(style)

    for sheet in workbook.sheetnames:
        for column in workbook[sheet].columns:
            workbook[sheet][f"{column[0].column_letter}1"].style = style
            if f"{column[0].column_letter}1" == "A1":
                workbook[sheet][f"{column[0].column_letter}1"].fill = PatternFill("solid", start_color="2FB5F2")
            else:
                workbook[sheet][f"{column[0].column_letter}1"].fill = PatternFill("solid", start_color="FFB202")
            workbook[sheet][f"{column[0].column_letter}1"].alignment = Alignment(horizontal="center", vertical="center")


def rules2graph_capturing_sheet(
    transformation_rules: TransformationRules,
    file_path: Path,
    no_rows: int = 1000,
    auto_identifier_type: str = "index",
    add_drop_down_list: bool = True,
):
    """Converts a TransformationRules object to a graph capturing sheet

    Parameters
    ----------
    transformation_rules : TransformationRules
        The TransformationRules object to convert to the graph capturing sheet
    file_path : Path
        File path to save the sheet to
    auto_identifier_type : str, optional
        Type of automatic identifier, by default "index" based
    no_rows : int, optional
        Number of rows for processing, by default 1000
    add_drop_down_list : bool, optional
        Add drop down selection for columns that contain linking properties, by default True

    Notes
    -----
    no_rows should be set to the maximum expected number of instances of any of the classes.
    By default, it is set to 1000, increase it accordingly if you have more instances.

    """

    workbook = Workbook()

    # Remove default sheet named "Sheet"
    workbook.remove(workbook["Sheet"])

    for class_, properties in transformation_rules.get_class_property_pairs().items():
        workbook.create_sheet(title=class_)

        # Add header rows
        workbook[class_].append(["identifier"] + list(properties.keys()))

        if auto_identifier_type == "index":  # default, easy to read
            logging.debug("Configuring index-based automatic identifiers")
            _add_index_identifiers(workbook, class_, no_rows)
        elif auto_identifier_type == "uuid":
            logging.debug("Configuring UUID-based automatic identifiers")
            _add_uuid_identifiers(workbook, class_, no_rows)
        else:
            logging.debug(f"{auto_identifier_type} based auto identifier not supported!")

        for i, property_ in enumerate(properties.values()):
            if property_.property_type == "ObjectProperty" and add_drop_down_list:
                _add_drop_down_list(
                    workbook,
                    class_,
                    get_column_letter(i + 2),
                    no_rows,
                    property_.expected_value_type,
                    "A",
                )

    _adjust_column_width(workbook)
    _set_header_style(workbook)

    logging.info(f"Graph capturing sheet generated and stored at {file_path}!")
    workbook.save(file_path)
    workbook.close()
