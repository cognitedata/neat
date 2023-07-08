from pathlib import Path

import pandas as pd

from cognite.neat.core.utils.auxiliary import local_import


def google_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if gspread is not installed
    local_import("gspread", "google")
    import gspread

    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    # To trigger ImportError if openpyxl is not installed
    local_import("openpyxl", "excel")

    from openpyxl import Workbook, load_workbook

    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].source = filepath

    return sheets


def yaml_file_to_mapping_by_name(dirpath: Path, expected_files: set[str] | None = None) -> dict[str, dict]:
    # To trigger ImportError if yaml is not installed
    local_import("yaml", "yaml")
    from yaml import safe_load

    mapping_by_name = {}
    for filepath in dirpath.iterdir():
        if expected_files is not None and filepath.stem not in expected_files:
            continue
        mapping_by_name[filepath.stem] = safe_load(filepath.read_text())
    return mapping_by_name
