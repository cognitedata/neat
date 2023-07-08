from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

__all__ = [
    "google_to_table_by_name",
    "excel_file_to_table_by_name",
]

from cognite.neat.core.utils.auxiliary import local_import


def google_to_table_by_name(sheet_id: str) -> dict[str, pd.DataFrame]:
    gspread = local_import("gspread", "google")

    client_google = gspread.service_account()
    spreadsheet = client_google.open_by_key(sheet_id)
    return {worksheet.title: pd.DataFrame(worksheet.get_all_records()) for worksheet in spreadsheet.worksheets()}


def excel_file_to_table_by_name(filepath: Path) -> dict[str, pd.DataFrame]:
    workbook: Workbook = load_workbook(filepath)

    sheets = {
        sheetname: pd.read_excel(
            filepath,
            sheet_name=sheetname,
            header=None if sheetname == "Metadata" else 0,
            skiprows=1 if sheetname in ["Classes", "Properties", "Instances"] else None,
        )
        for sheetname in workbook.sheetnames
    }

    for sheetname in sheets:
        sheets[sheetname].source = filepath

    return sheets
