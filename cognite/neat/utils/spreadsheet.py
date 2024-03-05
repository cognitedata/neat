from dataclasses import dataclass, field
from typing import Literal, overload

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SpreadsheetRead:
    """This class is used to store information about the source spreadsheet.

    It is used to adjust row numbers to account for header rows and empty rows
    such that the error/warning messages are accurate.
    """

    header_row: int = 0
    empty_rows: list[int] = field(default_factory=list)
    is_one_indexed: bool = True

    def __post_init__(self):
        self.empty_rows = sorted(self.empty_rows)

    def adjusted_row_number(self, row_no: int) -> int:
        output = row_no + self.header_row + (1 if self.is_one_indexed else 0)
        for empty_row in self.empty_rows:
            if empty_row <= output:
                output += 1
            else:
                break
        return output


@overload
def read_spreadsheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: Literal[True],
    expected_headers: list[str] | None = None,
) -> tuple[list[dict], SpreadsheetRead]:
    ...


@overload
def read_spreadsheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: Literal[False] = False,
    expected_headers: list[str] | None = None,
) -> list[dict]:
    ...


def read_spreadsheet(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    return_read_info: bool = False,
    expected_headers: list[str] | None = None,
) -> tuple[list[dict], SpreadsheetRead] | list[dict]:
    if expected_headers:
        target_row = _get_row_number(load_workbook(excel_file)[sheet_name], expected_headers)
        skiprows = target_row - 1 if target_row is not None else 0
    else:
        skiprows = 0

    raw = pd.read_excel(excel_file, sheet_name, skiprows=skiprows)
    is_na = raw.isnull().all(axis=1)
    empty_rows = is_na[is_na].index.tolist()

    output = raw.dropna(axis=0, how="all").replace(float("nan"), None).to_dict(orient="records")
    if return_read_info:
        return output, SpreadsheetRead(header_row=skiprows + 1, empty_rows=empty_rows, is_one_indexed=True)
    return output


def _get_row_number(sheet: Worksheet, values_to_find: list[str]) -> int | None:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(value in row for value in values_to_find):
            return row_number
    return None
