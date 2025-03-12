from pathlib import Path

from cognite.client import CogniteClient
from openpyxl import Workbook, load_workbook

from cognite.neat import NeatSession


class TestCreateAPI:
    def test_create_enterprise_maintain_reference(self, cognite_client: CogniteClient, tmp_path: Path) -> None:
        neat = NeatSession(cognite_client)
        neat.read.examples.pump_example()

        neat.template.enterprise_model(("my_space", "MySpace", "v1"))

        result_path = tmp_path / "result.xlsx"
        neat.to.excel(result_path, include_reference=True, include_properties="same-space")

        workbook: Workbook | None = None
        try:
            workbook = load_workbook(result_path)
            actual_sheets = set(workbook.sheetnames)
        finally:
            workbook.close()
        expected_sheets = {"Metadata", "Views", "Properties", "Containers", "Enum", "Nodes"}
        expected_sheets |= {f"Ref{name}" for name in expected_sheets}

        missing = expected_sheets - actual_sheets
        assert not missing, f"Missing sheets: {missing}"
