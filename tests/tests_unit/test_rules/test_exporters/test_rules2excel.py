from pathlib import Path

import pytest
from lxml import etree
from openpyxl import Workbook, load_workbook

from cognite.neat.v0.core._data_model.exporters import ExcelExporter
from cognite.neat.v0.core._data_model.models import (
    ConceptualDataModel,
    PhysicalDataModel,
)
from cognite.neat.v0.core._data_model.models._base_verified import RoleTypes
from cognite.neat.v0.core._issues.errors._general import NeatValueError


def compare_cells(expected: Workbook, resulted: Workbook) -> list[bool]:
    skip_rows = [header for header in ExcelExporter()._main_header_by_sheet_name.values()] + ["created", "updated"]
    comparison_results = []

    for sheet_name in expected.sheetnames:
        for row1, row2 in zip(
            expected[sheet_name].iter_rows(values_only=True),
            resulted[sheet_name].iter_rows(values_only=True),
            strict=False,
        ):
            if row1[0] in skip_rows:
                continue
            row1 = tuple("" if cell is None else cell for cell in row1)
            row2 = tuple("" if cell is None else cell for cell in row2)

            comparison_results += [row1 == row2]
    return comparison_results


def compare_data_validators(expected: Workbook, resulted: Workbook) -> list[bool]:
    comparison_results = []

    for sheet_name in expected.sheetnames:
        comparison_results += [expected[sheet_name].data_validations == resulted[sheet_name].data_validations]
    return comparison_results


class TestExcelExporter:
    def test_export_dms_rules(self, alice_rules: PhysicalDataModel):
        exporter = ExcelExporter(styling="maximal")
        workbook = exporter.export(alice_rules)
        assert "Metadata" in workbook.sheetnames
        assert "Containers" in workbook.sheetnames
        assert "Views" in workbook.sheetnames
        assert "Properties" in workbook.sheetnames

    def test_export_information_rules(self, david_rules: ConceptualDataModel):
        exporter = ExcelExporter()
        workbook = exporter.export(david_rules)

        assert "Metadata" in workbook.sheetnames
        assert "Concepts" in workbook.sheetnames
        assert "Properties" in workbook.sheetnames

    def test_conceptual_data_model_template(self, tmp_path: Path):
        exporter = ExcelExporter(total_concepts=10, base_model="CogniteCore")

        expected = exporter.template(RoleTypes.information)

        template_path = tmp_path / "template.xlsx"
        exporter.template(RoleTypes.information, template_path)

        resulted = load_workbook(template_path)

        assert expected.sheetnames == resulted.sheetnames
        assert all(compare_cells(expected, resulted))
        assert all(compare_data_validators(expected, resulted))

        # There should be two data validators in the Properties sheet
        # one for Class column another one for Value Type column

        assert expected["Properties"].data_validations.count == 2
        assert resulted["Properties"].data_validations.count == 2

        # Validation should be applied to 10000 rows for Class and Value Type columns
        assert '<dataValidation sqref="E3:E10003"' in etree.tostring(
            resulted["Properties"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        assert '<dataValidation sqref="A3:A10003"' in etree.tostring(
            resulted["Properties"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        assert f"<formula1>={exporter._helper_sheet_name}!A$1:A$101</formula1>" in etree.tostring(
            resulted["Properties"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        assert f"<formula1>={exporter._helper_sheet_name}!C$1:C$151</formula1>" in etree.tostring(
            resulted["Properties"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        assert expected["Concepts"].data_validations.count == 1
        assert resulted["Concepts"].data_validations.count == 1

        assert f"<formula1>={exporter._helper_sheet_name}!B$1:B$201</formula1>" in etree.tostring(
            resulted["Concepts"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        assert 'dataValidation sqref="D3:D103"' in etree.tostring(
            resulted["Concepts"].data_validations.to_tree(), pretty_print=True
        ).decode("utf-8")

        # check if only 10 Cognite concepts are present in the dropdown
        # by checking if 11th row contains a formula
        assert expected["_helper"]["B11"].value == '=IF(ISBLANK(Concepts!A3), "", Concepts!A3)'
        assert resulted["_helper"]["B11"].value == '=IF(ISBLANK(Concepts!A3), "", Concepts!A3)'

        # Get all values from the "Value Type" column in the _helper sheet
        # Check if "json" is in the values from the "Value Type" column in the _helper sheet
        value_type_col_idx = exporter._helper_sheet_column_indexes_by_names["Value Type"]
        has_json = any(row[value_type_col_idx - 1].value == "json" for row in resulted["_helper"].iter_rows())
        assert has_json, "Expected 'json' to be found in _helper sheet column C"

    def test_conceptual_data_model_template_fail(self):
        exporter = ExcelExporter(base_model="NeatModel")

        with pytest.raises(NeatValueError) as e:
            _ = exporter.template(RoleTypes.information)

        assert e.value == NeatValueError(raw_message="Base model <NeatModel> is not supported")
