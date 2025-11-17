from collections.abc import Iterable
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cognite.neat._data_model.exporters._table_exporter.exporter import (
    DMSExcelExporter,
    DMSTableYamlExporter,
)
from cognite.neat._data_model.exporters._table_exporter.workbook import WorkbookCreator
from cognite.neat._data_model.importers import DMSTableImporter
from cognite.neat._data_model.importers._table_importer.source import SpreadsheetReadContext, TableSource
from cognite.neat._data_model.models.dms import (
    BtreeIndex,
    ContainerPropertyDefinition,
    ContainerReference,
    ContainerRequest,
    DataModelRequest,
    DirectNodeRelation,
    EnumProperty,
    EnumValue,
    Float32Property,
    MultiEdgeProperty,
    MultiReverseDirectRelationPropertyRequest,
    NodeReference,
    RequestSchema,
    RequiresConstraintDefinition,
    SpaceRequest,
    TextProperty,
    UniquenessConstraintDefinition,
    ViewCorePropertyRequest,
    ViewDirectReference,
    ViewReference,
    ViewRequest,
)
from cognite.neat._exceptions import DataModelImportException
from cognite.neat._utils.useful_types import CellValueType, DataModelTableType

SOURCE = "pytest.xlsx"


def valid_dms_table_formats() -> Iterable[tuple]:
    yield pytest.param(
        {
            "Metadata": [
                {
                    "Key": "space",
                    "Value": "cdf_cdm",
                },
                {
                    "Key": "externalId",
                    "Value": "CogniteCore",
                },
                {
                    "Key": "version",
                    "Value": "v1",
                },
                {
                    "Key": "name",
                    "Value": "Cognite Core Data Model",
                },
                {
                    "Key": "description",
                    "Value": "The Cognite Core Data Model (CDM) is a standardized data model for industrial data.",
                },
            ],
            "Properties": [
                {
                    "View": "CogniteDescribable",
                    "View Property": "name",
                    "Connection": None,
                    "Value Type": "text(maxTextSize=400)",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": False,
                    "Container": "CogniteDescribable",
                    "Container Property": "name",
                    "Index": "btree:name(cursorable=False)",
                    "Constraint": "uniqueness:uniqueName(bySpace=True)",
                },
                {
                    "View": "CogniteAsset",
                    "View Property": "files",
                    "Connection": "reverse(property=assets)",
                    "Value Type": "CogniteFile",
                    "Min Count": 0,
                    "Max Count": None,
                },
                {
                    "View": "CogniteAsset",
                    "View Property": "name",
                    "Connection": None,
                    "Value Type": "text(maxTextSize=400)",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": False,
                    "Container": "CogniteDescribable",
                    "Container Property": "name",
                    "Index": "btree:name(cursorable=False)",
                    "Constraint": "uniqueness:uniqueName(bySpace=True)",
                },
                {
                    "View": "CogniteFile",
                    "View Property": "assets",
                    "Connection": "direct",
                    "Value Type": "CogniteAsset",
                    "Min Count": 0,
                    "Max Count": 1200,
                    "Immutable": False,
                    "Container": "CogniteFile",
                    "Container Property": "assets",
                },
                {
                    "View": "CogniteFile",
                    "View Property": "equipments",
                    "Connection": "direct",
                    "Value Type": "#N/A",
                    "Min Count": 0,
                    "Max Count": 1200,
                    "Immutable": False,
                    "Container": "CogniteFile",
                    "Container Property": "equipments",
                },
                {
                    "View": "CogniteFile",
                    "View Property": "assetAnnotations",
                    "Connection": "edge(edgeSource=FileAnnotation,type=diagramAnnotation)",
                    "Value Type": "CogniteAsset",
                    "Min Count": 0,
                    "Max Count": None,
                },
                {
                    "View": "CogniteFile",
                    "View Property": "category",
                    "Connection": None,
                    "Value Type": "enum(collection=CogniteFile.category,unknownValue=other)",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": False,
                    "Container": "CogniteFile",
                    "Container Property": "category",
                    "Container Property Name": "category_405",
                },
                {
                    "View": "FileAnnotation",
                    "View Property": "confidence",
                    "Connection": None,
                    "Value Type": "float32",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": True,
                    "Container": "FileAnnotation",
                    "Container Property": "confidence",
                },
            ],
            "Views": [
                {
                    "View": "CogniteDescribable",
                    "Name": "Cognite Describable",
                    "Description": "The describable core concept is used as a standard way of "
                    "holding the bare minimum of information about the instance",
                },
                {
                    "View": "CogniteAsset",
                    "Name": "Cognite Asset",
                    "Implements": "CogniteDescribable",
                },
                {
                    "View": "CogniteFile",
                    "Name": "Cognite File",
                    "Implements": "CogniteDescribable",
                },
                {
                    "View": "FileAnnotation",
                    "Name": "File Annotation",
                    "Implements": "CogniteDescribable",
                },
            ],
            "Containers": [
                {
                    "Container": "CogniteDescribable",
                    "Used For": "all",
                },
                {
                    "Container": "CogniteFile",
                    "Constraint": "requires:describablePresent(require=CogniteDescribable)",
                    "Used For": "node",
                },
                {
                    "Container": "FileAnnotation",
                    "Constraint": "requires:describablePresent(require=CogniteDescribable)",
                    "Used For": "edge",
                },
            ],
            "Enum": [
                {
                    "Collection": "CogniteFile.category",
                    "Value": "blueprint",
                    "Name": "Blueprint",
                    "Description": "A technical drawing",
                },
                {
                    "Collection": "CogniteFile.category",
                    "Value": "document",
                },
                {
                    "Collection": "CogniteFile.category",
                    "Value": "other",
                },
            ],
            "Nodes": [
                {
                    "Node": "diagramAnnotation",
                }
            ],
        },
        RequestSchema(
            dataModel=DataModelRequest(
                space="cdf_cdm",
                externalId="CogniteCore",
                version="v1",
                name="Cognite Core Data Model",
                description="The Cognite Core Data Model (CDM) is a standardized data model for industrial data.",
                views=[
                    ViewReference(space="cdf_cdm", external_id="CogniteDescribable", version="v1"),
                    ViewReference(space="cdf_cdm", external_id="CogniteAsset", version="v1"),
                    ViewReference(space="cdf_cdm", external_id="CogniteFile", version="v1"),
                    ViewReference(space="cdf_cdm", external_id="FileAnnotation", version="v1"),
                ],
            ),
            spaces=[SpaceRequest(space="cdf_cdm")],
            views=[
                ViewRequest(
                    space="cdf_cdm",
                    externalId="CogniteDescribable",
                    version="v1",
                    name="Cognite Describable",
                    description="The describable core concept is used as a standard way of holding the bare minimum "
                    "of information about the instance",
                    implements=None,
                    properties={
                        "name": ViewCorePropertyRequest(
                            name=None,
                            description=None,
                            container=ContainerReference(space="cdf_cdm", external_id="CogniteDescribable"),
                            containerPropertyIdentifier="name",
                        ),
                    },
                ),
                ViewRequest(
                    space="cdf_cdm",
                    externalId="CogniteAsset",
                    version="v1",
                    name="Cognite Asset",
                    description=None,
                    implements=[ViewReference(space="cdf_cdm", external_id="CogniteDescribable", version="v1")],
                    properties={
                        "files": MultiReverseDirectRelationPropertyRequest(
                            name=None,
                            description=None,
                            source=ViewReference(space="cdf_cdm", external_id="CogniteFile", version="v1"),
                            through=ViewDirectReference(
                                source=ViewReference(space="cdf_cdm", external_id="CogniteFile", version="v1"),
                                identifier="assets",
                            ),
                        ),
                        "name": ViewCorePropertyRequest(
                            container=ContainerReference(space="cdf_cdm", external_id="CogniteDescribable"),
                            containerPropertyIdentifier="name",
                        ),
                    },
                ),
                ViewRequest(
                    space="cdf_cdm",
                    externalId="CogniteFile",
                    version="v1",
                    name="Cognite File",
                    description=None,
                    implements=[ViewReference(space="cdf_cdm", external_id="CogniteDescribable", version="v1")],
                    properties={
                        "assets": ViewCorePropertyRequest(
                            name=None,
                            description=None,
                            container=ContainerReference(space="cdf_cdm", external_id="CogniteFile"),
                            containerPropertyIdentifier="assets",
                            source=ViewReference(space="cdf_cdm", external_id="CogniteAsset", version="v1"),
                        ),
                        "equipments": ViewCorePropertyRequest(
                            name=None,
                            description=None,
                            container=ContainerReference(space="cdf_cdm", external_id="CogniteFile"),
                            containerPropertyIdentifier="equipments",
                            source=None,
                        ),
                        "assetAnnotations": MultiEdgeProperty(
                            name=None,
                            description=None,
                            source=ViewReference(space="cdf_cdm", external_id="CogniteAsset", version="v1"),
                            edgeSource=ViewReference(space="cdf_cdm", external_id="FileAnnotation", version="v1"),
                            direction="outwards",
                            type=NodeReference(space="cdf_cdm", external_id="diagramAnnotation"),
                        ),
                        "category": ViewCorePropertyRequest(
                            name=None,
                            description=None,
                            container=ContainerReference(space="cdf_cdm", external_id="CogniteFile"),
                            containerPropertyIdentifier="category",
                        ),
                    },
                ),
                ViewRequest(
                    space="cdf_cdm",
                    externalId="FileAnnotation",
                    version="v1",
                    name="File Annotation",
                    description=None,
                    implements=[ViewReference(space="cdf_cdm", external_id="CogniteDescribable", version="v1")],
                    properties={
                        "confidence": ViewCorePropertyRequest(
                            name=None,
                            description=None,
                            container=ContainerReference(space="cdf_cdm", external_id="FileAnnotation"),
                            containerPropertyIdentifier="confidence",
                        ),
                    },
                ),
            ],
            containers=[
                ContainerRequest(
                    space="cdf_cdm",
                    externalId="CogniteDescribable",
                    usedFor="all",
                    properties={
                        "name": ContainerPropertyDefinition(
                            immutable=False,
                            nullable=True,
                            autoIncrement=None,
                            defaultValue=None,
                            description=None,
                            name=None,
                            type=TextProperty(list=False, maxTextSize=400),
                        )
                    },
                    indexes={
                        "name": BtreeIndex(properties=["name"], cursorable=False),
                    },
                    constraints={
                        "uniqueName": UniquenessConstraintDefinition(
                            constraint_type="uniqueness",
                            properties=["name"],
                            bySpace=True,
                        )
                    },
                ),
                ContainerRequest(
                    space="cdf_cdm",
                    externalId="CogniteFile",
                    usedFor="node",
                    properties={
                        "assets": ContainerPropertyDefinition(
                            immutable=False,
                            nullable=True,
                            autoIncrement=None,
                            defaultValue=None,
                            description=None,
                            name=None,
                            type=DirectNodeRelation(maxListSize=1200, list=True),
                        ),
                        "equipments": ContainerPropertyDefinition(
                            immutable=False,
                            nullable=True,
                            autoIncrement=None,
                            defaultValue=None,
                            description=None,
                            name=None,
                            type=DirectNodeRelation(maxListSize=1200, list=True),
                        ),
                        "category": ContainerPropertyDefinition(
                            immutable=False,
                            nullable=True,
                            autoIncrement=None,
                            defaultValue=None,
                            description=None,
                            name="category_405",
                            type=EnumProperty(
                                unknownValue="other",
                                values={
                                    "blueprint": EnumValue(name="Blueprint", description="A technical drawing"),
                                    "document": EnumValue(),
                                    "other": EnumValue(),
                                },
                            ),
                        ),
                    },
                    constraints={
                        "describablePresent": RequiresConstraintDefinition(
                            require=ContainerReference(space="cdf_cdm", external_id="CogniteDescribable")
                        )
                    },
                ),
                ContainerRequest(
                    space="cdf_cdm",
                    externalId="FileAnnotation",
                    usedFor="edge",
                    properties={
                        "confidence": ContainerPropertyDefinition(
                            immutable=True,
                            nullable=True,
                            autoIncrement=None,
                            defaultValue=None,
                            description=None,
                            name=None,
                            type=Float32Property(list=False),
                        )
                    },
                    constraints={
                        "describablePresent": RequiresConstraintDefinition(
                            require=ContainerReference(space="cdf_cdm", external_id="CogniteDescribable")
                        )
                    },
                ),
            ],
            nodeTypes=[NodeReference(space="cdf_cdm", external_id="diagramAnnotation")],
        ),
        id="Full example",
    )
    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "multiIndex",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiIndex",
                    "Index": "btree:compositeIdx(order=1)",
                },
                {
                    "View": "TestView",
                    "View Property": "multiIndex2",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiIndex2",
                    "Index": "btree:compositeIdx(order=2)",
                },
                {
                    "View": "TestView",
                    "View Property": "multiConstraint",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiConstraint",
                    "Constraint": "uniqueness:compositeUnique(order=1)",
                },
                {
                    "View": "TestView",
                    "View Property": "multiConstraint2",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiConstraint2",
                    "Constraint": "uniqueness:compositeUnique(order=2)",
                },
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [{"Container": "TestContainer", "Used For": "node"}],
        },
        RequestSchema(
            dataModel=DataModelRequest(
                space="test_space",
                externalId="TestModel",
                version="v1",
                views=[ViewReference(space="test_space", external_id="TestView", version="v1")],
            ),
            spaces=[SpaceRequest(space="test_space")],
            views=[
                ViewRequest(
                    space="test_space",
                    externalId="TestView",
                    version="v1",
                    properties={
                        "multiIndex": ViewCorePropertyRequest(
                            container=ContainerReference(space="test_space", external_id="TestContainer"),
                            containerPropertyIdentifier="multiIndex",
                        ),
                        "multiIndex2": ViewCorePropertyRequest(
                            container=ContainerReference(space="test_space", external_id="TestContainer"),
                            containerPropertyIdentifier="multiIndex2",
                        ),
                        "multiConstraint": ViewCorePropertyRequest(
                            container=ContainerReference(space="test_space", external_id="TestContainer"),
                            containerPropertyIdentifier="multiConstraint",
                        ),
                        "multiConstraint2": ViewCorePropertyRequest(
                            container=ContainerReference(space="test_space", external_id="TestContainer"),
                            containerPropertyIdentifier="multiConstraint2",
                        ),
                    },
                )
            ],
            containers=[
                ContainerRequest(
                    space="test_space",
                    externalId="TestContainer",
                    usedFor="node",
                    properties={
                        "multiIndex": ContainerPropertyDefinition(
                            type=TextProperty(list=False),
                            nullable=True,
                        ),
                        "multiIndex2": ContainerPropertyDefinition(
                            type=TextProperty(list=False),
                            nullable=True,
                        ),
                        "multiConstraint": ContainerPropertyDefinition(
                            type=TextProperty(list=False),
                            nullable=True,
                        ),
                        "multiConstraint2": ContainerPropertyDefinition(
                            type=TextProperty(list=False),
                            nullable=True,
                        ),
                    },
                    indexes={
                        "compositeIdx": BtreeIndex(properties=["multiIndex", "multiIndex2"], cursorable=None),
                    },
                    constraints={
                        "compositeUnique": UniquenessConstraintDefinition(
                            constraint_type="uniqueness",
                            properties=["multiConstraint", "multiConstraint2"],
                        )
                    },
                )
            ],
            nodeTypes=[],
        ),
        id="Multi-property indices and constraints",
    )


def invalid_dms_table_formats() -> Iterable[tuple]:
    yield pytest.param(
        {
            "Metadata": [
                {
                    "Key": "space",
                    "Value": "cdf_cdm",
                },
                {
                    "Key": "version",
                    "Value": "v1",
                },
            ],
            "Properties": [
                {
                    "View": "CogniteDescribable",
                    "View Property": "name",
                    "Name": None,
                    "Description": None,
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": False,
                    "Default": None,
                    "Container": "CogniteDescribable",
                    "Container Property": "name",
                    "Index": "btree:name(cursorable=invalid)",
                    "Constraint": None,
                }
            ],
            "Views": [
                {
                    "View": None,
                    "Name": "Cognite Describable",
                    "Description": "The describable core concept is used as a standard way of "
                    "holding the bare minimum of information about the instance",
                    "Implements": None,
                    "Filter": None,
                }
            ],
            "Containers": [
                {
                    "Container": "CogniteDescribable",
                    "Name": None,
                    "Description": None,
                    "Constraint": None,
                    "Used For": "Instances",
                }
            ],
        },
        {
            "In table 'Metadata' missing required value: 'externalId'.",
            "In table 'Properties' row 1 column 'Index' -> btree.cursorable input should be "
            "a valid boolean. Got 'invalid' of type str.",
            "In table 'Views' row 1 the column 'View' cannot be empty.",
            "In table 'Containers' row 1 column 'Used For' input should be 'node', 'edge' or 'all'. Got 'Instances'.",
        },
        id="Missing required metadata fields",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "prop1",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "sameProp",
                },
                {
                    "View": "TestView",
                    "View Property": "prop2",
                    "Connection": None,
                    "Value Type": "int32",
                    "Min Count": 1,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "sameProp",
                },
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [{"Container": "TestContainer", "Used For": "node"}],
        },
        {
            "In table 'Properties' when the column 'Container' and 'Container Property' "
            "are the same, all the container columns (Auto Increment, Connection, "
            "Constraint, Container Property Description, Container Property Name, "
            "Default, Index, Max Count, Min Count and Value Type) must be the same. "
            "Inconsistent definitions for container 'TestContainer and 'sameProp'' found "
            "in rows 1 and 2."
        },
        id="Inconsistent container property definitions",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "duplicate",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "duplicate",
                },
                {
                    "View": "TestView",
                    "View Property": "duplicate",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "duplicate",
                },
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [],
        },
        {
            "In table 'Properties' the combination of columns 'View' and 'View Property' "
            "must be unique. Duplicated entries for view 'TestView' and property "
            "'duplicate' found in rows 1 and 2."
        },
        id="Duplicate view properties",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "prop1",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "prop1",
                }
            ],
            "Views": [
                {"View": "TestView"},
                {"View": "TestView"},
            ],
            "Containers": [
                {"Container": "TestContainer", "Used For": "node"},
                {"Container": "TestContainer", "Used For": "edge"},
            ],
        },
        {
            "In table 'Containers' the values in column 'Container' must be unique. "
            "Duplicated entries for container 'TestContainer' found in rows 1 and 2.",
            "In table 'Views' the values in column 'View' must be unique. Duplicated "
            "entries for view 'TestView' found in rows 1 and 2.",
        },
        id="Duplicate views and containers",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "multiIdx",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiIdx",
                    "Index": "btree:compositeIdx",
                },
                {
                    "View": "TestView",
                    "View Property": "multiIdx2",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiIdx2",
                    "Index": "btree:compositeIdx",
                },
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [{"Container": "TestContainer", "Used For": "node"}],
        },
        {
            "In table 'Properties' column 'Index': the index 'compositeIdx' on container TestContainer "
            "is defined on multiple properties. This requires the 'order' attribute to be set. "
            "It is missing in rows 1 and 2.",
        },
        id="Multi-property index missing order",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "multiConst",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiConst",
                    "Constraint": "uniqueness:compositeConst",
                },
                {
                    "View": "TestView",
                    "View Property": "multiConst2",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "multiConst2",
                    "Constraint": "uniqueness:compositeConst",
                },
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [{"Container": "TestContainer", "Used For": "node", "Constraint": "requires:TestConst"}],
        },
        {
            "In table 'Containers' row 1 column 'Constraint' the constraint 'TestConst' "
            "on container 'TestContainer' is missing the 'require' property which is "
            "required for container level constraints.",
            "In table 'Properties' column 'Constraint': the uniqueness constraint "
            "'compositeConst' on container TestContainer is defined on multiple "
            "properties. This requires the 'order' attribute to be set. It is missing in "
            "rows 1 and 2.",
        },
        id="Multi-property constraint missing order",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [],
            "Views": [
                {
                    "View": "TestView",
                    "Filter": "invalid json{",
                }
            ],
            "Containers": [],
        },
        {
            "In table 'Views' row 1 column 'Filter' must be valid json. Got error "
            "Expecting value: line 1 column 1 (char 0)"
        },
        id="Invalid JSON filter",
    )

    yield pytest.param(
        {
            "Metadata": [
                {"Key": "space", "Value": "test_space"},
                {"Key": "externalId", "Value": "TestModel"},
                {"Key": "version", "Value": "v1"},
            ],
            "Properties": [
                {
                    "View": "TestView",
                    "View Property": "badOrder",
                    "Connection": None,
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Container": "TestContainer",
                    "Container Property": "badOrder",
                    "Index": "btree:testIdx(order=not_an_int)",
                    "Constraint": "uniqueness:testConst",
                }
            ],
            "Views": [{"View": "TestView"}],
            "Containers": [
                {
                    "Container": "TestContainer",
                    "Used For": "node",
                    "Constraint": "requires:testConst(require=cdf_cdm:CogniteDescribable)",
                }
            ],
        },
        {
            "In table 'Containers' row 1 column 'Constraint' the container "
            "'TestContainer' has constraints defined with the same identifier(s) as the "
            "uniqueness constraint defined in the Properties sheet. Ensure that the "
            "identifiers are unique. Conflicting identifiers: testConst. ",
            "In table 'Properties' row 1 column 'Index' invalid order value 'not_an_int'. Must be an integer.",
        },
        id="Invalid order value in index",
    )


def invalid_dms_table() -> Iterable[tuple]:
    """This tests cases are designed to fail the initial validation in the DMSTableImporter"""
    yield pytest.param(
        {
            "Metadata": [
                {
                    "Key": "space",
                    "Value": "my_space",
                }
            ],
            "properties": [
                {
                    "View": "MyView",
                    "View Property": "prop1",
                    "Value Type": "text",
                    "Min Count": 1,
                    "Max Count": 1,
                },
                {
                    "View": "asset:MyAsset(capacity=100,type=storage)trailing",
                    "View Property": "prop2",
                    "Value Type": "int32",
                    "Min Count": "not_an_int",
                    "Max Count": 1,
                },
            ],
        },
        {
            "In Properties sheet missing required column: 'Connection'.",
            "In Properties sheet row 2 column 'Min Count' input should be a valid "
            "integer, unable to parse string as an integer.",
            "In Properties sheet row 2 column 'View' invalid entity syntax: Unexpected "
            "characters after properties at position 40. Got 't'.",
            "Missing required column: 'Views'.",
        },
        id="Missing required column in Properties table",
    )

    yield pytest.param(
        {
            "Metadata": [
                {
                    "Key": "space",
                    "Value": "my_space",
                }
            ],
            "Properties": [
                {
                    "View": "MyView",
                    "View Property": "prop1",
                    "Value Type": "text",
                    "Connection": "MyConnection",
                    "Immutable": "not_a_boolean",
                    "Auto Increment": "maybe",
                }
            ],
            "Views": [
                {
                    "View": "MyView",
                    "Implements": "invalid[entity,list]syntax",
                }
            ],
        },
        {
            "In Properties sheet missing required column: 'Max Count'.",
            "In Properties sheet missing required column: 'Min Count'.",
            "In Properties sheet row 1 column 'Auto Increment' input should be a valid "
            "boolean, unable to interpret input.",
            "In Properties sheet row 1 column 'Immutable' input should be a valid boolean, unable to interpret input.",
        },
        id="Invalid boolean and entity list values",
    )

    yield pytest.param(
        {
            "Metadata": [
                {
                    "Value": "my_space",
                }
            ],
            "Properties": [
                {
                    "View": "MyView",
                    "Value Type": "text",  # Missing required "View Property" field
                    "Connection": "MyConnection",
                }
            ],
            "Views": [
                {
                    "Name": "Some View Name",
                }
            ],
        },
        {
            "In Metadata sheet missing required column: 'Key'.",
            "In Properties sheet missing required column: 'Max Count'.",
            "In Properties sheet missing required column: 'Min Count'.",
            "In Properties sheet missing required column: 'View Property'.",
            "In Views sheet missing required column: 'View'.",
        },
        id="Missing required fields in various sheets",
    )


@pytest.fixture
def max_count_infinity_table() -> dict:
    return {
        "Metadata": [
            {"Key": "space", "Value": "test_space"},
            {"Key": "externalId", "Value": "InfinityTest"},
            {"Key": "version", "Value": "v1"},
        ],
        "Properties": [
            {
                "View": "TestView",
                "View Property": "unlimitedList",
                "Connection": "reverse(property=limitedList)",
                "Value Type": "TestView",
                "Min Count": 0,
                "Max Count": "inf",  # Test infinity
            },
            {
                "View": "TestView",
                "View Property": "limitedList",
                "Connection": "direct",
                "Value Type": "TestTarget",
                "Min Count": 0,
                "Max Count": 100,  # Regular max count for comparison
                "Immutable": False,
                "Container": "TestContainer",
                "Container Property": "limitedList",
            },
        ],
        "Views": [
            {"View": "TestView"},
            {"View": "TestTarget"},
        ],
        "Containers": [
            {"Container": "TestContainer", "Used For": "node"},
        ],
    }


class TestDMSTableImporter:
    @pytest.mark.parametrize("data, expected_errors", list(invalid_dms_table()))
    def test_read_invalid_tables(
        self, data: dict[str, list[dict[str, CellValueType]]], expected_errors: set[str]
    ) -> None:
        importer = DMSTableImporter(data)
        with pytest.raises(DataModelImportException) as exc_info:
            importer._read_tables()
        actual_errors = {err.message for err in exc_info.value.errors}
        assert actual_errors == expected_errors

    @pytest.mark.parametrize("data,expected", list(valid_dms_table_formats()))
    def test_import(self, data: dict[str, list[dict[str, CellValueType]]], expected: RequestSchema) -> None:
        importer = DMSTableImporter(data)
        result = importer.to_data_model()
        assert result.model_dump() == expected.model_dump()

    @pytest.mark.parametrize("data,expected_errors", list(invalid_dms_table_formats()))
    def test_import_errors(self, data: dict[str, list[dict[str, CellValueType]]], expected_errors: set[str]) -> None:
        importer = DMSTableImporter(data, source=TableSource(source=SOURCE, table_read={}))
        with pytest.raises(DataModelImportException) as e:
            _ = importer.to_data_model()

        result_errors = {err.message for err in e.value.errors}

        assert result_errors == expected_errors

    def test_legacy_max_count_infinity(
        self, max_count_infinity_table: dict[str, list[dict[str, CellValueType]]]
    ) -> None:
        importer = DMSTableImporter(max_count_infinity_table)
        schema = importer.to_data_model()

        result = DMSTableYamlExporter().export(schema)
        assert result["Properties"][0]["Max Count"] is None  # infinity represented as None
        assert result["Properties"][1]["Max Count"] == 100


class TestDMSTableExporter:
    @pytest.mark.parametrize("expected,schema", list(valid_dms_table_formats()))
    def test_export(self, expected: dict[str, list[dict[str, CellValueType]]], schema: RequestSchema) -> None:
        result = DMSTableYamlExporter().export(schema)

        assert result == expected


class TestTableSource:
    @pytest.mark.parametrize(
        "path,table_read,expected",
        [
            pytest.param((), {}, "", id="empty_path"),
            pytest.param(("MyTable",), {}, "table 'MyTable'", id="table_only"),
            pytest.param(("MyTable", 5), {}, "table 'MyTable' row 6", id="table_and_row"),
            pytest.param(("MyTable", 5, "field"), {}, "table 'MyTable' row 6 column 'field'", id="table_row_column"),
            pytest.param(
                ("MyTable", 5),
                {"MyTable": SpreadsheetReadContext(header_row=2, empty_rows=[3, 5])},
                "table 'MyTable' row 11",
                id="with_spreadsheet_read",
            ),
            pytest.param(("Views", 1, "externalId"), {}, "table 'Views' row 2 column 'View'", id="with_field_mapping"),
            pytest.param(
                ("MyTable", 1, "field", "nested", "path"),
                {},
                "table 'MyTable' row 2 column 'field' -> nested.path",
                id="with_extra_path_elements",
            ),
            pytest.param((123, 5, "field"), {}, "row 6 column 'field'", id="non_string_table_id"),
            pytest.param(
                ("MyTable", "not_int", "field"), {}, "table 'MyTable' column 'field'", id="non_int_row_number"
            ),
            pytest.param((0, 5), {}, "row 6", id="row_only"),
            pytest.param((0, "not_int", "field"), {}, "column 'field'", id="column_only"),
            pytest.param(
                ("MyTable", 1, "field", "extra"), {}, "table 'MyTable' row 2 column 'field'", id="path_length_exactly_4"
            ),
            pytest.param(
                ("MyTable", 1, "field", "a", "b", "c"),
                {},
                "table 'MyTable' row 2 column 'field' -> a.b.c",
                id="path_length_greater_than_4",
            ),
        ],
    )
    def test_location(
        self, path: tuple[int | str, ...], table_read: dict[str, SpreadsheetReadContext], expected: str
    ) -> None:
        source = TableSource("test_source", table_read)
        assert source.location(path) == expected

    @pytest.mark.parametrize(
        "table_id,row_no,table_read,expected",
        [
            pytest.param(
                "MyTable",
                5,
                {"MyTable": SpreadsheetReadContext(header_row=2, empty_rows=[1, 3])},
                10,  # 5 + 2 (header) + 2 (empty rows before 5) + 1 (1-indexed)
                id="with_table_read",
            ),
            pytest.param("MyTable", 5, {}, 6, id="without_table_read"),
            pytest.param(None, 5, {}, 6, id="none_table_id"),
            pytest.param("", 3, {"": SpreadsheetReadContext(header_row=5)}, 4, id="falsy_table_id"),
        ],
    )
    def test_adjust_row_number(
        self, table_id: str | None, row_no: int, table_read: dict[str, SpreadsheetReadContext], expected: int
    ) -> None:
        source = TableSource("test_source", table_read)
        assert source.adjust_row_number(table_id, row_no) == expected

    @pytest.mark.parametrize(
        "table_id,field,expected",
        [
            pytest.param("Views", "externalId", "View", id="views_table_external_id"),
            pytest.param("Views", "space", "View", id="views_table_space"),
            pytest.param("UnknownTable", "someField", "someField", id="unknown_table"),
            pytest.param(None, "someField", "someField", id="none_table_id"),
            pytest.param("Views", "unmappedField", "unmappedField", id="unmapped_field_in_mapped_table"),
        ],
    )
    def test_field_to_column(self, table_id: str | None, field: str, expected: str) -> None:
        assert TableSource.field_to_column(table_id, field) == expected

    @pytest.mark.parametrize(
        "table_id,expected_has_mapping,expected_external_id",
        [
            pytest.param("Views", True, "View", id="string_table_id_views"),
            pytest.param(123, False, None, id="non_string_table_id"),
            pytest.param(None, False, None, id="none_table_id"),
            pytest.param("UnknownTable", False, None, id="unknown_table"),
        ],
    )
    def test_field_mapping(
        self, table_id: str | int | None, expected_has_mapping: bool, expected_external_id: str | None
    ) -> None:
        mapping = TableSource.field_mapping(table_id)
        if expected_has_mapping:
            assert mapping is not None
            assert "externalId" in mapping
            assert mapping["externalId"] == expected_external_id
        else:
            assert mapping is None


class TestSpreadsheetRead:
    @pytest.mark.parametrize(
        "row,read,expected",
        [
            pytest.param(
                5,
                SpreadsheetReadContext(header_row=1, empty_rows=[]),
                8,  # 5 + 1 (header) + 1 (one_indexed) + 1 (offset for rows after header)
                id="basic_case_with_one_indexing",
            ),
            pytest.param(
                5,
                SpreadsheetReadContext(header_row=2, empty_rows=[1, 3, 6]),
                11,  # 5->6 (empty 1)->7 (empty 3)->8 (empty 6) + 2 (header) + 1 (one_indexed)
                id="with_empty_rows",
            ),
            pytest.param(
                3,
                SpreadsheetReadContext(header_row=1, empty_rows=[2]),
                7,  # 3->4 (empty 2)->5 (skipped 1) + 1 (header) + 0 (zero_indexed)
                id="with_empty_and_skipped_rows_zero_indexed",
            ),
            pytest.param(
                4,
                SpreadsheetReadContext(header_row=1, empty_rows=[6, 12]),
                8,
                id="real_case",
            ),
        ],
    )
    def test_adjusted_row_number(self, row: int, read: SpreadsheetReadContext, expected: int) -> None:
        assert read.adjusted_row_number(row) == expected


def valid_dms_yaml_formats() -> Iterable[tuple]:
    yield pytest.param(
        """Metadata:
- Key: space
  Value: cdf_cdm
- Key: externalId
  Value: CogniteDataModel
- Key: version
  Value: v1
Properties:
- View: CogniteDescribable
  View Property: name
  Value Type: text
  Min Count: 0
  Max Count: 1
  Immutable: false
  Container: CogniteDescribable
  Container Property: name
  Index: btree:name(cursorable=True)
  Connection: null
Views:
- View: CogniteDescribable
Containers:
- Container: CogniteDescribable
  Used For: node
""",
        id="Minimal example",
    )


class TestYAMLTableFormat:
    @pytest.mark.parametrize("yaml_str", list(valid_dms_yaml_formats()))
    def test_roundtrip(self, yaml_str: str) -> None:
        yaml_file = MagicMock(spec=Path)
        yaml_file.read_text.return_value = yaml_str
        data_model = DMSTableImporter.from_yaml(yaml_file).to_data_model()

        yaml_file.read_text.assert_called_once()
        result_file = MagicMock(spec=Path)
        DMSTableYamlExporter().export_to_file(data_model, result_file)

        result_file.write_text.assert_called_once()
        written_yaml = result_file.write_text.call_args[0][0]
        assert written_yaml == yaml_str


def valid_dms_excel_formats() -> Iterable[tuple]:
    yield pytest.param(
        {
            "Metadata": [
                ["space", "cdf_cdm"],
                ["externalId", "CogniteDataModel"],
                ["version", "v1"],
            ],
            "Properties": [
                ["Definition of Properties", *[None] * 16],
                [
                    "View",
                    "View Property",
                    "Name",
                    "Description",
                    "Connection",
                    "Value Type",
                    "Min Count",
                    "Max Count",
                    "Immutable",
                    "Default",
                    "Auto Increment",
                    "Container",
                    "Container Property",
                    "Container Property Name",
                    "Container Property Description",
                    "Index",
                    "Constraint",
                ],
                [
                    "CogniteDescribable",
                    "name",
                    None,
                    None,
                    None,
                    "text",
                    0,
                    1,
                    False,
                    None,
                    None,
                    "CogniteDescribable",
                    "name",
                    None,
                    None,
                    "btree:name(cursorable=True)",
                    None,
                ],
            ],
            "Views": [
                ["Definition of Views", *[None] * 4],
                [
                    "View",
                    "Name",
                    "Description",
                    "Implements",
                    "Filter",
                ],
                [
                    "CogniteDescribable",
                    "Cognite Describable",
                    "The describable core concept is used as a standard way of holding the "
                    "bare minimum of information about the instance",
                    None,
                    None,
                ],
            ],
            "Containers": [
                ["Definition of Containers", *[None] * 4],
                [
                    "Container",
                    "Name",
                    "Description",
                    "Constraint",
                    "Used For",
                ],
                [
                    "CogniteDescribable",
                    None,
                    None,
                    None,
                    "node",
                ],
            ],
        },
        dict(
            Metadata=[
                {"Key": "space", "Value": "cdf_cdm"},
                {"Key": "externalId", "Value": "CogniteDataModel"},
                {"Key": "version", "Value": "v1"},
            ],
            Properties=[
                {
                    "View": "CogniteDescribable",
                    "View Property": "name",
                    "Value Type": "text",
                    "Min Count": 0,
                    "Max Count": 1,
                    "Immutable": False,
                    "Default": None,
                    "Auto Increment": None,
                    "Container": "CogniteDescribable",
                    "Container Property": "name",
                    "Container Property Description": None,
                    "Container Property Name": None,
                    "Index": "btree:name(cursorable=True)",
                    "Connection": None,
                    "Name": None,
                    "Description": None,
                    "Constraint": None,
                }
            ],
            Views=[
                {
                    "View": "CogniteDescribable",
                    "Name": "Cognite Describable",
                    "Description": "The describable core concept is used as a standard way of holding the "
                    "bare minimum of information about the instance",
                    "Implements": None,
                    "Filter": None,
                }
            ],
            Containers=[
                {
                    "Container": "CogniteDescribable",
                    "Name": None,
                    "Description": None,
                    "Constraint": None,
                    "Used For": "node",
                }
            ],
        ),
        TableSource(
            source="excel_file.xlsx",
            table_read={
                "Metadata": SpreadsheetReadContext(),
                "Properties": SpreadsheetReadContext(header_row=1),
                "Views": SpreadsheetReadContext(header_row=1),
                "Containers": SpreadsheetReadContext(header_row=1),
            },
        ),
        id="Minimal example",
    )


class TestExcelFormat:
    @pytest.mark.parametrize("excel_tables,expected_tables,expected_context", list(valid_dms_excel_formats()))
    def test_read_excel(
        self,
        excel_tables: dict[str, list[list[CellValueType]]],
        expected_tables: DataModelTableType,
        expected_context: TableSource,
    ) -> None:
        excel_file = self._create_excel_file_mock(expected_context.source)
        load_workbook_mock = self._create_load_workbook_mock(excel_tables)
        with patch(f"{DMSTableImporter.__module__}.load_workbook", load_workbook_mock):
            importer = DMSTableImporter.from_excel(excel_file)

        assert importer._table == expected_tables
        assert importer._source == expected_context

    @pytest.mark.parametrize("excel_tables,expected_tables,expected_context", list(valid_dms_excel_formats()))
    def test_roundtrip_excel(
        self,
        excel_tables: dict[str, list[list[CellValueType]]],
        expected_tables: DataModelTableType,
        expected_context: TableSource,
    ) -> None:
        excel_file = self._create_excel_file_mock(expected_context.source)
        load_workbook_mock = self._create_load_workbook_mock(excel_tables)
        with patch(f"{DMSTableImporter.__module__}.load_workbook", load_workbook_mock):
            importer = DMSTableImporter.from_excel(excel_file)
            data_model = importer.to_data_model()

            exported = DMSExcelExporter().export(data_model)
            created_workbook = WorkbookCreator().create_workbook(exported)

            read_tables = self._read_workbook(created_workbook)
            read_tables.pop(WorkbookCreator.Sheets.dropdown_source)
            assert excel_tables == read_tables

    @staticmethod
    def _create_load_workbook_mock(excel_tables: dict[str, list[list[CellValueType]]]) -> MagicMock:
        workbook = MagicMock(spec=Workbook)

        workbook.sheetnames = list(excel_tables.keys())
        sheet_by_name: dict[str, MagicMock] = {}
        for sheet_name, rows in excel_tables.items():
            sheet = MagicMock(spec=Worksheet)
            sheet.title = sheet_name
            sheet.iter_rows.return_value = rows
            sheet_by_name[sheet_name] = sheet

        def get_item(name: str) -> MagicMock:
            return sheet_by_name[name]

        workbook.__getitem__.side_effect = get_item
        load_workbook_mock = MagicMock(return_value=workbook)
        return load_workbook_mock

    @staticmethod
    def _create_excel_file_mock(excel_file_name: str) -> MagicMock:
        excel_file = MagicMock(spec=Path)
        excel_file.relative_to.return_value = Path(excel_file_name)
        excel_file.name = excel_file_name
        return excel_file

    @staticmethod
    def _read_workbook(workbook: Workbook) -> dict[str, list[list[CellValueType]]]:
        output: dict[str, list[list[CellValueType]]] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows: list[list[CellValueType]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))  # type: ignore[arg-type]
            output[sheet_name] = rows
        return output
