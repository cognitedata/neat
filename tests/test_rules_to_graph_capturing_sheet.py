from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from cognite.neat.core.extractors import rules2graph_capturing_sheet


def test_graph_capturing_sheet(simple_rules, graph_capturing_sheet):
    tmp_sheet = NamedTemporaryFile(suffix=".xlsx")
    rules2graph_capturing_sheet(tmp_sheet.name, simple_rules, add_drop_down_list=True, use_uuid_id=True)
    resulting_sheet = load_workbook(tmp_sheet.name)
    tmp_sheet.close()

    assert resulting_sheet.sheetnames == graph_capturing_sheet.sheetnames

    for worksheet in resulting_sheet.sheetnames:
        source = resulting_sheet[worksheet]
        target = graph_capturing_sheet[worksheet]
        for row in range(1, source.max_row + 1):
            for col in range(1, source.max_column + 1):
                cell1 = source.cell(row, col)
                cell2 = target.cell(row, col)
                assert cell1.value == cell2.value
